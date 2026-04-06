import calendar
import csv
import io
import json
import logging
import os
import re
import requests
import tempfile
import threading
import time
import uuid
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from decimal import Decimal
from functools import wraps

from dotenv import load_dotenv
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, flash, make_response, Response
from flask.json.provider import DefaultJSONProvider
from flask_cors import CORS
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from mysql.connector import Error
from werkzeug.security import generate_password_hash, check_password_hash

from services.hnb_exchange_rate_service import get_hnb_exchange_rate_service
from services.pb_exchange_rate_service import get_pb_exchange_rate_service
from services.sampath_exchange_rate_service import get_sampath_exchange_rate_service
from services.gemini_bill_scanner import get_gemini_bill_scanner
from services.gemini_exchange_analyzer import get_gemini_exchange_analyzer
from services.backup_service import get_backup_service
from services.appwrite_file_service import get_appwrite_file_service
from services.exchange_rate_routes import register_exchange_rate_routes
from services.tax_service import register_tax_routes

# Load environment variables from .env file
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Try importing optional dependencies
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export will use CSV format.")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab not installed. PDF export will use text format.")

try:
    from appwrite.client import Client
    from appwrite.services.storage import Storage
    from appwrite.input_file import InputFile
    from appwrite.id import ID

    APPWRITE_AVAILABLE = True
except ImportError:
    APPWRITE_AVAILABLE = False
    logger.warning("appwrite not installed. Bill image upload will be disabled.")

try:
    from PIL import Image

    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    logger.warning("Pillow not installed. Image optimization will be disabled.")

# PyPDF2 removed - PDFs are converted to images on frontend before upload

logger.info("Environment variables loaded")

# Initialize Appwrite client (will be set up after helper functions are defined)
appwrite_client = None
appwrite_storage = None
APPWRITE_BUCKET_ID = None

# Initialize Appwrite file service
appwrite_file_service = get_appwrite_file_service()


# Custom JSON provider to handle Decimal objects
class DecimalJSONProvider(DefaultJSONProvider):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        return super().default(obj)


# Flask app configuration
app = Flask(__name__)
app.json = DecimalJSONProvider(app)

# Require SECRET_KEY from environment (no fallback for security)
if not os.environ.get('SECRET_KEY'):
    raise ValueError("SECRET_KEY environment variable is required. Please set it in your .env file.")
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=365)  # Remember me for 1 year

# Session cookie configuration for mobile browser compatibility
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # Allow cookies in same-site context
app.config['SESSION_COOKIE_HTTPONLY'] = True  # Prevent JavaScript access for security
# Enable secure cookies only when HTTPS is available (production)
IS_HTTPS = os.environ.get('HTTPS_ENABLED', 'False').lower() == 'true'
app.config['SESSION_COOKIE_SECURE'] = IS_HTTPS
app.config['SESSION_COOKIE_NAME'] = 'session'  # Explicit session cookie name
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB max upload size

# Configure CORS with restricted origins for security
allowed_origins = []
cors_origins_env = os.environ.get('CORS_ALLOWED_ORIGINS', '')

if cors_origins_env:
    # Use explicitly configured origins from environment variable
    allowed_origins = [origin.strip() for origin in cors_origins_env.split(',') if origin.strip()]
else:
    # Fallback: Allow localhost for development, or require explicit configuration for production
    if IS_HTTPS:
        logger.warning("HTTPS enabled but CORS_ALLOWED_ORIGINS not set. CORS will be restrictive.")
        allowed_origins = []  # No origins allowed - must be explicitly configured
    else:
        # Development mode - allow localhost
        allowed_origins = [
            'http://localhost:5000',
            'http://127.0.0.1:5000'
        ]

if allowed_origins:
    CORS(app, origins=allowed_origins, supports_credentials=True)
    logger.info(f"CORS configured with allowed origins: {allowed_origins}")
else:
    logger.warning("CORS not configured - no origins allowed. Set CORS_ALLOWED_ORIGINS in environment.")

# ---------------------------------------------------------------------------
# Rate Limiting Configuration
# ---------------------------------------------------------------------------
# Initialize Flask-Limiter for rate limiting (brute force protection)
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=[],  # No global limits, we'll set per-endpoint
    storage_uri="memory://",  # Use in-memory storage (switch to Redis for production scaling)
    strategy="fixed-window",  # Fixed window strategy
)

# Load rate limit thresholds from environment variables with fallback defaults
RATE_LIMIT_LOGIN = os.environ.get('RATE_LIMIT_LOGIN', '5 per 15 minutes')
RATE_LIMIT_REGISTER = os.environ.get('RATE_LIMIT_REGISTER', '3 per hour')
RATE_LIMIT_CHANGE_PASSWORD = os.environ.get('RATE_LIMIT_CHANGE_PASSWORD', '3 per hour')
RATE_LIMIT_ADMIN = os.environ.get('RATE_LIMIT_ADMIN', '30 per minute')
RATE_LIMIT_API = os.environ.get('RATE_LIMIT_API', '100 per minute')

logger.info(
    f"Rate limiting configured - Login: {RATE_LIMIT_LOGIN}, Register: {RATE_LIMIT_REGISTER}, Admin: {RATE_LIMIT_ADMIN}")


# ---------------------------------------------------------------------------
# Rate Limit Error Handler
# ---------------------------------------------------------------------------
@app.errorhandler(429)
def ratelimit_handler(e):
    """Handle rate limit exceeded errors."""
    logger.warning(f"Rate limit exceeded for IP: {get_remote_address()} - {request.path}")
    return jsonify({
        'error': 'Rate limit exceeded',
        'message': 'Too many requests. Please try again later.',
        'retry_after': e.description
    }), 429


# ---------------------------------------------------------------------------
# Cache Control - Disable caching during development
# ---------------------------------------------------------------------------
@app.after_request
def add_header(response):
    """
    Add headers to disable caching for HTML, CSS, and JS files.
    This ensures changes are immediately visible during development.

    Remove or modify this in production for better performance.
    """
    # Don't cache HTML, CSS, JS, or JSON responses
    if response.content_type and any(ct in response.content_type for ct in
                                     ['text/html', 'text/css', 'application/javascript',
                                      'application/json', 'text/javascript']):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


# ---------------------------------------------------------------------------
# Database connection pool (centralised in db.py)
# ---------------------------------------------------------------------------
from db import get_db_connection, DB_CONFIG  # noqa: E402

if DB_CONFIG is None:
    logger.warning(
        "Database is NOT configured. The application will start but all "
        "database operations will fail. Copy .env.example to .env and fill "
        "in your database credentials, then restart the application.")


# ---------------------------------------------------------------------------
# File Optimization Utilities
# ---------------------------------------------------------------------------

def fix_image_orientation(img):
    """
    Fix image orientation based on EXIF data.

    Mobile devices often store images in the wrong orientation and use EXIF
    tags to indicate how they should be rotated. This function reads the EXIF
    orientation tag and physically rotates/transposes the image accordingly.

    Args:
        img: PIL Image object

    Returns:
        PIL Image object with corrected orientation
    """
    try:
        # Get EXIF data
        exif = img.getexif()

        if exif is not None:
            # EXIF orientation tag is 0x0112 (274 in decimal)
            orientation = exif.get(0x0112, 1)

            # Apply rotation/transpose based on orientation tag
            # See: http://sylvana.net/jpegcrop/exif_orientation.html
            if orientation == 2:
                # Mirrored horizontally
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
            elif orientation == 3:
                # Rotated 180 degrees
                img = img.rotate(180, expand=True)
            elif orientation == 4:
                # Mirrored vertically
                img = img.transpose(Image.FLIP_TOP_BOTTOM)
            elif orientation == 5:
                # Mirrored horizontally then rotated 90 CCW
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
                img = img.rotate(90, expand=True)
            elif orientation == 6:
                # Rotated 90 degrees CCW (or 270 CW)
                img = img.rotate(270, expand=True)
            elif orientation == 7:
                # Mirrored horizontally then rotated 90 CW
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
                img = img.rotate(270, expand=True)
            elif orientation == 8:
                # Rotated 90 degrees CW
                img = img.rotate(90, expand=True)

            # Reset orientation tag to normal after applying the rotation
            if orientation != 1:
                logger.info(f"Fixed image orientation (EXIF tag: {orientation})")
                # Remove EXIF orientation tag by creating new EXIF data
                exif[0x0112] = 1

    except (AttributeError, KeyError, IndexError) as e:
        # Image doesn't have EXIF data or orientation tag, that's fine
        logger.debug(f"No EXIF orientation data found: {str(e)}")

    return img


def optimize_file_for_upload(file_data, file_ext, original_filename):
    """
    Optimize file size for upload while maintaining quality.

    For images: Fixes orientation, resizes and compresses if >1MB

    Args:
        file_data: bytes - The original file data
        file_ext: str - File extension (e.g., 'jpg', 'pdf')
        original_filename: str - Original filename for logging

    Returns:
        tuple: (optimized_data: bytes, was_optimized: bool, original_size: int, new_size: int)
    """
    original_size = len(file_data)
    original_size_mb = original_size / (1024 * 1024)

    # Define size thresholds
    IMAGE_THRESHOLD_MB = 1.0  # Optimize images larger than 1MB

    # Image optimization
    if file_ext in ['jpg', 'jpeg', 'png', 'gif', 'webp'] and PIL_AVAILABLE:
        try:
            # ALWAYS load and fix orientation for images, regardless of size
            # This ensures mobile photos are displayed correctly
            logger.info(f"Processing image {original_filename}: {original_size_mb:.2f}MB")

            # Load image
            img = Image.open(io.BytesIO(file_data))

            # Fix orientation FIRST (critical for mobile photos)
            img = fix_image_orientation(img)

            needs_optimization = original_size_mb > IMAGE_THRESHOLD_MB

            # Resize if needed
            if needs_optimization:
                # Calculate new dimensions (max 2000px on longest side)
                max_dimension = 2000
                ratio = min(max_dimension / img.width, max_dimension / img.height, 1.0)

                if ratio < 1.0:
                    new_size = (int(img.width * ratio), int(img.height * ratio))
                    img = img.resize(new_size, Image.Resampling.LANCZOS)
                    logger.info(f"  Resized from original to {new_size}")

            # Convert to RGB if necessary (for JPEG)
            if img.mode in ('RGBA', 'P', 'LA'):
                # Create white background
                rgb_img = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                rgb_img.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = rgb_img

            # Save processed version (with orientation fix applied)
            output = io.BytesIO()

            # Use appropriate format and quality
            if file_ext in ['jpg', 'jpeg']:
                img.save(output, format='JPEG', quality=85, optimize=True)
            elif file_ext == 'png':
                img.save(output, format='PNG', optimize=True, compress_level=6)
            elif file_ext == 'webp':
                img.save(output, format='WEBP', quality=85)
            else:
                img.save(output, format=img.format or 'JPEG', quality=85, optimize=True)

            processed_data = output.getvalue()
            new_size = len(processed_data)
            new_size_mb = new_size / (1024 * 1024)

            # Return processed image (with orientation fix always applied)
            if needs_optimization and new_size < original_size:
                reduction_pct = ((original_size - new_size) / original_size) * 100
                logger.info(
                    f"  ✓ Image optimized: {original_size_mb:.2f}MB → {new_size_mb:.2f}MB ({reduction_pct:.1f}% reduction)")
                return (processed_data, True, original_size, new_size)
            else:
                # Even if size didn't reduce, we still fixed orientation
                logger.info(f"  ✓ Image processed (orientation fixed): {original_size_mb:.2f}MB")
                return (processed_data, True, original_size, new_size)

        except Exception as e:
            logger.warning(f"Image processing failed: {str(e)}, using original")
            return (file_data, False, original_size, original_size)

    # No optimization needed or possible
    return (file_data, False, original_size, original_size)


def apply_category_percentage_markup(amount, category_id, connection):
    """
    Apply percentage markup to an amount based on category settings.

    For example, if Fuel category has 1% markup and amount is 1000:
    - Result will be 1000 + (1000 * 1%) = 1010

    Args:
        amount: Decimal - The base amount
        category_id: int or None - Category ID to check for markup
        connection: MySQL connection object

    Returns:
        Decimal - Amount with markup applied (or original amount if no markup)
    """
    if not category_id or not amount or amount == 0:
        return amount

    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute(
            "SELECT percentage_markup FROM categories WHERE id = %s",
            (category_id,)
        )
        result = cursor.fetchone()
        cursor.close()

        if result and result.get('percentage_markup'):
            markup_percentage = Decimal(str(result['percentage_markup']))
            if markup_percentage > 0:
                markup_amount = (amount * markup_percentage) / Decimal('100')
                final_amount = amount + markup_amount
                logger.info(
                    f"Applied {markup_percentage}% markup to category {category_id}: "
                    f"{amount} + {markup_amount} = {final_amount}"
                )
                return final_amount
    except Exception as e:
        logger.error(f"Error applying category markup: {str(e)}")

    return amount


def get_setting(key, default=None):
    """Read a single value from the app_settings table. Returns *default* when
    the table does not exist yet, the key is missing, or the DB is unreachable."""
    connection = get_db_connection()
    if not connection:
        return default
    cursor = None
    try:
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT value FROM app_settings WHERE setting_key = %s", (key,))
        row = cursor.fetchone()
        return row['value'] if row else default
    except Exception as e:
        logger.warning(f"get_setting('{key}'): {e} — using default {default}")
        return default
    finally:
        if cursor:
            cursor.close()
        connection.close()


# Initialize Appwrite client
if APPWRITE_AVAILABLE:
    try:
        # Load from environment variables
        appwrite_endpoint = os.environ.get('APPWRITE_ENDPOINT')
        appwrite_project_id = os.environ.get('APPWRITE_PROJECT_ID')
        appwrite_api_key = os.environ.get('APPWRITE_API_KEY')
        APPWRITE_BUCKET_ID = os.environ.get('APPWRITE_BUCKET_ID')

        if all([appwrite_endpoint, appwrite_project_id, appwrite_api_key, APPWRITE_BUCKET_ID]):
            appwrite_client = Client()
            appwrite_client.set_endpoint(appwrite_endpoint)
            appwrite_client.set_project(appwrite_project_id)
            appwrite_client.set_key(appwrite_api_key)

            appwrite_storage = Storage(appwrite_client)
            logger.info(f"✅ Appwrite storage initialized (endpoint: {appwrite_endpoint})")
        else:
            logger.warning("Appwrite credentials incomplete. Set them in .env file.")
            appwrite_storage = None
    except Exception as e:
        logger.error(f"Failed to initialize Appwrite client: {e}")
        appwrite_storage = None


def login_required(f):
    """Decorator to require login for routes."""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    """Decorator to require admin privileges for routes.

    Uses the ``is_admin`` flag cached in the session at login time so that
    no extra database round-trip is needed on every admin request.
    """

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please login first.', 'warning')
            return redirect(url_for('login'))

        if not session.get('is_admin'):
            flash('Access denied. Admin privileges required.', 'danger')
            return redirect(url_for('dashboard'))

        return f(*args, **kwargs)

    return decorated_function


def token_required(f):
    """Decorator to require token authentication for routes.
    Extracts token from Authorization header: Bearer <token>

    IMPORTANT: This must be defined BEFORE any endpoints that use it.
    """

    @wraps(f)
    def decorated_function(*args, **kwargs):
        import jwt

        token = None

        # Get token from Authorization header
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                # Expected format: "Bearer <token>"
                token = auth_header.split(' ')[1]
            except IndexError:
                return jsonify({'error': 'Invalid authorization header format. Use: Bearer <token>'}), 401

        if not token:
            return jsonify({'error': 'Token is missing. Please provide token in Authorization header.'}), 401

        try:
            # Decode token (also verifies signature and expiry)
            payload = jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token has expired. Please generate a new token.'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token. Please provide a valid token.'}), 401
        except Exception as e:
            logger.error(f"Error validating token: {str(e)}")
            return jsonify({'error': 'Token validation failed'}), 401

        # Validate token against the tokens table
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Token validation failed'}), 401
        cursor = connection.cursor(dictionary=True)
        try:
            cursor.execute(
                "SELECT id, is_revoked, expires_at < UTC_TIMESTAMP() AS is_expired FROM tokens WHERE token = %s AND user_id = %s",
                (token, payload['user_id'])
            )
            token_record = cursor.fetchone()

            if not token_record:
                return jsonify({'error': 'Token not recognized. Please generate a new token.'}), 401
            if token_record['is_revoked']:
                return jsonify({'error': 'Token has been revoked. Please generate a new token.'}), 401
            if token_record['is_expired']:
                return jsonify({'error': 'Token has expired. Please generate a new token.'}), 401

            cursor.execute(
                "UPDATE tokens SET last_used_at = CURRENT_TIMESTAMP WHERE id = %s",
                (token_record['id'],)
            )
            connection.commit()
        except Exception as e:
            logger.error(f"Error validating token in database: {str(e)}")
            return jsonify({'error': 'Token validation failed'}), 401
        finally:
            cursor.close()
            connection.close()

        # Store user info in request context
        request.current_user = {
            'user_id': payload['user_id'],
            'username': payload['username'],
            'is_admin': payload.get('is_admin', False)
        }

        return f(*args, **kwargs)

    return decorated_function


# Utility function to serialize Decimal for JSON
def decimal_default(obj):
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError


def log_transaction_audit(cursor, transaction_id, user_id, action, field_name=None, old_value=None, new_value=None):
    """
    Log transaction changes to audit trail.

    Args:
        cursor: Database cursor
        transaction_id: ID of the transaction (None for DELETE after completion)
        user_id: ID of the user making the change
        action: Type of action (CREATE, UPDATE, DELETE)
        field_name: Name of the field changed (None for CREATE/DELETE)
        old_value: Previous value (None for CREATE)
        new_value: New value (None for DELETE)
    """
    try:
        # Get request metadata
        ip_address = request.remote_addr if request else None
        user_agent = request.headers.get('User-Agent') if request else None

        # Convert values to strings for storage
        old_value_str = str(old_value) if old_value is not None else None
        new_value_str = str(new_value) if new_value is not None else None

        cursor.execute("""
                       INSERT INTO transaction_audit_logs
                       (transaction_id, user_id, action, field_name, old_value, new_value, ip_address, user_agent)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                       """, (transaction_id, user_id, action, field_name, old_value_str, new_value_str, ip_address,
                             user_agent))

        logger.info(f"Audit log created: {action} on transaction {transaction_id} by user {user_id}")
    except Exception as e:
        logger.error(f"Failed to create audit log: {e}")
        # Don't fail the main transaction if audit logging fails


# Routes


@app.before_request
def make_session_permanent():
    """Ensure authenticated sessions always use a persistent cookie.

    Without this, non-'Remember Me' sessions use browser-session cookies
    that are deleted when the tab or browser is closed. By always marking
    authenticated sessions as permanent, the cookie is sent with a Max-Age
    equal to PERMANENT_SESSION_LIFETIME (365 days) so the user stays
    logged in across tab/browser restarts.
    """
    if 'user_id' in session:
        session.permanent = True


@app.route('/')
def index():
    """Landing page - redirect to login or dashboard/mobile based on device."""
    logger.info(f"Index route accessed - User logged in: {'user_id' in session}")
    if 'user_id' in session:
        # Detect if mobile device from user agent
        user_agent = request.headers.get('User-Agent', '').lower()
        is_mobile = any(device in user_agent for device in
                        ['android', 'webos', 'iphone', 'ipad', 'ipod', 'blackberry', 'windows phone'])

        if is_mobile:
            logger.info(f"Redirecting user {session.get('user_id')} to mobile view")
            return redirect(url_for('mobile'))
        logger.info(f"Redirecting user {session.get('user_id')} to dashboard")
        return redirect(url_for('dashboard'))
    logger.info("No user session found, redirecting to login")
    return redirect(url_for('login'))


@app.route('/register', methods=['GET', 'POST'])
@limiter.limit(RATE_LIMIT_REGISTER)
def register():
    """User registration."""
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username')
        email = data.get('email')
        password = data.get('password')

        connection = get_db_connection()
        if connection:
            cursor = connection.cursor()
            try:
                # Check if user already exists
                cursor.execute("SELECT id FROM users WHERE username = %s OR email = %s",
                               (username, email))
                if cursor.fetchone():
                    return jsonify({'error': 'Username or email already exists'}), 400

                # Create new user (deactivated by default, requires admin activation)
                password_hash = generate_password_hash(password)
                cursor.execute(
                    "INSERT INTO users (username, email, password_hash, is_active) VALUES (%s, %s, %s, %s)",
                    (username, email, password_hash, False)
                )
                connection.commit()

                logger.info(f"New user registered: {username} ({email}) - Account created in deactivated state")

                return jsonify({'message': 'Registration successful. Your account is pending admin approval.'}), 201
            except Error as e:
                return jsonify({'error': str(e)}), 500
            finally:
                cursor.close()
                connection.close()

        return jsonify({'error': 'Database connection failed'}), 500

    return render_template('register.html')


@app.route('/login', methods=['GET', 'POST'])
@limiter.limit(RATE_LIMIT_LOGIN)
def login():
    """User login."""
    if request.method == 'POST':
        data = request.get_json() if request.is_json else request.form
        username = data.get('username')
        password = data.get('password')
        remember_me = bool(data.get('remember_me', False))

        logger.info(f"Login attempt for username: {username}, remember_me: {remember_me}")

        connection = get_db_connection()
        if connection:
            cursor = connection.cursor(dictionary=True)
            try:
                # Check if username is an email or username
                cursor.execute("SELECT * FROM users WHERE username = %s OR email = %s",
                               (username, username))
                user = cursor.fetchone()

                if user and check_password_hash(user['password_hash'], password):
                    # Check if user account is active
                    if not user.get('is_active', True):
                        logger.warning(f"Login failed for username: {username} - Account is deactivated")
                        return jsonify(
                            {'error': 'Your account has been deactivated. Please contact an administrator.'}), 403

                    # Update last_login timestamp
                    cursor.execute("UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE id = %s", (user['id'],))
                    connection.commit()

                    # Always set session as permanent so the cookie is sent
                    # with Max-Age (persists across tab/browser restarts).
                    session.permanent = True
                    session['user_id'] = user['id']
                    session['username'] = user['username']
                    session['is_admin'] = user.get('is_admin', False)
                    session.modified = True
                    logger.info(
                        f"Login successful for user: {username} (ID: {user['id']}), is_admin: {user.get('is_admin', False)}")

                    return jsonify({'message': 'Login successful'}), 200
                else:
                    logger.warning(f"Login failed for username: {username} - Invalid credentials")
                    return jsonify({'error': 'Invalid credentials'}), 401
            except Error as e:
                logger.error(f"Database error during login: {str(e)}")
                return jsonify({'error': str(e)}), 500
            finally:
                cursor.close()
                connection.close()
        else:
            logger.error("Failed to establish database connection during login")
            return jsonify({'error': 'Database connection failed'}), 500

    # If user is already authenticated, redirect to dashboard/mobile
    if 'user_id' in session:
        user_agent = request.headers.get('User-Agent', '').lower()
        is_mobile = any(device in user_agent for device in
                        ['android', 'webos', 'iphone', 'ipad', 'ipod', 'blackberry', 'windows phone'])
        if is_mobile:
            return redirect(url_for('mobile'))
        return redirect(url_for('dashboard'))

    return render_template('login.html')


@app.route('/logout')
def logout():
    """User logout."""
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))


@app.route('/api/change-password', methods=['POST'])
@login_required
@limiter.limit(RATE_LIMIT_CHANGE_PASSWORD)
def change_password():
    """Change user password."""
    data = request.get_json()
    current_password = data.get('current_password')
    new_password = data.get('new_password')

    if not current_password or not new_password:
        return jsonify({'error': 'Current and new passwords are required'}), 400

    if len(new_password) < 6:
        return jsonify({'error': 'New password must be at least 6 characters long'}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        # Get current user
        cursor.execute("SELECT password_hash FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Verify current password
        if not check_password_hash(user['password_hash'], current_password):
            return jsonify({'error': 'Current password is incorrect'}), 401

        # Update password
        new_password_hash = generate_password_hash(new_password)
        cursor.execute(
            "UPDATE users SET password_hash = %s WHERE id = %s",
            (new_password_hash, user_id)
        )
        connection.commit()

        return jsonify({'message': 'Password changed successfully'}), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/user-preferences', methods=['GET'])
@login_required
@limiter.limit(RATE_LIMIT_API)
def get_user_preferences():
    """Get current user's preferences."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        cursor.execute("SELECT default_page FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'error': 'User not found'}), 404

        return jsonify({
            'default_page': user.get('default_page', 'transactions')
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/user-preferences', methods=['PUT'])
@login_required
@limiter.limit(RATE_LIMIT_API)
def update_user_preferences():
    """Update current user's preferences."""
    data = request.get_json()
    default_page = data.get('default_page')

    # Validate default_page value
    valid_pages = ['transactions', 'tax', 'reports', 'rateTrends']
    if default_page not in valid_pages:
        return jsonify({
            'error': f'Invalid default_page. Must be one of: {", ".join(valid_pages)}'
        }), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        cursor.execute(
            "UPDATE users SET default_page = %s WHERE id = %s",
            (default_page, user_id)
        )
        connection.commit()

        logger.info(f"User {session.get('username')} updated default_page to {default_page}")
        return jsonify({
            'message': 'Preferences updated successfully',
            'default_page': default_page
        }), 200

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard."""
    return render_template('dashboard.html', username=session.get('username'))


@app.route('/mobile')
@login_required
def mobile():
    """Mobile view."""
    return render_template('mobile.html', username=session.get('username'))


@app.route('/admin')
@admin_required
def admin_dashboard():
    """Admin dashboard with server-side data."""
    connection = get_db_connection()
    users = []
    audit_logs = []
    settings = {}
    error_message = None

    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            # Fetch users
            cursor.execute("""
                           SELECT u.id,
                                  u.username,
                                  u.email,
                                  u.is_admin,
                                  u.is_active,
                                  u.last_login,
                                  u.default_page,
                                  u.created_at,
                                  COUNT(DISTINCT mr.id) as monthly_records_count,
                                  COUNT(DISTINCT t.id)  as transactions_count
                           FROM users u
                                    LEFT JOIN monthly_records mr ON u.id = mr.user_id
                                    LEFT JOIN transactions t ON mr.id = t.monthly_record_id
                           GROUP BY u.id, u.username, u.email, u.is_admin, u.is_active, u.last_login, u.default_page, u.created_at
                           ORDER BY u.created_at DESC
                           """)
            users = cursor.fetchall()

            # Fetch audit logs
            cursor.execute("""
                           SELECT al.id,
                                  al.action,
                                  al.details,
                                  al.created_at,
                                  au.username as admin_username,
                                  tu.username as target_username
                           FROM audit_logs al
                                    JOIN users au ON al.admin_user_id = au.id
                                    LEFT JOIN users tu ON al.target_user_id = tu.id
                           ORDER BY al.created_at DESC LIMIT 50
                           """)
            audit_logs = cursor.fetchall()

            # Fetch app settings
            cursor.execute("SELECT setting_key, value, description FROM app_settings ORDER BY setting_key")
            settings = {row['setting_key']: row for row in cursor.fetchall()}

        except Error as e:
            logger.error(f"Error fetching admin data: {str(e)}")
            error_message = str(e)
        finally:
            cursor.close()
            connection.close()
    else:
        error_message = "Database connection failed"

    return render_template('admin.html',
                           username=session.get('username'),
                           users=users,
                           audit_logs=audit_logs,
                           settings=settings,
                           error_message=error_message,
                           current_user_id=session.get('user_id'))


@app.route('/api/admin/users', methods=['GET'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def get_all_users():
    """Get all users with their details."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)

    try:
        cursor.execute("""
                       SELECT u.id,
                              u.username,
                              u.email,
                              u.is_admin,
                              u.is_active,
                              u.last_login,
                              u.created_at,
                              COUNT(DISTINCT mr.id) as monthly_records_count,
                              COUNT(DISTINCT t.id)  as transactions_count
                       FROM users u
                                LEFT JOIN monthly_records mr ON u.id = mr.user_id
                                LEFT JOIN transactions t ON mr.id = t.monthly_record_id
                       GROUP BY u.id, u.username, u.email, u.is_admin, u.is_active, u.last_login, u.created_at
                       ORDER BY u.created_at DESC
                       """)

        users = cursor.fetchall()
        return jsonify(users)

    except Error as e:
        logger.error(f"Error fetching users: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def log_audit(admin_user_id, action, target_user_id=None, details=None):
    """Helper function to log admin actions."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor()
        try:
            cursor.execute("""
                           INSERT INTO audit_logs (admin_user_id, action, target_user_id, details)
                           VALUES (%s, %s, %s, %s)
                           """, (admin_user_id, action, target_user_id, details))
            connection.commit()
        except Error as e:
            logger.error(f"Error logging audit: {str(e)}")
        finally:
            cursor.close()
            connection.close()


# ==================================================
# REGISTER SERVICE ROUTES
# ==================================================
# Register all service routes from the dedicated service modules
# This must be called after all decorators and helper functions are defined
register_exchange_rate_routes(app, login_required, admin_required, token_required, log_audit)
register_tax_routes(app, login_required)


@app.route('/api/admin/users/<int:user_id>/toggle-active', methods=['POST'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def toggle_user_active(user_id):
    """Activate or deactivate a user."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    admin_id = session['user_id']

    try:
        # Prevent admin from deactivating themselves
        if user_id == admin_id:
            return jsonify({'error': 'You cannot deactivate your own account'}), 400

        # Get current user status
        cursor.execute("SELECT username, is_active FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Toggle active status
        new_status = not user['is_active']
        cursor.execute("UPDATE users SET is_active = %s WHERE id = %s", (new_status, user_id))
        connection.commit()

        # Log the action
        action = f"{'Activated' if new_status else 'Deactivated'} user"
        log_audit(admin_id, action, user_id,
                  f"User '{user['username']}' status changed to {'active' if new_status else 'inactive'}")

        logger.info(f"Admin {admin_id} {action.lower()} user {user_id} ({user['username']})")

        return jsonify({
            'message': f"User {'activated' if new_status else 'deactivated'} successfully",
            'is_active': new_status
        })

    except Error as e:
        logger.error(f"Error toggling user active status: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/users/<int:user_id>/toggle-admin', methods=['POST'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def toggle_user_admin(user_id):
    """Grant or revoke admin privileges."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    admin_id = session['user_id']

    try:
        # Prevent admin from revoking their own admin status
        if user_id == admin_id:
            return jsonify({'error': 'You cannot modify your own admin status'}), 400

        # Get current user status
        cursor.execute("SELECT username, is_admin FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Toggle admin status
        new_status = not user['is_admin']
        cursor.execute("UPDATE users SET is_admin = %s WHERE id = %s", (new_status, user_id))
        connection.commit()

        # Log the action
        action = f"{'Granted' if new_status else 'Revoked'} admin privileges"
        log_audit(admin_id, action, user_id, f"User '{user['username']}' admin status changed to {new_status}")

        logger.info(f"Admin {admin_id} {action.lower()} for user {user_id} ({user['username']})")

        return jsonify({
            'message': f"Admin privileges {'granted' if new_status else 'revoked'} successfully",
            'is_admin': new_status
        })

    except Error as e:
        logger.error(f"Error toggling user admin status: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def delete_user(user_id):
    """Delete a user and all their data."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    admin_id = session['user_id']

    try:
        # Prevent admin from deleting themselves
        if user_id == admin_id:
            return jsonify({'error': 'You cannot delete your own account'}), 400

        # Get user details before deletion
        cursor.execute("SELECT username, email FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Get all attachments from user's transactions before deletion
        cursor.execute("""
            SELECT t.attachments 
            FROM transactions t
            INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
            WHERE mr.user_id = %s AND t.attachments IS NOT NULL
        """, (user_id,))
        transactions_with_attachments = cursor.fetchall()

        # Delete all attachments from Appwrite before deleting user
        if appwrite_file_service.is_available() and transactions_with_attachments:
            deleted_count = 0
            for txn in transactions_with_attachments:
                attachments_value = txn['attachments']
                # Split comma-separated GUIDs
                attachment_guids = [guid.strip() for guid in attachments_value.split(',') if guid.strip()]

                for attachment_guid in attachment_guids:
                    success, error = appwrite_file_service.delete_file(attachment_guid)
                    if success:
                        deleted_count += 1
                    else:
                        logger.warning(f"Failed to delete attachment {attachment_guid}: {error}")

            if deleted_count > 0:
                logger.info(f"Deleted {deleted_count} attachment(s) from Appwrite for user {user_id}")

        # Delete user (cascading will handle related records)
        cursor.execute("DELETE FROM users WHERE id = %s", (user_id,))
        connection.commit()

        # Log the action
        log_audit(admin_id, 'Deleted user', None, f"User '{user['username']}' ({user['email']}) permanently deleted")

        logger.info(f"Admin {admin_id} deleted user {user_id} ({user['username']})")

        return jsonify({'message': 'User deleted successfully'})

    except Error as e:
        logger.error(f"Error deleting user: {str(e)}")
        connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/users/<int:user_id>/default-page', methods=['PUT'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def admin_update_user_default_page(user_id):
    """Update a user's default landing page (admin only)."""
    data = request.get_json()
    default_page = data.get('default_page')

    # Validate default_page value
    valid_pages = ['transactions', 'tax', 'reports', 'rateTrends']
    if default_page not in valid_pages:
        return jsonify({
            'error': f'Invalid default_page. Must be one of: {", ".join(valid_pages)}'
        }), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    admin_id = session['user_id']

    try:
        # Get user details
        cursor.execute("SELECT username FROM users WHERE id = %s", (user_id,))
        user = cursor.fetchone()

        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Update default page
        cursor.execute(
            "UPDATE users SET default_page = %s WHERE id = %s",
            (default_page, user_id)
        )
        connection.commit()

        # Log the action
        log_audit(admin_id, 'Updated user preferences', user_id,
                  f"User '{user['username']}' default page changed to '{default_page}'")

        logger.info(f"Admin {admin_id} updated default_page for user {user_id} ({user['username']}) to {default_page}")

        return jsonify({
            'message': 'Default page updated successfully',
            'default_page': default_page
        }), 200

    except Error as e:
        logger.error(f"Error updating user default page: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/audit-logs', methods=['GET'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def get_audit_logs():
    """Get audit logs of admin actions."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)

    try:
        limit = request.args.get('limit', 100, type=int)

        cursor.execute("""
                       SELECT al.id,
                              al.action,
                              al.details,
                              al.created_at,
                              au.username as admin_username,
                              tu.username as target_username
                       FROM audit_logs al
                                JOIN users au ON al.admin_user_id = au.id
                                LEFT JOIN users tu ON al.target_user_id = tu.id
                       ORDER BY al.created_at DESC
                           LIMIT %s
                       """, (limit,))

        logs = cursor.fetchall()
        return jsonify(logs)

    except Error as e:
        logger.error(f"Error fetching audit logs: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/settings', methods=['GET'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def get_admin_settings():
    """Get all application settings."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("SELECT setting_key, value, description, updated_at FROM app_settings ORDER BY setting_key")
        settings = cursor.fetchall()
        for row in settings:
            if row.get('updated_at'):
                row['updated_at'] = str(row['updated_at'])
        return jsonify(settings), 200
    except Error as e:
        logger.error(f"Error fetching settings: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/settings/upload-mode', methods=['GET'])
def get_upload_mode():
    """Public endpoint to get bill upload mode (sequential vs batch).
    No authentication required - used by client-side to determine upload strategy."""
    mode = get_setting('bill_upload_mode', 'sequential')
    logger.info(f"Upload mode requested: returning '{mode}'")
    return jsonify({'upload_mode': mode}), 200


@app.route('/api/admin/settings/<string:key>', methods=['PUT'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def update_admin_setting(key):
    """Update a single application setting.  Only keys that already exist in
    app_settings may be written — arbitrary keys are rejected."""
    data = request.get_json()
    if not data or 'value' not in data:
        return jsonify({'error': "'value' field is required"}), 400

    new_value = str(data['value'])

    # Key-specific validation
    if key == 'bill_upload_mode':
        if new_value not in ('sequential', 'batch'):
            return jsonify({'error': "Value must be 'sequential' or 'batch'"}), 400

    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    try:
        # Guard: only allow updating pre-existing keys
        cursor.execute("SELECT setting_key FROM app_settings WHERE setting_key = %s", (key,))
        if not cursor.fetchone():
            return jsonify({'error': f'Unknown setting: {key}'}), 404

        cursor.execute("UPDATE app_settings SET value = %s WHERE setting_key = %s", (new_value, key))
        connection.commit()

        username = session.get('username')
        logger.info(f"Admin setting updated: {key} = {new_value} (by {username})")
        log_audit(session['user_id'], 'UPDATE_SETTING', details=f'{key} = {new_value}')

        return jsonify({'message': 'Setting updated', 'key': key, 'value': new_value}), 200
    except Error as e:
        logger.error(f"Error updating setting '{key}': {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/db-backup', methods=['GET'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def admin_db_backup():
    """Generate a full MySQL database backup in MySQL Workbench compatible format.

    Includes: table structures, data, views, stored procedures, functions,
    triggers, and events.  The output mirrors the format produced by
    mysqldump / MySQL Workbench so that it can be imported back seamlessly.
    """
    import subprocess
    import shutil

    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    # Log the backup action
    log_audit(session['user_id'], 'DATABASE_BACKUP', details='Full database backup downloaded')

    db_name = DB_CONFIG['database']
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{db_name}_{timestamp}.sql"

    # Try mysqldump first (produces the most compatible output)
    mysqldump_path = shutil.which('mysqldump')
    if mysqldump_path:
        try:
            cmd = [
                mysqldump_path,
                '--host', DB_CONFIG['host'],
                '--port', str(DB_CONFIG['port']),
                '--user', DB_CONFIG['user'],
                f'--password={DB_CONFIG["password"]}',
                '--routines',
                '--events',
                '--triggers',
                '--single-transaction',
                '--set-gtid-purged=OFF',
                '--column-statistics=0',
                '--skip-lock-tables',
                '--default-character-set=utf8mb4',
                db_name
            ]
            result = subprocess.run(cmd, capture_output=True, timeout=300)
            if result.returncode == 0 and result.stdout:
                response = make_response(result.stdout)
                response.headers['Content-Type'] = 'application/sql'
                response.headers['Content-Disposition'] = f'attachment; filename={filename}'
                return response
            logger.warning(f"mysqldump failed (rc={result.returncode}): {result.stderr[:500]}")
        except Exception as e:
            logger.warning(f"mysqldump error: {e}")

    # Fallback: pure-Python dump in MySQL Workbench format
    cursor = connection.cursor()
    try:
        output = io.StringIO()

        # ── Header ──────────────────────────────────────────────────
        cursor.execute("SELECT VERSION()")
        mysql_version = cursor.fetchone()[0]

        output.write("-- MySQL dump\n")
        output.write(f"-- Host: {DB_CONFIG['host']}    Database: {db_name}\n")
        output.write("-- ------------------------------------------------------\n")
        output.write(f"-- Server version\t{mysql_version}\n\n")

        output.write("/*!40101 SET @OLD_CHARACTER_SET_CLIENT=@@CHARACTER_SET_CLIENT */;\n")
        output.write("/*!40101 SET @OLD_CHARACTER_SET_RESULTS=@@CHARACTER_SET_RESULTS */;\n")
        output.write("/*!40101 SET @OLD_COLLATION_CONNECTION=@@COLLATION_CONNECTION */;\n")
        output.write("/*!40101 SET NAMES utf8mb4 */;\n")
        output.write("/*!40103 SET @OLD_TIME_ZONE=@@TIME_ZONE */;\n")
        output.write("/*!40103 SET TIME_ZONE='+00:00' */;\n")
        output.write("/*!40014 SET @OLD_UNIQUE_CHECKS=@@UNIQUE_CHECKS, UNIQUE_CHECKS=0 */;\n")
        output.write("/*!40014 SET @OLD_FOREIGN_KEY_CHECKS=@@FOREIGN_KEY_CHECKS, FOREIGN_KEY_CHECKS=0 */;\n")
        output.write("/*!40101 SET @OLD_SQL_MODE=@@SQL_MODE, SQL_MODE='NO_AUTO_VALUE_ON_ZERO' */;\n")
        output.write("/*!40111 SET @OLD_SQL_NOTES=@@SQL_NOTES, SQL_NOTES=0 */;\n")
        output.write(
            "/*!80000 SET @OLD_SQL_REQUIRE_PRIMARY_KEY=@@SQL_REQUIRE_PRIMARY_KEY, SQL_REQUIRE_PRIMARY_KEY=0 */;\n\n")

        logger.info(f"Starting comprehensive database backup for: {db_name}")

        # ── Collect object lists ────────────────────────────────────
        cursor.execute("SHOW FULL TABLES WHERE Table_type = 'BASE TABLE'")
        tables = [row[0] for row in cursor.fetchall()]
        logger.info(f"Found {len(tables)} tables to backup")

        cursor.execute("SHOW FULL TABLES WHERE Table_type = 'VIEW'")
        views = [row[0] for row in cursor.fetchall()]
        logger.info(f"Found {len(views)} views to backup")

        # Get stored procedures
        cursor.execute(
            "SELECT ROUTINE_NAME FROM INFORMATION_SCHEMA.ROUTINES "
            "WHERE ROUTINE_SCHEMA = %s AND ROUTINE_TYPE = 'PROCEDURE'",
            (db_name,)
        )
        procedures = [row[0] for row in cursor.fetchall()]
        logger.info(f"Found {len(procedures)} stored procedures to backup")

        # Get functions
        cursor.execute(
            "SELECT ROUTINE_NAME FROM INFORMATION_SCHEMA.ROUTINES "
            "WHERE ROUTINE_SCHEMA = %s AND ROUTINE_TYPE = 'FUNCTION'",
            (db_name,)
        )
        functions = [row[0] for row in cursor.fetchall()]
        logger.info(f"Found {len(functions)} functions to backup")

        # Get triggers
        cursor.execute(
            "SELECT TRIGGER_NAME FROM INFORMATION_SCHEMA.TRIGGERS "
            "WHERE TRIGGER_SCHEMA = %s",
            (db_name,)
        )
        triggers = [row[0] for row in cursor.fetchall()]
        logger.info(f"Found {len(triggers)} triggers to backup")

        # Get events
        cursor.execute(
            "SELECT EVENT_NAME FROM INFORMATION_SCHEMA.EVENTS "
            "WHERE EVENT_SCHEMA = %s",
            (db_name,)
        )
        events = [row[0] for row in cursor.fetchall()]
        logger.info(f"Found {len(events)} events to backup")

        # ── Helper: escape a Python value for a SQL INSERT ─────────
        def sql_escape(val):
            if val is None:
                return 'NULL'
            if isinstance(val, bool):
                return '1' if val else '0'
            if isinstance(val, (int, float, Decimal)):
                return str(val)
            if isinstance(val, (bytes, bytearray)):
                hex_str = val.hex()
                return f"X'{hex_str}'" if hex_str else "''"
            if isinstance(val, datetime):
                return f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'"
            if isinstance(val, timedelta):
                total = int(val.total_seconds())
                h, rem = divmod(abs(total), 3600)
                m, s = divmod(rem, 60)
                sign = '-' if total < 0 else ''
                return f"'{sign}{h:02d}:{m:02d}:{s:02d}'"
            # String — escape special characters
            s = str(val)
            s = s.replace('\\', '\\\\')
            s = s.replace("'", "\\'")
            s = s.replace('\n', '\\n')
            s = s.replace('\r', '\\r')
            s = s.replace('\x00', '\\0')
            s = s.replace('\x1a', '\\Z')
            return f"'{s}'"

        # ── Tables: structure + data ────────────────────────────────
        for table in tables:
            output.write(f"--\n-- Table structure for table `{table}`\n--\n\n")
            output.write(f"DROP TABLE IF EXISTS `{table}`;\n")
            output.write("/*!40101 SET @saved_cs_client     = @@character_set_client */;\n")
            output.write("/*!40101 SET character_set_client = utf8 */;\n")

            cursor.execute(f"SHOW CREATE TABLE `{table}`")
            create_stmt = cursor.fetchone()[1]
            output.write(f"{create_stmt};\n")
            output.write("/*!40101 SET character_set_client = @saved_cs_client */;\n\n")

            # Data
            cursor.execute(f"SELECT * FROM `{table}`")
            rows = cursor.fetchall()
            if rows:
                # Get column names
                col_names = [desc[0] for desc in cursor.description]
                col_list = ', '.join(f'`{c}`' for c in col_names)

                output.write(f"--\n-- Dumping data for table `{table}`\n--\n\n")
                output.write(f"LOCK TABLES `{table}` WRITE;\n")
                output.write(f"/*!40000 ALTER TABLE `{table}` DISABLE KEYS */;\n")

                # Write INSERT statements in batches (extended inserts)
                batch_size = 100
                for i in range(0, len(rows), batch_size):
                    batch = rows[i:i + batch_size]
                    output.write(f"INSERT INTO `{table}` ({col_list}) VALUES\n")
                    value_lines = []
                    for row in batch:
                        vals = ', '.join(sql_escape(v) for v in row)
                        value_lines.append(f"({vals})")
                    output.write(',\n'.join(value_lines))
                    output.write(';\n')

                output.write(f"/*!40000 ALTER TABLE `{table}` ENABLE KEYS */;\n")
                output.write("UNLOCK TABLES;\n\n")

        logger.info(f"Completed backing up {len(tables)} tables and their data")

        # ── Temporary Tables (if any exist) ─────────────────────────
        output.write("--\n-- Temporary tables\n--\n\n")

        # ── Views ───────────────────────────────────────────────────
        if views:
            output.write("--\n-- Final view structure for views\n--\n\n")
            for view in views:
                output.write(f"--\n-- Final view structure for view `{view}`\n--\n\n")
                output.write(f"/*!50001 DROP VIEW IF EXISTS `{view}`*/;\n")
                output.write("/*!50001 SET @saved_cs_client          = @@character_set_client */;\n")
                output.write("/*!50001 SET @saved_cs_results         = @@character_set_results */;\n")
                output.write("/*!50001 SET @saved_col_connection     = @@collation_connection */;\n")
                output.write("/*!50001 SET character_set_client      = utf8mb4 */;\n")
                output.write("/*!50001 SET character_set_results     = utf8mb4 */;\n")
                output.write("/*!50001 SET collation_connection      = utf8mb4_0900_ai_ci */;\n")
                try:
                    cursor.execute(f"SHOW CREATE VIEW `{view}`")
                    result = cursor.fetchone()
                    if result and len(result) >= 2:
                        create_view = result[1]
                        # Format as MySQL Workbench does
                        output.write(f"/*!50001 {create_view} */;\n")
                        output.write("/*!50001 SET character_set_client      = @saved_cs_client */;\n")
                        output.write("/*!50001 SET character_set_results     = @saved_cs_results */;\n")
                        output.write("/*!50001 SET collation_connection      = @saved_col_connection */;\n\n")
                except Exception as e:
                    logger.warning(f"Error exporting view `{view}`: {e}")
                    output.write(f"-- Error exporting view `{view}`: {e}\n\n")
            logger.info(f"Completed backing up {len(views)} views")

        # ── Stored Procedures ───────────────────────────────────────
        if procedures:
            output.write("--\n-- Dumping routines for database '" + db_name + "'\n--\n")
            for proc in procedures:
                output.write(f"--\n-- Procedure `{proc}`\n--\n\n")
                output.write(f"/*!50003 DROP PROCEDURE IF EXISTS `{proc}` */;\n")
                output.write("/*!50003 SET @saved_cs_client      = @@character_set_client */ ;\n")
                output.write("/*!50003 SET @saved_cs_results     = @@character_set_results */ ;\n")
                output.write("/*!50003 SET @saved_col_connection = @@collation_connection */ ;\n")
                output.write("/*!50003 SET character_set_client  = utf8mb4 */ ;\n")
                output.write("/*!50003 SET character_set_results = utf8mb4 */ ;\n")
                output.write("/*!50003 SET collation_connection  = utf8mb4_0900_ai_ci */ ;\n")
                output.write("/*!50003 SET @saved_sql_mode       = @@sql_mode */ ;\n")
                output.write(
                    "/*!50003 SET sql_mode              = 'ONLY_FULL_GROUP_BY,STRICT_TRANS_TABLES,NO_ZERO_IN_DATE,NO_ZERO_DATE,ERROR_FOR_DIVISION_BY_ZERO,NO_ENGINE_SUBSTITUTION' */ ;\n")
                output.write("DELIMITER ;;\n")
                try:
                    cursor.execute(f"SHOW CREATE PROCEDURE `{proc}`")
                    result = cursor.fetchone()
                    if result and len(result) >= 3:
                        create_proc = result[2]
                        output.write(f"{create_proc} ;;\n")
                except Exception as e:
                    logger.warning(f"Error exporting procedure `{proc}`: {e}")
                    output.write(f"-- Error exporting procedure `{proc}`: {e}\n")
                output.write("DELIMITER ;\n")
                output.write("/*!50003 SET sql_mode              = @saved_sql_mode */ ;\n")
                output.write("/*!50003 SET character_set_client  = @saved_cs_client */ ;\n")
                output.write("/*!50003 SET character_set_results = @saved_cs_results */ ;\n")
                output.write("/*!50003 SET collation_connection  = @saved_col_connection */ ;\n\n")
            logger.info(f"Completed backing up {len(procedures)} stored procedures")

        # ── Functions ───────────────────────────────────────────────
        if functions:
            for func in functions:
                output.write(f"--\n-- Function `{func}`\n--\n\n")
                output.write(f"/*!50003 DROP FUNCTION IF EXISTS `{func}` */;\n")
                output.write("/*!50003 SET @saved_cs_client      = @@character_set_client */ ;\n")
                output.write("/*!50003 SET @saved_cs_results     = @@character_set_results */ ;\n")
                output.write("/*!50003 SET @saved_col_connection = @@collation_connection */ ;\n")
                output.write("/*!50003 SET character_set_client  = utf8mb4 */ ;\n")
                output.write("/*!50003 SET character_set_results = utf8mb4 */ ;\n")
                output.write("/*!50003 SET collation_connection  = utf8mb4_0900_ai_ci */ ;\n")
                output.write("/*!50003 SET @saved_sql_mode       = @@sql_mode */ ;\n")
                output.write(
                    "/*!50003 SET sql_mode              = 'ONLY_FULL_GROUP_BY,STRICT_TRANS_TABLES,NO_ZERO_IN_DATE,NO_ZERO_DATE,ERROR_FOR_DIVISION_BY_ZERO,NO_ENGINE_SUBSTITUTION' */ ;\n")
                output.write("DELIMITER ;;\n")
                try:
                    cursor.execute(f"SHOW CREATE FUNCTION `{func}`")
                    result = cursor.fetchone()
                    if result and len(result) >= 3:
                        create_func = result[2]
                        output.write(f"{create_func} ;;\n")
                except Exception as e:
                    logger.warning(f"Error exporting function `{func}`: {e}")
                    output.write(f"-- Error exporting function `{func}`: {e}\n")
                output.write("DELIMITER ;\n")
                output.write("/*!50003 SET sql_mode              = @saved_sql_mode */ ;\n")
                output.write("/*!50003 SET character_set_client  = @saved_cs_client */ ;\n")
                output.write("/*!50003 SET character_set_results = @saved_cs_results */ ;\n")
                output.write("/*!50003 SET collation_connection  = @saved_col_connection */ ;\n\n")
            logger.info(f"Completed backing up {len(functions)} functions")

        # ── Triggers ────────────────────────────────────────────────
        if triggers:
            for trigger in triggers:
                output.write(f"--\n-- Trigger `{trigger}`\n--\n\n")
                output.write("/*!50003 SET @saved_cs_client      = @@character_set_client */ ;\n")
                output.write("/*!50003 SET @saved_cs_results     = @@character_set_results */ ;\n")
                output.write("/*!50003 SET @saved_col_connection = @@collation_connection */ ;\n")
                output.write("/*!50003 SET character_set_client  = utf8mb4 */ ;\n")
                output.write("/*!50003 SET character_set_results = utf8mb4 */ ;\n")
                output.write("/*!50003 SET collation_connection  = utf8mb4_0900_ai_ci */ ;\n")
                output.write("/*!50003 SET @saved_sql_mode       = @@sql_mode */ ;\n")
                output.write(
                    "/*!50003 SET sql_mode              = 'ONLY_FULL_GROUP_BY,STRICT_TRANS_TABLES,NO_ZERO_IN_DATE,NO_ZERO_DATE,ERROR_FOR_DIVISION_BY_ZERO,NO_ENGINE_SUBSTITUTION' */ ;\n")
                output.write("DELIMITER ;;\n")
                try:
                    cursor.execute(f"SHOW CREATE TRIGGER `{trigger}`")
                    result = cursor.fetchone()
                    if result and len(result) >= 3:
                        create_trigger = result[2]
                        output.write(f"/*!50003 {create_trigger} */;;\n")
                except Exception as e:
                    logger.warning(f"Error exporting trigger `{trigger}`: {e}")
                    output.write(f"-- Error exporting trigger `{trigger}`: {e}\n")
                output.write("DELIMITER ;\n")
                output.write("/*!50003 SET sql_mode              = @saved_sql_mode */ ;\n")
                output.write("/*!50003 SET character_set_client  = @saved_cs_client */ ;\n")
                output.write("/*!50003 SET character_set_results = @saved_cs_results */ ;\n")
                output.write("/*!50003 SET collation_connection  = @saved_col_connection */ ;\n\n")
            logger.info(f"Completed backing up {len(triggers)} triggers")

        # ── Events ──────────────────────────────────────────────────
        if events:
            output.write("--\n-- Dumping events for database '" + db_name + "'\n--\n")
            for event in events:
                output.write(f"--\n-- Event `{event}`\n--\n\n")
                output.write(f"/*!50106 DROP EVENT IF EXISTS `{event}` */;\n")
                output.write("DELIMITER ;;\n")
                output.write("/*!50106 SET @save_time_zone= @@TIME_ZONE */ ;;\n")
                output.write("/*!50106 SET TIME_ZONE= 'SYSTEM' */ ;;\n")
                output.write("/*!50106 SET @saved_cs_client      = @@character_set_client */ ;;\n")
                output.write("/*!50106 SET @saved_cs_results     = @@character_set_results */ ;;\n")
                output.write("/*!50106 SET @saved_col_connection = @@collation_connection */ ;;\n")
                output.write("/*!50106 SET character_set_client  = utf8mb4 */ ;;\n")
                output.write("/*!50106 SET character_set_results = utf8mb4 */ ;;\n")
                output.write("/*!50106 SET collation_connection  = utf8mb4_0900_ai_ci */ ;;\n")
                output.write("/*!50106 SET @saved_sql_mode       = @@sql_mode */ ;;\n")
                output.write(
                    "/*!50106 SET sql_mode              = 'ONLY_FULL_GROUP_BY,STRICT_TRANS_TABLES,NO_ZERO_IN_DATE,NO_ZERO_DATE,ERROR_FOR_DIVISION_BY_ZERO,NO_ENGINE_SUBSTITUTION' */ ;;\n")
                try:
                    cursor.execute(f"SHOW CREATE EVENT `{event}`")
                    result = cursor.fetchone()
                    if result and len(result) >= 4:
                        create_event = result[3]
                        output.write(f"/*!50106 {create_event} */ ;;\n")
                except Exception as e:
                    logger.warning(f"Error exporting event `{event}`: {e}")
                    output.write(f"-- Error exporting event `{event}`: {e}\n")
                output.write("/*!50106 SET sql_mode              = @saved_sql_mode */ ;;\n")
                output.write("/*!50106 SET character_set_client  = @saved_cs_client */ ;;\n")
                output.write("/*!50106 SET character_set_results = @saved_cs_results */ ;;\n")
                output.write("/*!50106 SET collation_connection  = @saved_col_connection */ ;;\n")
                output.write("/*!50106 SET TIME_ZONE= @save_time_zone */ ;;\n")
                output.write("DELIMITER ;\n\n")
            logger.info(f"Completed backing up {len(events)} events")

        # ── Footer ──────────────────────────────────────────────────
        output.write("/*!40103 SET TIME_ZONE=@OLD_TIME_ZONE */;\n")
        output.write("/*!40101 SET SQL_MODE=@OLD_SQL_MODE */;\n")
        output.write("/*!40014 SET FOREIGN_KEY_CHECKS=@OLD_FOREIGN_KEY_CHECKS */;\n")
        output.write("/*!40014 SET UNIQUE_CHECKS=@OLD_UNIQUE_CHECKS */;\n")
        output.write("/*!40101 SET CHARACTER_SET_CLIENT=@OLD_CHARACTER_SET_CLIENT */;\n")
        output.write("/*!40101 SET CHARACTER_SET_RESULTS=@OLD_CHARACTER_SET_RESULTS */;\n")
        output.write("/*!40101 SET COLLATION_CONNECTION=@OLD_COLLATION_CONNECTION */;\n")
        output.write("/*!40111 SET SQL_NOTES=@OLD_SQL_NOTES */;\n")
        output.write("/*!80000 SET SQL_REQUIRE_PRIMARY_KEY=@OLD_SQL_REQUIRE_PRIMARY_KEY */;\n\n")
        output.write(f"-- Dump completed on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        logger.info(f"Database backup completed successfully: {len(tables)} tables, {len(views)} views, "
                    f"{len(procedures)} procedures, {len(functions)} functions, {len(triggers)} triggers, "
                    f"{len(events)} events")

        # Build response
        sql_content = output.getvalue()
        output.close()

        response = make_response(sql_content)
        response.headers['Content-Type'] = 'application/sql; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        return response

    except Error as e:
        logger.error(f"Error generating database backup: {str(e)}")
        return jsonify({'error': f'Backup failed: {str(e)}'}), 500
    except Exception as e:
        logger.error(f"Unexpected error during database backup: {str(e)}", exc_info=True)
        return jsonify({'error': f'Backup failed: {str(e)}'}), 500
    finally:
        cursor.close()
        connection.close()


# ---------------------------------------------------------------------------
# Async DB backup → Appwrite upload (triggered by external cron)
# ---------------------------------------------------------------------------

def _run_backup_and_upload():
    """Generate a MySQL database backup and upload it to Appwrite.

    Uses the BackupService to create, compress, encrypt and upload backups.
    """
    backup_service = get_backup_service()
    return backup_service.create_and_upload_backup()


@app.route('/api/admin/trigger-backup', methods=['GET'])
def trigger_db_backup():
    """Trigger a database backup that uploads to Appwrite.

    Can run synchronously (for Vercel) or asynchronously (for local dev).
    Set BACKUP_MODE=sync in environment for Vercel/serverless platforms.
    """
    # Get allowed origins from environment variable
    backup_origins_env = os.environ.get('BACKUP_ALLOWED_ORIGINS', '')
    allowed_origins = [origin.strip() for origin in backup_origins_env.split(',') if origin.strip()]

    # Get local patterns from environment variable
    local_patterns_env = os.environ.get('BACKUP_LOCAL_PATTERNS', 'localhost,127.0.0.1')
    local_patterns = [pattern.strip() for pattern in local_patterns_env.split(',') if pattern.strip()]

    origin = request.headers.get('Origin', '')
    referer = request.headers.get('Referer', '')
    user_agent = request.headers.get('User-Agent', '')
    remote_addr = request.remote_addr or ''

    origin_ok = any(origin.startswith(a) for a in allowed_origins)
    referer_ok = any(referer.startswith(a) for a in allowed_origins)
    ua_ok = 'cron-job.org' in user_agent.lower()
    local_origin = any(p in origin for p in local_patterns) or origin.startswith(
        'http://localhost') or origin.startswith('http://127.0.0.1')
    local_referer = any(p in referer for p in local_patterns) or referer.startswith(
        'http://localhost') or referer.startswith('http://127.0.0.1')
    local_addr = any(remote_addr.startswith(p) for p in local_patterns)

    if not (origin_ok or referer_ok or ua_ok or local_origin or local_referer or local_addr):
        logger.warning("Unauthorized backup trigger — Origin: %s, Referer: %s, UA: %s, Remote: %s",
                       origin, referer, user_agent, remote_addr)
        return jsonify({'error': 'Access denied',
                        'message': 'This endpoint is only accessible from authorized sources'}), 403

    # Determine if we should run synchronously (for Vercel) or async (for local)
    backup_mode = os.environ.get('BACKUP_MODE', 'async').lower()

    if backup_mode == 'sync':
        # Run synchronously (Vercel-friendly)
        logger.info("Starting synchronous database backup")
        try:
            success, message = _run_backup_and_upload()
            if success:
                return jsonify({
                    'status': 'completed',
                    'message': message
                }), 200
            else:
                return jsonify({
                    'status': 'failed',
                    'error': message
                }), 500
        except Exception as e:
            logger.error("Backup failed with exception: %s", e, exc_info=True)
            return jsonify({
                'status': 'failed',
                'error': str(e)
            }), 500
    else:
        # Run asynchronously (local dev)
        def _async_backup_wrapper():
            try:
                success, message = _run_backup_and_upload()
                if success:
                    logger.info("Background backup completed: %s", message)
                else:
                    logger.error("Background backup failed: %s", message)
            except Exception as e:
                logger.error("Background backup exception: %s", e, exc_info=True)

        thread = threading.Thread(target=_async_backup_wrapper, daemon=True)
        thread.start()
        logger.info("Database backup triggered (background thread started)")

        return jsonify({
            'status': 'triggered',
            'message': 'Database backup started in background. It will be uploaded to Appwrite upon completion.',
        }), 202


@app.route('/api/admin/users/<int:user_id>/payment-methods', methods=['GET'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def admin_get_user_payment_methods(user_id):
    """Return active payment methods for a given user (admin only)."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("""
                       SELECT id, name, type, color
                       FROM payment_methods
                       WHERE user_id = %s AND is_active = TRUE
                       ORDER BY name
                       """, (user_id,))
        return jsonify(cursor.fetchall())
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/import-csv', methods=['POST'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def admin_import_csv():
    """Import transactions from a CSV file for a specific user/month/year.

    CSV format: Description, Credit, Debit, Note, Method
    Amounts may have a currency prefix (e.g. "Rs") and thousand-separators.

    Accepts an optional ``method_mapping`` JSON field (form-data string):
        { "CsvMethodName": <payment_method_id | "__create__"  | "__skip__"> , ... }

    - An integer id means "map to this existing payment method".
    - ``"__create__"`` means "create a new payment method with that name".
    - ``"__skip__"`` means "leave payment_method_id NULL for those rows".
    - If a CSV method name is not in the mapping it falls back to
      auto-matching by name (case-insensitive) or creating a new method.
    """
    # ── Validate inputs ──────────────────────────────────────────
    if 'file' not in request.files:
        return jsonify({'error': 'No CSV file provided'}), 400

    csv_file = request.files['file']
    if not csv_file.filename or not csv_file.filename.lower().endswith('.csv'):
        return jsonify({'error': 'File must be a .csv file'}), 400

    target_user_id = request.form.get('user_id', type=int)
    year = request.form.get('year', type=int)
    month = request.form.get('month', type=int)

    if not target_user_id or not year or not month:
        return jsonify({'error': 'user_id, year and month are required'}), 400

    if month < 1 or month > 12:
        return jsonify({'error': 'month must be between 1 and 12'}), 400

    # Optional method mapping from the frontend
    method_mapping_raw = request.form.get('method_mapping')
    method_mapping = {}
    if method_mapping_raw:
        try:
            method_mapping = json.loads(method_mapping_raw)
        except (json.JSONDecodeError, TypeError):
            return jsonify({'error': 'Invalid method_mapping JSON'}), 400

    # ── Parse CSV ────────────────────────────────────────────────
    import re

    def parse_amount(raw):
        """Strip currency prefix/symbols and thousand-separators, return Decimal or None."""
        if raw is None:
            return None
        raw = str(raw).strip()
        if not raw:
            return None
        # Remove common currency prefixes (Rs, Rs., LKR, $, etc.) and spaces
        raw = re.sub(r'^[A-Za-z$.\s]+', '', raw)
        # Remove thousand-separators
        raw = raw.replace(',', '')
        if not raw:
            return None
        try:
            val = Decimal(raw)
            return val if val > 0 else None
        except Exception:
            return None

    try:
        stream = io.StringIO(csv_file.read().decode('utf-8-sig'))
        reader = csv.DictReader(stream)

        # Normalise header names (strip whitespace, lowercase)
        if reader.fieldnames:
            reader.fieldnames = [f.strip() for f in reader.fieldnames]

        rows = []
        for line_no, row in enumerate(reader, start=2):
            # Normalise keys
            row = {k.strip().lower(): v for k, v in row.items() if k}
            description = (row.get('description') or '').strip()
            if not description:
                continue  # skip blank rows

            debit = parse_amount(row.get('debit'))
            credit = parse_amount(row.get('credit'))
            note = (row.get('note') or row.get('notes') or '').strip()
            method = (row.get('method') or row.get('payment method') or '').strip()

            rows.append({
                'description': description,
                'debit': debit,
                'credit': credit,
                'note': note,
                'method': method,
            })

        if not rows:
            return jsonify({'error': 'CSV file contains no valid rows'}), 400

    except UnicodeDecodeError:
        return jsonify({'error': 'File encoding not supported. Please use UTF-8.'}), 400
    except Exception as e:
        logger.error(f"CSV parse error: {e}")
        return jsonify({'error': f'Failed to parse CSV: {str(e)}'}), 400

    # ── Database operations ──────────────────────────────────────
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    admin_id = session['user_id']

    try:
        # Verify target user exists
        cursor.execute("SELECT id, username FROM users WHERE id = %s", (target_user_id,))
        target_user = cursor.fetchone()
        if not target_user:
            return jsonify({'error': 'Target user not found'}), 404

        # Get or create monthly record
        month_name = calendar.month_name[month]
        cursor.execute("""
                       INSERT INTO monthly_records (user_id, year, month, month_name)
                       VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                       UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                       """, (target_user_id, year, month, month_name))
        monthly_record = {'id': cursor.lastrowid}

        # Pre-load existing payment methods for the target user
        cursor.execute("""
                       SELECT id, name FROM payment_methods
                       WHERE user_id = %s AND is_active = TRUE
                       """, (target_user_id,))
        existing_methods = {pm['name'].lower(): pm['id'] for pm in cursor.fetchall()}

        # Get current max display_order for appending at the end
        cursor.execute("""
                       SELECT COALESCE(MAX(display_order), 0) AS max_order
                       FROM transactions WHERE monthly_record_id = %s
                       """, (monthly_record['id'],))
        current_max_order = cursor.fetchone()['max_order']

        transaction_date = datetime(year, month, 1).date()

        imported_count = 0
        skipped_count = 0
        methods_created = []

        # ── Phase 1: resolve payment methods for ALL rows first ──
        # This avoids interleaving method-creation queries with
        # transaction inserts and reduces round-trips.
        resolved_pm_ids = []  # parallel list — one entry per row
        for row in rows:
            payment_method_id = None
            if row['method']:
                mapping_value = method_mapping.get(row['method'])

                if mapping_value == '__skip__':
                    payment_method_id = None
                elif mapping_value == '__create__':
                    method_key = row['method'].lower()
                    if method_key in existing_methods:
                        payment_method_id = existing_methods[method_key]
                    else:
                        cursor.execute("""
                                       INSERT INTO payment_methods (user_id, name, type, color)
                                       VALUES (%s, %s, %s, %s)
                                       """, (target_user_id, row['method'], 'other', '#6c757d'))
                        payment_method_id = cursor.lastrowid
                        existing_methods[method_key] = payment_method_id
                        methods_created.append(row['method'])
                elif mapping_value is not None:
                    try:
                        payment_method_id = int(mapping_value)
                    except (ValueError, TypeError):
                        payment_method_id = None
                else:
                    method_key = row['method'].lower()
                    if method_key in existing_methods:
                        payment_method_id = existing_methods[method_key]
                    else:
                        cursor.execute("""
                                       INSERT INTO payment_methods (user_id, name, type, color)
                                       VALUES (%s, %s, %s, %s)
                                       """, (target_user_id, row['method'], 'other', '#6c757d'))
                        payment_method_id = cursor.lastrowid
                        existing_methods[method_key] = payment_method_id
                        methods_created.append(row['method'])
            resolved_pm_ids.append(payment_method_id)

        # ── Phase 2: batch-insert all transactions in one query ──
        txn_values = []
        txn_params = []
        for idx, row in enumerate(rows):
            current_max_order += 1
            txn_values.append("(%s, %s, %s, %s, %s, %s, %s, %s, %s, TRUE, TRUE, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)")
            txn_params.extend([
                monthly_record['id'],
                row['description'],
                None,
                row['debit'],
                row['credit'],
                transaction_date,
                row['note'] if row['note'] else None,
                resolved_pm_ids[idx],
                current_max_order,
            ])

        if txn_values:
            cursor.execute(
                "INSERT INTO transactions "
                "(monthly_record_id, description, category_id, debit, credit, "
                "transaction_date, notes, payment_method_id, display_order, "
                "is_done, is_paid, marked_done_at, paid_at) VALUES "
                + ", ".join(txn_values),
                txn_params,
            )
            first_txn_id = cursor.lastrowid
            imported_count = len(txn_values)

            # ── Phase 3: batch-insert audit logs in one query ──
            ip_address = request.remote_addr if request else None
            user_agent = request.headers.get('User-Agent') if request else None
            audit_note = f"Imported for user {target_user['username']}"

            audit_values = []
            audit_params = []
            for i in range(imported_count):
                audit_values.append("(%s, %s, %s, %s, %s, %s, %s)")
                audit_params.extend([
                    first_txn_id + i,
                    admin_id,
                    'CREATE',
                    'csv_import',
                    audit_note,
                    ip_address,
                    user_agent,
                ])

            try:
                cursor.execute(
                    "INSERT INTO transaction_audit_logs "
                    "(transaction_id, user_id, action, field_name, new_value, ip_address, user_agent) VALUES "
                    + ", ".join(audit_values),
                    audit_params,
                )
            except Exception as e:
                logger.error(f"Failed to create batch audit logs: {e}")

        connection.commit()

        # Audit log
        details = (f"Imported {imported_count} transactions from CSV "
                   f"for {month_name} {year} (user: {target_user['username']})")
        if methods_created:
            details += f". Created payment methods: {', '.join(methods_created)}"
        log_audit(admin_id, 'CSV_IMPORT', target_user_id=target_user_id, details=details)

        return jsonify({
            'message': f'Successfully imported {imported_count} transactions for {month_name} {year}',
            'imported': imported_count,
            'skipped': skipped_count,
            'methods_created': methods_created,
        }), 201

    except Error as e:
        connection.rollback()
        logger.error(f"CSV import DB error: {e}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/users/<int:user_id>/monthly-records', methods=['GET'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def admin_get_user_monthly_records(user_id):
    """Return monthly records for a given user with transaction counts (admin only)."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    try:
        cursor.execute("""
                       SELECT mr.id, mr.year, mr.month, mr.month_name,
                              COUNT(t.id) AS transaction_count,
                              COALESCE(SUM(t.debit), 0) AS total_debit,
                              COALESCE(SUM(t.credit), 0) AS total_credit
                       FROM monthly_records mr
                       LEFT JOIN transactions t ON mr.id = t.monthly_record_id
                       WHERE mr.user_id = %s
                       GROUP BY mr.id, mr.year, mr.month, mr.month_name
                       ORDER BY mr.year DESC, mr.month DESC
                       """, (user_id,))
        return jsonify(cursor.fetchall())
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/admin/users/<int:user_id>/monthly-records/<int:record_id>', methods=['DELETE'])
@admin_required
@limiter.limit(RATE_LIMIT_ADMIN)
def admin_delete_monthly_record(user_id, record_id):
    """Delete a monthly record and all its transactions for a given user (admin only)."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    admin_id = session['user_id']

    try:
        # Verify the record belongs to the specified user
        cursor.execute("""
                       SELECT mr.id, mr.year, mr.month, mr.month_name, u.username
                       FROM monthly_records mr
                       JOIN users u ON mr.user_id = u.id
                       WHERE mr.id = %s AND mr.user_id = %s
                       """, (record_id, user_id))
        record = cursor.fetchone()
        if not record:
            return jsonify({'error': 'Monthly record not found for this user'}), 404

        # Count transactions that will be deleted
        cursor.execute("SELECT COUNT(*) AS cnt FROM transactions WHERE monthly_record_id = %s", (record_id,))
        txn_count = cursor.fetchone()['cnt']

        # Get all attachments from transactions in this monthly record before deletion
        cursor.execute("""
            SELECT attachments 
            FROM transactions 
            WHERE monthly_record_id = %s AND attachments IS NOT NULL
        """, (record_id,))
        transactions_with_attachments = cursor.fetchall()

        # Delete all attachments from Appwrite before deleting transactions
        if appwrite_file_service.is_available() and transactions_with_attachments:
            for txn in transactions_with_attachments:
                attachments_value = txn['attachments']
                # Split comma-separated GUIDs
                attachment_guids = [guid.strip() for guid in attachments_value.split(',') if guid.strip()]

                for attachment_guid in attachment_guids:
                    success, error = appwrite_file_service.delete_file(attachment_guid)
                    if success:
                        logger.info(f"Deleted attachment {attachment_guid} from Appwrite (monthly record cleanup)")
                    else:
                        logger.warning(f"Failed to delete attachment {attachment_guid}: {error}")

        # Delete the monthly record (CASCADE will remove transactions)
        cursor.execute("DELETE FROM monthly_records WHERE id = %s", (record_id,))
        connection.commit()

        details = (f"Deleted {record['month_name']} {record['year']} "
                   f"({txn_count} transactions) for user {record['username']}")
        log_audit(admin_id, 'DELETE_MONTHLY_RECORD', target_user_id=user_id, details=details)

        return jsonify({
            'message': f"Deleted {record['month_name']} {record['year']} ({txn_count} transactions)",
            'deleted_transactions': txn_count,
        }), 200

    except Error as e:
        connection.rollback()
        logger.error(f"Delete monthly record error: {e}")
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/dashboard-stats')
@login_required
@limiter.limit(RATE_LIMIT_API)
def dashboard_stats():
    """Get dashboard statistics."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            user_id = session['user_id']

            # Get current month stats
            current_year = datetime.now().year
            current_month = datetime.now().month

            # Fetch current-month stats, YTD stats, and top categories
            # in a single round-trip using UNION ALL.
            cursor.execute("""
                           SELECT 'current' AS _section,
                                  SUM(debit) AS total_income,
                                  SUM(credit) AS total_expenses,
                                  NULL AS ytd_income,
                                  NULL AS ytd_expenses,
                                  NULL AS category,
                                  NULL AS amount,
                                  NULL AS cat_type
                           FROM transactions t
                                    JOIN monthly_records mr ON t.monthly_record_id = mr.id
                           WHERE mr.user_id = %s AND mr.year = %s AND mr.month = %s

                           UNION ALL

                           SELECT 'ytd', NULL, NULL,
                                  SUM(debit), SUM(credit),
                                  NULL, NULL, NULL
                           FROM transactions t
                                    JOIN monthly_records mr ON t.monthly_record_id = mr.id
                           WHERE mr.user_id = %s AND mr.year = %s

                           UNION ALL

                           SELECT 'income_cat', NULL, NULL, NULL, NULL,
                                  c.name, SUM(t.debit), NULL
                           FROM transactions t
                                    JOIN monthly_records mr ON t.monthly_record_id = mr.id
                                    LEFT JOIN categories c ON t.category_id = c.id
                           WHERE mr.user_id = %s AND mr.year = %s AND mr.month = %s AND t.debit > 0
                           GROUP BY c.name
                           ORDER BY amount DESC LIMIT 5
                           """, (user_id, current_year, current_month,
                                 user_id, current_year,
                                 user_id, current_year, current_month))

            # Parse the UNION ALL result set
            current_stats = {'total_income': 0, 'total_expenses': 0}
            ytd_stats = {'ytd_income': 0, 'ytd_expenses': 0}
            income_categories = []

            for row in cursor.fetchall():
                section = row['_section']
                if section == 'current':
                    current_stats = {
                        'total_income': row['total_income'] or 0,
                        'total_expenses': row['total_expenses'] or 0,
                    }
                elif section == 'ytd':
                    ytd_stats = {
                        'ytd_income': row['ytd_income'] or 0,
                        'ytd_expenses': row['ytd_expenses'] or 0,
                    }
                elif section == 'income_cat':
                    income_categories.append({'category': row['category'], 'amount': row['amount']})

            current_stats['current_balance'] = (current_stats['total_income'] or 0) - (
                    current_stats['total_expenses'] or 0)

            # Second query: expense categories, monthly trend, and recent
            # transactions (these need independent ORDER BY / LIMIT clauses).
            cursor.execute("""
                           SELECT c.name        as category,
                                  SUM(t.credit) as amount
                           FROM transactions t
                                    JOIN monthly_records mr ON t.monthly_record_id = mr.id
                                    LEFT JOIN categories c ON t.category_id = c.id
                           WHERE mr.user_id = %s
                             AND mr.year = %s
                             AND mr.month = %s
                             AND t.credit > 0
                           GROUP BY c.name
                           ORDER BY amount DESC LIMIT 5
                           """, (user_id, current_year, current_month))
            expense_categories = cursor.fetchall()

            cursor.execute("""
                           SELECT mr.year,
                                  mr.month,
                                  mr.month_name,
                                  SUM(t.debit)  as income,
                                  SUM(t.credit) as expenses
                           FROM monthly_records mr
                                    LEFT JOIN transactions t ON mr.id = t.monthly_record_id
                           WHERE mr.user_id = %s
                           GROUP BY mr.year, mr.month, mr.month_name
                           ORDER BY mr.year DESC, mr.month DESC LIMIT 12
                           """, (user_id,))
            monthly_trend = cursor.fetchall()

            cursor.execute("""
                           SELECT t.description,
                                  t.debit,
                                  t.credit,
                                  t.transaction_date,
                                  c.name as category
                           FROM transactions t
                                    JOIN monthly_records mr ON t.monthly_record_id = mr.id
                                    LEFT JOIN categories c ON t.category_id = c.id
                           WHERE mr.user_id = %s
                           ORDER BY t.created_at DESC LIMIT 10
                           """, (user_id,))
            recent_transactions = cursor.fetchall()

            return jsonify({
                'current_stats': current_stats,
                'ytd_stats': ytd_stats,
                'recent_transactions': recent_transactions,
                'monthly_trend': monthly_trend,
                'income_categories': income_categories,
                'expense_categories': expense_categories
            })

        except Error as e:
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/transactions', methods=['GET', 'POST'])
@login_required
@limiter.limit(RATE_LIMIT_API)
def transactions():
    """Get or create transactions."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        if request.method == 'GET':
            # Get query parameters
            year = request.args.get('year', datetime.now().year, type=int)
            month = request.args.get('month', datetime.now().month, type=int)
            search_all = request.args.get('searchAll', 'false').lower() == 'true'

            # Get filter parameters
            description = request.args.get('description', '')
            notes_filter = request.args.get('notes', '')
            categories = request.args.get('categories', '')  # comma-separated IDs
            payment_methods = request.args.get('paymentMethods', '')  # comma-separated IDs
            types = request.args.get('types', '')  # comma-separated: income,expense
            statuses = request.args.get('statuses', '')  # comma-separated: done,not_done,paid,unpaid
            min_amount = request.args.get('minAmount', type=float)
            max_amount = request.args.get('maxAmount', type=float)
            start_date = request.args.get('startDate', '')
            end_date = request.args.get('endDate', '')

            # Check if any filters are active
            has_filters = any([
                description, notes_filter, categories, payment_methods,
                types, statuses, min_amount is not None, max_amount is not None,
                start_date, end_date
            ])

            # Build dynamic WHERE clause
            where_clauses = []
            params = []

            # Use a JOIN on monthly_records instead of a separate
            # query to fetch IDs first (eliminates one round-trip).
            # The mr.user_id filter is always applied via the JOIN.
            where_clauses.append("mr.user_id = %s")
            params.append(user_id)

            if search_all or has_filters:
                # Parse date range to extract year and month if provided
                if start_date:
                    try:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        where_clauses.append("(mr.year > %s OR (mr.year = %s AND mr.month >= %s))")
                        params.extend([start_dt.year, start_dt.year, start_dt.month])
                    except ValueError:
                        pass

                if end_date:
                    try:
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        where_clauses.append("(mr.year < %s OR (mr.year = %s AND mr.month <= %s))")
                        params.extend([end_dt.year, end_dt.year, end_dt.month])
                    except ValueError:
                        pass
            else:
                # Normal behavior - limit to specific month
                where_clauses.append("mr.year = %s")
                where_clauses.append("mr.month = %s")
                params.extend([year, month])

            # Continue with filter building only if we have WHERE clauses
            if where_clauses:

                # Description filter
                if description:
                    where_clauses.append("LOWER(t.description) LIKE %s")
                    params.append(f"%{description.lower()}%")

                # Notes filter
                if notes_filter:
                    where_clauses.append("LOWER(t.notes) LIKE %s")
                    params.append(f"%{notes_filter.lower()}%")

                # Category filter
                if categories:
                    cat_ids = [int(cid) for cid in categories.split(',') if cid.strip()]
                    if cat_ids:
                        placeholders = ','.join(['%s'] * len(cat_ids))
                        where_clauses.append(f"t.category_id IN ({placeholders})")
                        params.extend(cat_ids)

                # Payment method filter
                if payment_methods:
                    pm_ids = [int(pmid) for pmid in payment_methods.split(',') if pmid.strip()]
                    if pm_ids:
                        placeholders = ','.join(['%s'] * len(pm_ids))
                        where_clauses.append(f"t.payment_method_id IN ({placeholders})")
                        params.extend(pm_ids)

                # Transaction type filter
                if types:
                    type_list = [t.strip() for t in types.split(',') if t.strip()]
                    type_conditions = []
                    if 'income' in type_list:
                        type_conditions.append("t.debit > 0")
                    if 'expense' in type_list:
                        type_conditions.append("t.credit > 0")
                    if type_conditions:
                        where_clauses.append(f"({' OR '.join(type_conditions)})")

                # Status filter
                if statuses:
                    status_list = [s.strip() for s in statuses.split(',') if s.strip()]
                    status_conditions = []
                    if 'done' in status_list:
                        status_conditions.append("t.is_done = TRUE")
                    if 'not_done' in status_list:
                        status_conditions.append("t.is_done = FALSE OR t.is_done IS NULL")
                    if 'paid' in status_list:
                        status_conditions.append("t.is_paid = TRUE")
                    if 'unpaid' in status_list:
                        status_conditions.append("t.is_paid = FALSE OR t.is_paid IS NULL")
                    if status_conditions:
                        where_clauses.append(f"({' OR '.join(status_conditions)})")

                # Amount range filter
                if min_amount is not None:
                    where_clauses.append("(COALESCE(t.debit, 0) >= %s OR COALESCE(t.credit, 0) >= %s)")
                    params.extend([min_amount, min_amount])

                if max_amount is not None:
                    where_clauses.append("(COALESCE(t.debit, 0) <= %s OR COALESCE(t.credit, 0) <= %s)")
                    params.extend([max_amount, max_amount])

                # Date range filter is now handled via monthly_records filtering above
                # No need to filter on transaction dates here

                # Combine WHERE clauses
                where_sql = " AND ".join(where_clauses)

                query = f"""
                    SELECT
                        t.*,
                        c.name as category_name,
                        pm.name as payment_method_name,
                        pm.type as payment_method_type,
                        pm.color as payment_method_color,
                        COALESCE(t.is_paid, FALSE) as is_paid
                    FROM transactions t
                    INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                    LEFT JOIN categories c ON t.category_id = c.id
                    LEFT JOIN payment_methods pm ON t.payment_method_id = pm.id
                    WHERE {where_sql}
                    ORDER BY t.display_order ASC, t.id ASC
                """

                cursor.execute(query, params)
                transactions = cursor.fetchall()
                return jsonify(transactions)
            else:
                return jsonify([])

        else:  # POST - Create new transaction
            # Check if request is multipart/form-data (with image) or JSON (without image)
            is_multipart = request.content_type and 'multipart/form-data' in request.content_type

            if is_multipart:
                # Get form data
                data = request.form.to_dict()
                # Convert string numbers to proper types
                for key in ['debit', 'credit', 'category_id', 'year', 'month', 'payment_method_id']:
                    if key in data and data[key]:
                        try:
                            if key in ['debit', 'credit']:
                                data[key] = float(data[key]) if data[key] else None
                            else:
                                data[key] = int(data[key]) if data[key] else None
                        except (ValueError, TypeError):
                            data[key] = None
            else:
                data = request.get_json()

            print(f"[DEBUG] Received transaction data: {data}")

            # Handle bill image(s) upload to Appwrite (if provided)
            attachment_guids = []
            attachments_value = None

            # Check if attachments were already uploaded (new sequential upload flow)
            if 'attachments' in data and data['attachments']:
                # Images already uploaded sequentially, use provided GUIDs
                logger.info(f"Using pre-uploaded attachments: {data['attachments']}")
                attachments_value = data['attachments']
            else:
                # Legacy flow: handle multipart file uploads
                # Check for multiple images first, then fall back to single image (backward compatible)
                bill_images = []
                if is_multipart:
                    if 'bill_images' in request.files:
                        bill_images = request.files.getlist('bill_images')
                    elif 'bill_image' in request.files:
                        bill_images = [request.files['bill_image']]

                # Process each image
                for idx, bill_image in enumerate(bill_images):
                    if not bill_image or not bill_image.filename:
                        continue

                    # Validate file type
                    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}
                    file_ext = bill_image.filename.rsplit('.', 1)[-1].lower() if '.' in bill_image.filename else ''

                    if file_ext not in allowed_extensions:
                        logger.warning(f"Skipping file {idx + 1} with invalid extension: {file_ext}")
                        continue

                    # Read file data
                    file_data = bill_image.read()

                    if len(file_data) == 0:
                        logger.warning(f"Skipping empty file {idx + 1}")
                        continue

                    if appwrite_file_service.is_available():
                        try:
                            # Optimize file if needed (resize images, compress PDFs)
                            optimized_data, was_optimized, original_size, final_size = optimize_file_for_upload(
                                file_data,
                                file_ext,
                                bill_image.filename
                            )

                            # Use optimized data for upload
                            image_data = optimized_data

                            # Log optimization results
                            if was_optimized:
                                reduction_pct = ((original_size - final_size) / original_size) * 100
                                logger.info(
                                    f"✓ Image {idx + 1}/{len(bill_images)} optimized: {original_size / (1024 * 1024):.2f}MB → {final_size / (1024 * 1024):.2f}MB ({reduction_pct:.1f}% smaller)")
                            else:
                                logger.info(
                                    f"Image {idx + 1}/{len(bill_images)} size OK, no optimization needed: {original_size / (1024 * 1024):.2f}MB")

                            # Generate pure GUID for filename
                            attachment_guid = str(uuid.uuid4())
                            filename = f"{attachment_guid}.{file_ext}"

                            # Check file size
                            file_size_mb = len(image_data) / (1024 * 1024)
                            logger.info(
                                f"Uploading image {idx + 1}/{len(bill_images)} to Appwrite: {filename}, type: {file_ext}, size: {len(image_data)} bytes ({file_size_mb:.2f}MB)")

                            # Check first bytes for validation
                            first_bytes = image_data[:8] if len(image_data) >= 8 else image_data
                            logger.info(f"First 8 bytes: {first_bytes} (hex: {first_bytes.hex()})")

                            # Validate PDF header if uploading PDF
                            if file_ext == 'pdf':
                                if not first_bytes.startswith(b'%PDF'):
                                    logger.error(
                                        f"WARNING: PDF file does NOT start with %PDF header before upload!")
                                    logger.error(f"First 20 bytes: {image_data[:20]}")
                                else:
                                    logger.info(f"✓ PDF header valid before upload")

                            # Warn if file is still unusually large after optimization
                            if file_size_mb > 10:
                                logger.warning(
                                    f"Large file upload: {file_size_mb:.2f}MB - may take time to process")

                            # Upload to Appwrite using the file service
                            success, error, result = appwrite_file_service.upload_file(
                                image_data,
                                attachment_guid,
                                filename
                            )

                            if success:
                                stored_size = result.get('sizeOriginal', 'N/A') if result else 'N/A'
                                logger.info(
                                    f"✓ Image {idx + 1}/{len(bill_images)} uploaded successfully: {attachment_guid}, stored size: {stored_size}")
                                # Add to list of uploaded GUIDs
                                attachment_guids.append(attachment_guid)
                            else:
                                logger.error(f"Failed to upload bill image {idx + 1} to Appwrite: {error}")
                                # Continue with other images even if one fails
                        except Exception as e:
                            logger.error(f"Failed to process and upload bill image {idx + 1}: {str(e)}")
                            # Continue with other images even if one fails
                    else:
                        logger.warning(f"Appwrite storage not configured for image {idx + 1}")

                # Store comma-separated GUIDs in attachments field (or None if no uploads succeeded)
                attachments_value = ','.join(attachment_guids) if attachment_guids else None
                if attachments_value:
                    logger.info(f"Stored {len(attachment_guids)} attachment(s): {attachments_value}")

            # Get or create monthly record
            year = data.get('year', datetime.now().year)
            month = data.get('month', datetime.now().month)
            month_name = calendar.month_name[month]

            cursor.execute("""
                           INSERT INTO monthly_records (user_id, year, month, month_name)
                           VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                           UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                           """, (user_id, year, month, month_name))

            monthly_record = {'id': cursor.lastrowid}

            # Convert to Decimal to avoid float/Decimal arithmetic errors
            debit_value = data.get('debit')
            credit_value = data.get('credit')

            debit = Decimal(str(debit_value)) if debit_value else Decimal('0')
            credit = Decimal(str(credit_value)) if credit_value else Decimal('0')

            # Auto-categorize if no category provided (same logic as token endpoint)
            category_id = data.get('category_id') or None
            if not category_id:
                category_id = auto_categorize_transaction(data.get('description'))

            # Apply category percentage markup if applicable
            if category_id:
                if credit > 0:
                    credit = apply_category_percentage_markup(credit, category_id, connection)
                elif debit > 0:
                    debit = apply_category_percentage_markup(debit, category_id, connection)

            print(f"[DEBUG] Debit: {debit}, Credit: {credit}")

            # Use current date if no transaction_date provided
            transaction_date = data.get('transaction_date')
            if not transaction_date:
                transaction_date = datetime.now().date()

            # Push all existing transactions down by incrementing their display_order
            # This makes room for the new transaction at position 1 (top)
            cursor.execute("""
                           UPDATE transactions
                           SET display_order = display_order + 1
                           WHERE monthly_record_id = %s
                           """, (monthly_record['id'],))

            # New transaction gets display_order = 1 (appears at top)
            next_display_order = 1

            # Auto-categorize if no category provided (same logic as token endpoint)
            category_id = data.get('category_id') or None
            if not category_id:
                category_id = auto_categorize_transaction(data.get('description'))

            # Get bill content if provided (from scanned bills)
            bill_content = data.get('bill_content')

            # Insert transaction with attachments field
            insert_values = (
                monthly_record['id'],
                data.get('description'),
                category_id,
                debit if debit > 0 else None,
                credit if credit > 0 else None,
                transaction_date,
                data.get('notes'),
                next_display_order,
                bill_content,
                attachments_value,  # Store comma-separated GUIDs in attachments column
                data.get('payment_method_id')  # Add payment method
            )
            print(f"[DEBUG] Inserting transaction with values: {insert_values}")

            cursor.execute("""
                           INSERT INTO transactions
                           (monthly_record_id, description, category_id, debit, credit, transaction_date, notes,
                            display_order, bill_content, attachments, payment_method_id)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                           """, insert_values)

            transaction_id = cursor.lastrowid
            print(f"[DEBUG] Transaction inserted with ID: {transaction_id}")

            connection.commit()
            print(f"[DEBUG] Transaction committed successfully")

            response = {'message': 'Transaction created successfully', 'id': transaction_id}
            if attachments_value:
                response['attachments'] = attachments_value

            return jsonify(response), 201

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/filter', methods=['GET'])
@login_required
def filter_transactions():
    """Filter transactions across all data with advanced criteria."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session.get('user_id')

        # Get filter parameters
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        category_id = request.args.get('category')
        payment_method_id = request.args.get('paymentMethod')
        amount_min = request.args.get('amountMin')
        amount_max = request.args.get('amountMax')
        transaction_type = request.args.get('transactionType')  # 'debit' or 'credit'
        search_text = request.args.get('searchText')
        done_status = request.args.get('doneStatus')  # 'done', 'not_done', or empty
        paid_status = request.args.get('paidStatus')  # 'paid', 'not_paid', or empty

        # Build SQL query with filters
        query = """
                SELECT t.*,
                       c.name   as category_name,
                       pm.name  as payment_method_name,
                       pm.type  as payment_method_type,
                       pm.color as payment_method_color,
                       mr.year,
                       mr.month,
                       mr.month_name
                FROM transactions t
                         LEFT JOIN categories c ON t.category_id = c.id
                         LEFT JOIN payment_methods pm ON t.payment_method_id = pm.id
                         INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                WHERE mr.user_id = %s \
                """

        params = [user_id]

        # Add date range filter (based on monthly_records year/month)
        if date_from:
            try:
                start_dt = datetime.strptime(date_from, '%Y-%m-%d')
                start_year = start_dt.year
                start_month = start_dt.month
                query += " AND (mr.year > %s OR (mr.year = %s AND mr.month >= %s))"
                params.extend([start_year, start_year, start_month])
            except ValueError:
                pass

        if date_to:
            try:
                end_dt = datetime.strptime(date_to, '%Y-%m-%d')
                end_year = end_dt.year
                end_month = end_dt.month
                query += " AND (mr.year < %s OR (mr.year = %s AND mr.month <= %s))"
                params.extend([end_year, end_year, end_month])
            except ValueError:
                pass

        # Add category filter
        if category_id:
            query += " AND t.category_id = %s"
            params.append(int(category_id))

        # Add payment method filter
        if payment_method_id:
            query += " AND t.payment_method_id = %s"
            params.append(int(payment_method_id))

        # Add amount filter based on transaction type
        if transaction_type == 'debit':
            # Income transactions (debit > 0)
            query += " AND t.debit > 0"
            if amount_min:
                query += " AND t.debit >= %s"
                params.append(float(amount_min))
            if amount_max:
                query += " AND t.debit <= %s"
                params.append(float(amount_max))
        elif transaction_type == 'credit':
            # Expense transactions (credit > 0)
            query += " AND t.credit > 0"
            if amount_min:
                query += " AND t.credit >= %s"
                params.append(float(amount_min))
            if amount_max:
                query += " AND t.credit <= %s"
                params.append(float(amount_max))
        else:
            # Both types - check either debit or credit
            if amount_min or amount_max:
                amount_conditions = []
                if amount_min:
                    amount_conditions.append("(t.debit >= %s OR t.credit >= %s)")
                    params.extend([float(amount_min), float(amount_min)])
                if amount_max:
                    amount_conditions.append("(t.debit <= %s OR t.credit <= %s)")
                    params.extend([float(amount_max), float(amount_max)])
                if amount_conditions:
                    query += " AND " + " AND ".join(amount_conditions)

        # Add text search filter (search in description and notes)
        if search_text:
            query += " AND (t.description LIKE %s OR t.notes LIKE %s)"
            search_pattern = f"%{search_text}%"
            params.extend([search_pattern, search_pattern])

        # Add done status filter
        if done_status == 'done':
            query += " AND t.is_done = 1"
        elif done_status == 'not_done':
            query += " AND t.is_done = 0"

        # Add paid status filter
        if paid_status == 'paid':
            query += " AND t.is_paid = 1"
        elif paid_status == 'not_paid':
            query += " AND t.is_paid = 0"

        # Order by date descending (most recent first)
        query += " ORDER BY t.transaction_date DESC, t.id DESC"

        # Limit results to prevent overload (max 500 transactions)
        query += " LIMIT 500"

        cursor.execute(query, params)
        transactions = cursor.fetchall()

        # Calculate running balance for filtered transactions
        running_balance = 0
        for trans in reversed(transactions):
            running_balance += (trans['debit'] or 0) - (trans['credit'] or 0)
            trans['balance'] = running_balance

        # Reverse back to show most recent first
        transactions.reverse()

        return jsonify(transactions)

    except Exception as e:
        print(f"Error filtering transactions: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>', methods=['PUT', 'DELETE'])
@login_required
def manage_transaction(transaction_id):
    """Update or delete a transaction."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        if request.method == 'PUT':
            data = request.get_json()

            # Validate request data
            if not data:
                return jsonify({'error': 'No data provided'}), 400

            print(f"[DEBUG] Updating transaction {transaction_id} with data: {data}")

            # Get the current transaction data for audit trail
            dict_cursor = connection.cursor(dictionary=True)
            dict_cursor.execute("""
                                SELECT t.*
                                FROM transactions t
                                WHERE t.id = %s
                                  AND t.monthly_record_id IN
                                      (SELECT id FROM monthly_records WHERE user_id = %s)
                                """, (transaction_id, session['user_id']))

            old_transaction = dict_cursor.fetchone()
            dict_cursor.close()

            if not old_transaction:
                print(f"[DEBUG] Transaction {transaction_id} not found for user {session['user_id']}")
                return jsonify({'error': 'Transaction not found'}), 404

            monthly_record_id = old_transaction['monthly_record_id']
            print(f"[DEBUG] Found monthly_record_id: {monthly_record_id}")

            # Convert to Decimal to avoid float/Decimal arithmetic errors
            debit_value = data.get('debit')
            credit_value = data.get('credit')

            debit = Decimal(str(debit_value)) if debit_value else Decimal('0')
            credit = Decimal(str(credit_value)) if credit_value else Decimal('0')

            # Get category_id
            category_id = data.get('category_id')

            # Apply category percentage markup if applicable
            if category_id:
                if credit > 0:
                    credit = apply_category_percentage_markup(credit, category_id, connection)
                elif debit > 0:
                    debit = apply_category_percentage_markup(debit, category_id, connection)

            print(f"[DEBUG] Debit: {debit}, Credit: {credit}")

            # Handle transaction_date - use current date if not provided or empty
            transaction_date = data.get('transaction_date')
            if not transaction_date or transaction_date == '':
                transaction_date = datetime.now().date()

            # Handle payment method and is_done logic
            # If payment_method_id is provided (and not empty), set is_done to TRUE
            # If payment_method_id is None/empty, set is_done to FALSE
            payment_method_id = data.get('payment_method_id')
            if payment_method_id:
                is_done = True
            else:
                is_done = False
                payment_method_id = None

            # Update transaction (balance will be calculated on frontend)
            cursor.execute("""
                           UPDATE transactions
                           SET description      = %s,
                               category_id      = %s,
                               debit            = %s,
                               credit           = %s,
                               transaction_date = %s,
                               notes            = %s,
                               payment_method_id = %s,
                               is_done          = %s
                           WHERE id = %s
                           """, (
                data.get('description'),
                data.get('category_id'),
                debit if debit > 0 else None,
                credit if credit > 0 else None,
                transaction_date,
                data.get('notes'),
                payment_method_id,
                is_done,
                transaction_id
            ))

            print(f"[DEBUG] Transaction {transaction_id} updated successfully")

            # Log audit trail for each changed field
            user_id = session['user_id']

            # Track field changes
            new_debit = debit if debit > 0 else None
            new_credit = credit if credit > 0 else None

            # Normalize values for comparison to avoid false positives
            # Convert category_id to int for comparison (handle None)
            old_category = old_transaction['category_id']
            new_category = int(data.get('category_id')) if data.get('category_id') else None

            # Compare description
            if old_transaction['description'] != data.get('description'):
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'description',
                                      old_transaction['description'], data.get('description'))

            # Compare category_id (normalized)
            if old_category != new_category:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'category_id',
                                      old_category, new_category)

            # Compare debit (Decimal comparison)
            old_debit_normalized = old_transaction['debit'] if old_transaction['debit'] else None
            if old_debit_normalized != new_debit:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'debit',
                                      old_transaction['debit'], new_debit)

            # Compare credit (Decimal comparison)
            old_credit_normalized = old_transaction['credit'] if old_transaction['credit'] else None
            if old_credit_normalized != new_credit:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'credit',
                                      old_transaction['credit'], new_credit)

            # Compare transaction_date
            if str(old_transaction['transaction_date']) != str(transaction_date):
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'transaction_date',
                                      old_transaction['transaction_date'], transaction_date)

            # Compare notes (handle None/empty string)
            old_notes = old_transaction['notes'] if old_transaction['notes'] else None
            new_notes = data.get('notes') if data.get('notes') else None
            if old_notes != new_notes:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'notes',
                                      old_transaction['notes'], data.get('notes'))

            # Compare payment_method_id (handle None)
            old_payment_method = old_transaction['payment_method_id']
            new_payment_method = int(data.get('payment_method_id')) if data.get('payment_method_id') else None
            if old_payment_method != new_payment_method:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'payment_method_id',
                                      old_payment_method, new_payment_method)

            # Compare is_done status
            old_is_done = old_transaction.get('is_done', False)
            # Calculate new is_done based on payment method
            new_is_done = True if new_payment_method else False
            if old_is_done != new_is_done:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'is_done',
                                      old_is_done, new_is_done)

            connection.commit()
            print(f"[DEBUG] Transaction update committed successfully")
            return jsonify({'message': 'Transaction updated successfully'})

        else:  # DELETE
            # Log audit trail before deleting
            user_id = session['user_id']

            # Check if transaction has an attachment and delete it from Appwrite
            dict_cursor = connection.cursor(dictionary=True)
            dict_cursor.execute("""
                SELECT t.attachments
                FROM transactions t
                WHERE t.id = %s
                  AND t.monthly_record_id IN
                      (SELECT id FROM monthly_records WHERE user_id = %s)
            """, (transaction_id, user_id))

            transaction = dict_cursor.fetchone()
            dict_cursor.close()

            if transaction and transaction['attachments']:
                attachments_value = transaction['attachments']
                # Split comma-separated GUIDs (supports both single and multiple attachments)
                attachment_guids = [guid.strip() for guid in attachments_value.split(',') if guid.strip()]

                # Delete all attachments from Appwrite bucket
                if appwrite_file_service.is_available():
                    for attachment_guid in attachment_guids:
                        success, error = appwrite_file_service.delete_file(attachment_guid)
                        if success:
                            logger.info(
                                f"Deleted attachment {attachment_guid} from Appwrite for transaction {transaction_id}")
                        else:
                            logger.warning(f"Failed to delete attachment {attachment_guid}: {error}")
                            # Continue with other deletions even if one fails

            log_transaction_audit(cursor, transaction_id, user_id, 'DELETE')

            cursor.execute("""
                           DELETE
                           FROM transactions
                           WHERE id = %s
                             AND monthly_record_id IN
                                 (SELECT id FROM monthly_records WHERE user_id = %s)
                           """, (transaction_id, user_id))

            connection.commit()
            return jsonify({'message': 'Transaction deleted successfully'})

    except Error as e:
        print(f"[ERROR] Database error in manage_transaction: {str(e)}")
        connection.rollback()
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    except Exception as e:
        print(f"[ERROR] Unexpected error in manage_transaction: {str(e)}")
        connection.rollback()
        return jsonify({'error': f'Server error: {str(e)}'}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/audit-logs', methods=['GET'])
@login_required
def get_transaction_audit_logs(transaction_id):
    """Get audit logs for a specific transaction."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']

        # Verify the transaction belongs to the user
        cursor.execute("""
                       SELECT t.id
                       FROM transactions t
                                INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                       WHERE t.id = %s
                         AND mr.user_id = %s
                       """, (transaction_id, user_id))

        transaction = cursor.fetchone()
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404

        # Fetch audit logs for this transaction
        cursor.execute("""
                       SELECT tal.id,
                              tal.action,
                              tal.field_name,
                              tal.old_value,
                              tal.new_value,
                              tal.created_at,
                              u.username
                       FROM transaction_audit_logs tal
                                INNER JOIN users u ON tal.user_id = u.id
                       WHERE tal.transaction_id = %s
                       ORDER BY tal.created_at DESC
                       """, (transaction_id,))

        audit_logs = cursor.fetchall()
        return jsonify(audit_logs)

    except Error as e:
        logger.error(f"Error fetching audit logs: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/move', methods=['POST'])
@login_required
def move_transaction(transaction_id):
    """Move a transaction to a different month."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']
        data = request.get_json()

        target_year = data.get('target_year')
        target_month = data.get('target_month')

        if not target_year or not target_month:
            return jsonify({'error': 'Target year and month are required'}), 400

        # Verify the transaction belongs to the user
        cursor.execute("""
                       SELECT t.*, mr.year, mr.month
                       FROM transactions t
                                INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                       WHERE t.id = %s
                         AND mr.user_id = %s
                       """, (transaction_id, user_id))

        transaction = cursor.fetchone()
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404

        # Check if moving to the same month
        if transaction['year'] == target_year and transaction['month'] == target_month:
            return jsonify({'error': 'Transaction is already in this month'}), 400

        # Get or create target monthly record
        month_name = calendar.month_name[target_month]
        cursor.execute("""
                       INSERT INTO monthly_records (user_id, year, month, month_name)
                       VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                       UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                       """, (user_id, target_year, target_month, month_name))

        target_record_id = cursor.lastrowid

        # Update transaction's monthly_record_id and date
        new_date = datetime(target_year, target_month, 1).date()
        cursor.execute("""
                       UPDATE transactions
                       SET monthly_record_id = %s,
                           transaction_date  = %s
                       WHERE id = %s
                       """, (target_record_id, new_date, transaction_id))

        # Log audit trail
        log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'moved_to_month',
                              f"{transaction['year']}-{transaction['month']:02d}",
                              f"{target_year}-{target_month:02d}")

        connection.commit()

        return jsonify({
            'message': f'Transaction moved to {month_name} {target_year} successfully',
            'target_year': target_year,
            'target_month': target_month
        })

    except Error as e:
        logger.error(f"Error moving transaction: {e}")
        connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/copy', methods=['POST'])
@login_required
def copy_transaction(transaction_id):
    """Copy a transaction to a different month."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']
        data = request.get_json()

        target_year = data.get('target_year')
        target_month = data.get('target_month')

        if not target_year or not target_month:
            return jsonify({'error': 'Target year and month are required'}), 400

        # Verify the transaction belongs to the user and get its data
        cursor.execute("""
                       SELECT t.*
                       FROM transactions t
                                INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                       WHERE t.id = %s
                         AND mr.user_id = %s
                       """, (transaction_id, user_id))

        transaction = cursor.fetchone()
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404

        # Get or create target monthly record
        month_name = calendar.month_name[target_month]
        cursor.execute("""
                       INSERT INTO monthly_records (user_id, year, month, month_name)
                       VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                       UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                       """, (user_id, target_year, target_month, month_name))

        target_record_id = cursor.lastrowid

        # Push all existing transactions down in the target month
        cursor.execute("""
                       UPDATE transactions
                       SET display_order = display_order + 1
                       WHERE monthly_record_id = %s
                       """, (target_record_id,))

        # Create a copy of the transaction in the target month at position 1 (top)
        new_date = datetime(target_year, target_month, 1).date()
        debit = Decimal(str(transaction['debit'])) if transaction['debit'] else None
        credit = Decimal(str(transaction['credit'])) if transaction['credit'] else None

        cursor.execute("""
                       INSERT INTO transactions
                       (monthly_record_id, description, category_id, debit, credit,
                        transaction_date, notes, payment_method_id, display_order)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                       """, (
            target_record_id,
            transaction['description'],
            transaction['category_id'],
            debit,
            credit,
            new_date,
            transaction['notes'],
            transaction['payment_method_id'],
            1  # Display at top
        ))

        new_transaction_id = cursor.lastrowid

        # Log audit trail for the new transaction
        log_transaction_audit(cursor, new_transaction_id, user_id, 'CREATE')
        log_transaction_audit(cursor, new_transaction_id, user_id, 'UPDATE', 'copied_from_transaction',
                              None, str(transaction_id))

        connection.commit()

        return jsonify({
            'message': f'Transaction copied to {month_name} {target_year} successfully',
            'new_transaction_id': new_transaction_id,
            'target_year': target_year,
            'target_month': target_month
        })

    except Error as e:
        logger.error(f"Error copying transaction: {e}")
        connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/attachment', methods=['GET', 'DELETE'])
@login_required
def manage_transaction_attachment(transaction_id):
    """Get or delete a transaction's attachment."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']

        # Verify transaction belongs to user
        cursor.execute("""
            SELECT t.attachments, t.monthly_record_id
            FROM transactions t
            INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
            WHERE t.id = %s AND mr.user_id = %s
        """, (transaction_id, user_id))

        transaction = cursor.fetchone()

        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404

        attachments_value = transaction['attachments']

        if not attachments_value:
            return jsonify({'error': 'No attachment found for this transaction'}), 404

        # Split comma-separated GUIDs (supports both single and multiple attachments)
        attachment_guids = [guid.strip() for guid in attachments_value.split(',') if guid.strip()]

        if request.method == 'GET':
            # Return attachment info(s) with DIRECT Appwrite URLs (no proxy)
            # This avoids Vercel's 4.5MB response body limit
            if appwrite_file_service.is_available():
                appwrite_endpoint = os.environ.get('APPWRITE_ENDPOINT')
                appwrite_project_id = os.environ.get('APPWRITE_PROJECT_ID')

                if not appwrite_endpoint or not appwrite_project_id:
                    return jsonify({'error': 'Appwrite configuration incomplete'}), 500

                attachments_list = []

                for attachment_guid in attachment_guids:
                    # Get file metadata using the service
                    file_name, mime_type = appwrite_file_service.get_file_metadata(attachment_guid)

                    # Return proxy URLs that stream files (avoids loading into memory)
                    # Proxy handles Appwrite authentication and streams response
                    proxy_url = url_for('serve_attachment', transaction_id=transaction_id,
                                        attachment_guid=attachment_guid, _external=True)

                    attachments_list.append({
                        'attachment_guid': attachment_guid,
                        'file_url': proxy_url,
                        'download_url': proxy_url + '?download=1',
                        'file_name': file_name,
                        'mime_type': mime_type
                    })

                    logger.info(f"Generated proxy URL for {attachment_guid}: {proxy_url}")

                # Return list of attachments (backward compatible: single item list for single attachment)
                return jsonify({
                    'attachments': attachments_list,
                    'count': len(attachments_list)
                }), 200
            else:
                return jsonify({'error': 'Appwrite storage not available'}), 500

        elif request.method == 'DELETE':
            # Delete attachment(s) from Appwrite and update transaction
            # If specific_guid is provided in request, delete only that one; otherwise delete all
            specific_guid = request.args.get('guid')

            if appwrite_file_service.is_available():
                guids_to_delete = [specific_guid] if specific_guid else attachment_guids

                for guid in guids_to_delete:
                    success, error = appwrite_file_service.delete_file(guid)
                    if success:
                        logger.info(f"Deleted attachment {guid} from Appwrite")
                    else:
                        logger.error(f"Failed to delete attachment {guid}: {error}")
                        # Continue with other deletions even if one fails

            # Update database: remove deleted GUID(s) from attachments field
            if specific_guid and specific_guid in attachment_guids:
                # Remove only the specific GUID
                remaining_guids = [g for g in attachment_guids if g != specific_guid]
                new_attachments_value = ','.join(remaining_guids) if remaining_guids else None
            else:
                # Delete all attachments
                new_attachments_value = None

            cursor.execute("""
                UPDATE transactions
                SET attachments = %s
                WHERE id = %s
            """, (new_attachments_value, transaction_id))

            connection.commit()

            logger.info(f"Updated attachments for transaction {transaction_id}")

            return jsonify({'message': 'Attachment(s) deleted successfully'}), 200

    except Error as e:
        logger.error(f"Error managing attachment: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/attachment/view')
@login_required
def serve_attachment(transaction_id):
    """Proxy endpoint to serve attachment files with authentication."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = None
    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']

        # Verify transaction belongs to user and get attachment GUID(s)
        cursor.execute("""
            SELECT t.attachments
            FROM transactions t
            INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
            WHERE t.id = %s AND mr.user_id = %s
        """, (transaction_id, user_id))

        transaction = cursor.fetchone()

        if not transaction:
            return "Transaction not found", 404

        attachments_value = transaction['attachments']

        if not attachments_value:
            return "No attachment found", 404

        # Get specific GUID from query parameter, or default to first one
        requested_guid = request.args.get('attachment_guid')
        attachment_guids = [guid.strip() for guid in attachments_value.split(',') if guid.strip()]

        if requested_guid:
            # Validate requested GUID is in the list
            if requested_guid not in attachment_guids:
                return "Attachment not found", 404
            attachment_guid = requested_guid
        else:
            # Default to first attachment
            attachment_guid = attachment_guids[0]

        if not appwrite_file_service.is_available():
            return "Storage not available", 500

        # Get file metadata using the service
        file_name, mime_type = appwrite_file_service.get_file_metadata(attachment_guid)
        logger.info(f"Fetching attachment {attachment_guid}, mime: {mime_type}")

        # Download the file content using the service
        file_content, status_code, error = appwrite_file_service.download_file(attachment_guid)

        if not file_content:
            return error or "Failed to download file", status_code

        # Create response with the file content
        flask_response = Response(file_content, content_type=mime_type)
        flask_response.headers['Cache-Control'] = 'public, max-age=31536000'  # Cache for 1 year

        # Check if download parameter is present
        if request.args.get('download') == '1':
            flask_response.headers['Content-Disposition'] = f'attachment; filename="{file_name}"'
        else:
            flask_response.headers['Content-Disposition'] = 'inline'

        logger.info(
            f"Serving attachment {attachment_guid}: Content-Type={mime_type}, Mode={'download' if request.args.get('download') == '1' else 'inline'}, Size={len(file_content)} bytes")
        return flask_response


    except Error as e:
        logger.error(f"Error serving attachment: {str(e)}")
        return f"Database error: {str(e)}", 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/export', methods=['GET'])
@login_required
def export_transactions():
    """Export transactions to CSV, PDF, or Excel format."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']

        # Get parameters
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        export_format = request.args.get('format', 'csv')

        # Get monthly record
        cursor.execute("""
                       SELECT id
                       FROM monthly_records
                       WHERE user_id = %s AND year = %s AND month = %s
                       """, (user_id, year, month))

        monthly_record = cursor.fetchone()

        if not monthly_record:
            # Return empty file if no transactions
            transactions = []
        else:
            # Fetch transactions (DESC order for downloads - oldest first)
            cursor.execute("""
                           SELECT t.id,
                                  t.transaction_date,
                                  t.description,
                                  c.name  as category,
                                  t.debit,
                                  t.credit,
                                  t.notes,
                                  pm.name as payment_method,
                                  t.is_done,
                                  t.is_paid,
                                  t.paid_at
                           FROM transactions t
                                    LEFT JOIN categories c ON t.category_id = c.id
                                    LEFT JOIN payment_methods pm ON t.payment_method_id = pm.id
                           WHERE t.monthly_record_id = %s
                           ORDER BY t.display_order DESC, t.id DESC
                           """, (monthly_record['id'],))

            transactions = cursor.fetchall()

        # Generate file based on format
        if export_format == 'csv':
            return generate_csv(transactions, year, month)
        elif export_format == 'excel':
            return generate_excel(transactions, year, month)
        elif export_format == 'pdf':
            return generate_pdf(transactions, year, month)
        else:
            return jsonify({'error': 'Invalid format'}), 400

    except Error as e:
        logger.error(f"Error exporting transactions: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def generate_csv(transactions, year, month):
    """Generate CSV file from transactions."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(
        ['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance', 'Notes', 'Payment Method', 'Done', 'Paid',
         'Paid At'])

    # Transactions come in DESC order (oldest first in downloads)
    # Calculate running balance sequentially from oldest to newest
    balance = 0
    for t in transactions:
        debit = float(t['debit']) if t['debit'] else 0
        credit = float(t['credit']) if t['credit'] else 0
        balance += debit - credit

        writer.writerow([
            t['transaction_date'],
            t['description'],
            t['category'] or '',
            f"{debit:.2f}" if debit > 0 else '',
            f"{credit:.2f}" if credit > 0 else '',
            f"{balance:.2f}",
            t['notes'] or '',
            t['payment_method'] or '',
            'Yes' if t['is_done'] else 'No',
            'Yes' if t['is_paid'] else 'No',
            t['paid_at'] if t['paid_at'] else ''
        ])

    # Create response
    output.seek(0)
    month_name = calendar.month_name[month]
    filename = f'transactions_{month_name}_{year}.csv'

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


def generate_excel(transactions, year, month):
    """Generate Excel file from transactions."""
    if not EXCEL_AVAILABLE:
        # Fallback to CSV
        return generate_csv(transactions, year, month)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    month_name = calendar.month_name[month]
    ws.title = f"{month_name} {year}"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write header row
    headers = ['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance', 'Notes', 'Payment Method', 'Done',
               'Paid', 'Paid At']
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Transactions come in DESC order (oldest first in downloads)
    # Calculate running balance sequentially from oldest to newest
    balance = 0
    for t in transactions:
        debit = float(t['debit']) if t['debit'] else 0
        credit = float(t['credit']) if t['credit'] else 0
        balance += debit - credit

        ws.append([
            str(t['transaction_date']),
            t['description'],
            t['category'] or '',
            debit if debit > 0 else '',
            credit if credit > 0 else '',
            balance,
            t['notes'] or '',
            t['payment_method'] or '',
            'Yes' if t['is_done'] else 'No',
            'Yes' if t['is_paid'] else 'No',
            str(t['paid_at']) if t['paid_at'] else ''
        ])

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 20

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'transactions_{month_name}_{year}.xlsx'

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


def generate_pdf(transactions, year, month):
    """Generate PDF file from transactions."""
    if not PDF_AVAILABLE:
        # Fallback to CSV
        return generate_csv(transactions, year, month)

    month_name = calendar.month_name[month]
    output = io.BytesIO()

    # Create the PDF document
    doc = SimpleDocTemplate(output, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Add title
    title = Paragraph(f"<b>Transaction Report - {month_name} {year}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))

    # Create table data
    table_data = [['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance']]

    # Transactions come in DESC order (oldest first in downloads)
    # Calculate running balance sequentially from oldest to newest
    balance = 0
    for t in transactions:
        debit = float(t['debit']) if t['debit'] else 0
        credit = float(t['credit']) if t['credit'] else 0
        balance += debit - credit

        table_data.append([
            str(t['transaction_date']),
            t['description'][:30],  # Truncate long descriptions
            (t['category'] or '')[:15],
            f"{debit:.2f}" if debit > 0 else '',
            f"{credit:.2f}" if credit > 0 else '',
            f"{balance:.2f}"
        ])

    # Create table
    table = Table(table_data, colWidths=[1.0 * inch, 2.5 * inch, 1.2 * inch, 0.9 * inch, 0.9 * inch, 1.0 * inch])

    # Style the table
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

        # Body styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Date
        ('ALIGN', (1, 1), (2, -1), 'LEFT'),  # Description, Category
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Amounts
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)
    output.seek(0)

    filename = f'transactions_{month_name}_{year}.pdf'

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


@app.route('/api/transactions/reorder', methods=['POST'])
@login_required
def reorder_transactions():
    """Reorder transactions based on new order array of transaction IDs."""
    data = request.get_json()
    transaction_ids = data.get('transaction_ids', [])

    if not transaction_ids or not isinstance(transaction_ids, list):
        return jsonify({'error': 'Invalid transaction_ids. Must be a non-empty array'}), 400

    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            user_id = session.get('user_id')

            # Build a single UPDATE … CASE statement to set all
            # display_order values in one round-trip instead of N queries.
            case_clauses = []
            params = []
            for index, transaction_id in enumerate(transaction_ids):
                case_clauses.append("WHEN %s THEN %s")
                params.extend([transaction_id, index + 1])

            # Append the IN-list params
            params.extend(transaction_ids)
            placeholders = ','.join(['%s'] * len(transaction_ids))

            cursor.execute(
                f"UPDATE transactions SET display_order = CASE id "
                f"{' '.join(case_clauses)} END "
                f"WHERE id IN ({placeholders})",
                params,
            )

            connection.commit()
            return jsonify({
                'success': True,
                'message': 'Transaction order updated successfully'
            })

        except Error as e:
            connection.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/categories')
@login_required
def get_categories():
    """Get all categories."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            cursor.execute("SELECT * FROM categories ORDER BY type, name")
            categories = cursor.fetchall()
            return jsonify(categories)
        except Error as e:
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/categories', methods=['POST'])
@login_required
def add_category():
    """Add a new category."""
    data = request.get_json()

    # Validate input
    if not data or 'name' not in data or 'type' not in data:
        return jsonify({'error': 'Missing required fields: name and type'}), 400

    name = data['name'].strip()
    category_type = data['type'].strip().lower()
    percentage_markup = data.get('percentage_markup', 0.00)

    if not name:
        return jsonify({'error': 'Category name cannot be empty'}), 400

    if category_type not in ['income', 'expense']:
        return jsonify({'error': 'Category type must be either "income" or "expense"'}), 400

    # Validate percentage_markup
    try:
        percentage_markup = float(percentage_markup)
        if percentage_markup < 0 or percentage_markup > 100:
            return jsonify({'error': 'Percentage markup must be between 0 and 100'}), 400
    except (ValueError, TypeError):
        percentage_markup = 0.00

    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            # Check if category with same name and type already exists
            cursor.execute(
                "SELECT id FROM categories WHERE name = %s AND type = %s",
                (name, category_type)
            )
            existing = cursor.fetchone()

            if existing:
                return jsonify({'error': 'Category with this name and type already exists', 'id': existing['id']}), 409

            # Insert new category
            cursor.execute(
                "INSERT INTO categories (name, type, percentage_markup) VALUES (%s, %s, %s)",
                (name, category_type, percentage_markup)
            )
            connection.commit()

            # Get the newly created category
            category_id = cursor.lastrowid
            cursor.execute("SELECT * FROM categories WHERE id = %s", (category_id,))
            new_category = cursor.fetchone()

            return jsonify(new_category), 201
        except Error as e:
            connection.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/categories/<int:category_id>', methods=['PUT'])
@login_required
def update_category(category_id):
    """Update a category."""
    data = request.get_json()

    # Validate input
    if not data or 'name' not in data or 'type' not in data:
        return jsonify({'error': 'Missing required fields: name and type'}), 400

    name = data['name'].strip()
    category_type = data['type'].strip().lower()
    percentage_markup = data.get('percentage_markup', 0.00)

    if not name:
        return jsonify({'error': 'Category name cannot be empty'}), 400

    if category_type not in ['income', 'expense']:
        return jsonify({'error': 'Category type must be either "income" or "expense"'}), 400

    # Validate percentage_markup
    try:
        percentage_markup = float(percentage_markup)
        if percentage_markup < 0 or percentage_markup > 100:
            return jsonify({'error': 'Percentage markup must be between 0 and 100'}), 400
    except (ValueError, TypeError):
        percentage_markup = 0.00

    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            # Check if category exists
            cursor.execute("SELECT * FROM categories WHERE id = %s", (category_id,))
            category = cursor.fetchone()

            if not category:
                return jsonify({'error': 'Category not found'}), 404

            # Check if another category with same name and type already exists (excluding current category)
            cursor.execute(
                "SELECT id FROM categories WHERE name = %s AND type = %s AND id != %s",
                (name, category_type, category_id)
            )
            existing = cursor.fetchone()

            if existing:
                return jsonify({'error': 'Another category with this name and type already exists'}), 409

            # Update the category
            cursor.execute(
                "UPDATE categories SET name = %s, type = %s, percentage_markup = %s WHERE id = %s",
                (name, category_type, percentage_markup, category_id)
            )
            connection.commit()

            # Get the updated category
            cursor.execute("SELECT * FROM categories WHERE id = %s", (category_id,))
            updated_category = cursor.fetchone()

            return jsonify(updated_category), 200
        except Error as e:
            connection.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/categories/<int:category_id>', methods=['DELETE'])
@login_required
def delete_category(category_id):
    """Delete a category."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            # Check if category exists
            cursor.execute("SELECT * FROM categories WHERE id = %s", (category_id,))
            category = cursor.fetchone()

            if not category:
                return jsonify({'error': 'Category not found'}), 404

            # Check if category is being used in transactions
            cursor.execute(
                "SELECT COUNT(*) as count FROM transactions WHERE category_id = %s",
                (category_id,)
            )
            result = cursor.fetchone()

            if result and result['count'] > 0:
                return jsonify({
                    'error': f'Cannot delete category. It is being used by {result["count"]} transaction(s).',
                    'transaction_count': result['count']
                }), 409

            # Delete the category
            cursor.execute("DELETE FROM categories WHERE id = %s", (category_id,))
            connection.commit()

            return jsonify({'message': 'Category deleted successfully'}), 200
        except Error as e:
            connection.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/recalculate-balances', methods=['POST'])
@login_required
def recalculate_balances():
    """Deprecated: Balance calculation now happens on frontend."""
    return jsonify({
        'message': 'Balance calculation now happens on frontend',
        'transactions_updated': 0
    })


@app.route('/api/reports/monthly-summary')
@login_required
def monthly_summary_report():
    """Get monthly summary report."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            user_id = session['user_id']
            year = request.args.get('year', datetime.now().year, type=int)

            cursor.execute("""
                           SELECT mr.year,
                                  mr.month,
                                  mr.month_name,
                                  COALESCE(SUM(t.debit), 0)                              as total_income,
                                  COALESCE(SUM(t.credit), 0)                             as total_expenses,
                                  COALESCE(SUM(t.debit), 0) - COALESCE(SUM(t.credit), 0) as net_savings
                           FROM monthly_records mr
                                    LEFT JOIN transactions t ON mr.id = t.monthly_record_id
                           WHERE mr.user_id = %s
                             AND mr.year = %s
                           GROUP BY mr.year, mr.month, mr.month_name
                           ORDER BY mr.month
                           """, (user_id, year))

            summary = cursor.fetchall()
            return jsonify(summary)

        except Error as e:
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/reports/category-breakdown')
@login_required
def category_breakdown_report():
    """Get category breakdown report with weekly, monthly, or yearly views."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            user_id = session['user_id']
            range_type = request.args.get('range', 'monthly')  # weekly, monthly, yearly
            year = request.args.get('year', datetime.now().year, type=int)
            month = request.args.get('month', datetime.now().month, type=int)

            if range_type == 'weekly':
                # Get category spending by week for the specified month
                cursor.execute("""
                               SELECT WEEK(t.transaction_date, 1)                                             as week_num,
                                      DATE_FORMAT(MIN(t.transaction_date), '%%Y-%%m-%%d')                     as week_start,
                                      DATE_FORMAT(MAX(t.transaction_date), '%%Y-%%m-%%d')                     as week_end,
                                      c.name                                                                  as category,
                                      c.type,
                                      COALESCE(SUM(CASE WHEN c.type = 'income' THEN t.debit ELSE 0 END), 0)   as income,
                                      COALESCE(SUM(CASE WHEN c.type = 'expense' THEN t.credit ELSE 0 END), 0) as expense
                               FROM transactions t
                                        INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                                        INNER JOIN categories c ON t.category_id = c.id
                               WHERE mr.user_id = %s
                                 AND mr.year = %s
                                 AND mr.month = %s
                                 AND t.category_id IS NOT NULL
                               GROUP BY WEEK(t.transaction_date, 1), c.id, c.name, c.type
                               HAVING income > 0
                                   OR expense > 0
                               ORDER BY week_num, c.type, expense DESC, income DESC
                               """, (user_id, year, month))
            elif range_type == 'yearly':
                # Get category spending by year (all years)
                cursor.execute("""
                               SELECT mr.year,
                                      c.name                                                                  as category,
                                      c.type,
                                      COALESCE(SUM(CASE WHEN c.type = 'income' THEN t.debit ELSE 0 END), 0)   as income,
                                      COALESCE(SUM(CASE WHEN c.type = 'expense' THEN t.credit ELSE 0 END), 0) as expense
                               FROM transactions t
                                        INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                                        INNER JOIN categories c ON t.category_id = c.id
                               WHERE mr.user_id = %s
                                 AND t.category_id IS NOT NULL
                               GROUP BY mr.year, c.id, c.name, c.type
                               HAVING income > 0
                                   OR expense > 0
                               ORDER BY mr.year DESC, c.type, expense DESC, income DESC
                               """, (user_id,))
            else:  # monthly (default)
                # Get category spending for the specific selected month
                cursor.execute("""
                               SELECT mr.year,
                                      mr.month,
                                      mr.month_name,
                                      c.name                                                                  as category,
                                      c.type,
                                      COALESCE(SUM(CASE WHEN c.type = 'income' THEN t.debit ELSE 0 END), 0)   as income,
                                      COALESCE(SUM(CASE WHEN c.type = 'expense' THEN t.credit ELSE 0 END), 0) as expense
                               FROM transactions t
                                        INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                                        INNER JOIN categories c ON t.category_id = c.id
                               WHERE mr.user_id = %s
                                 AND mr.year = %s
                                 AND mr.month = %s
                                 AND t.category_id IS NOT NULL
                               GROUP BY mr.year, mr.month, mr.month_name, c.id, c.name, c.type
                               HAVING income > 0
                                   OR expense > 0
                               ORDER BY c.type, expense DESC, income DESC
                               """, (user_id, year, month))

            breakdown = cursor.fetchall()
            return jsonify(breakdown)

        except Error as e:
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/reports/cash-flow')
@login_required
def cash_flow_report():
    """Get cash flow analysis report with customizable date ranges using view."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            user_id = session['user_id']
            range_type = request.args.get('range', 'monthly')  # weekly, monthly, yearly
            year = request.args.get('year', datetime.now().year, type=int)
            month = request.args.get('month', datetime.now().month, type=int)

            if range_type == 'weekly':
                # Get weekly cash flow for the specified month
                cursor.execute("""
                               SELECT WEEK(t.transaction_date)                               as week_num,
                                      DATE_FORMAT(MIN(t.transaction_date), '%%Y-%%m-%%d')    as week_start,
                                      DATE_FORMAT(MAX(t.transaction_date), '%%Y-%%m-%%d')    as week_end,
                                      COALESCE(SUM(t.debit), 0)                              as cash_in,
                                      COALESCE(SUM(t.credit), 0)                             as cash_out,
                                      COALESCE(SUM(t.debit), 0) - COALESCE(SUM(t.credit), 0) as net_flow
                               FROM transactions t
                                        JOIN monthly_records mr ON t.monthly_record_id = mr.id
                               WHERE mr.user_id = %s
                                 AND mr.year = %s
                                 AND mr.month = %s
                               GROUP BY WEEK(t.transaction_date)
                               ORDER BY week_num
                               """, (user_id, year, month))
            elif range_type == 'yearly':
                # Get yearly cash flow using view
                cursor.execute("""
                               SELECT
                                   year, SUM (cash_in) as cash_in, SUM (cash_out) as cash_out, SUM (net_flow) as net_flow
                               FROM v_cash_flow
                               WHERE user_id = %s
                               GROUP BY year
                               ORDER BY year
                               """, (user_id,))
            else:  # monthly (default)
                # Get monthly cash flow using view
                cursor.execute("""
                               SELECT
                                   year, month, month_name, cash_in, cash_out, net_flow
                               FROM v_cash_flow
                               WHERE user_id = %s
                                 AND year = %s
                               ORDER BY month
                               """, (user_id, year))

            cash_flow = cursor.fetchall()
            return jsonify(cash_flow)

        except Error as e:
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/reports/top-spending')
@login_required
def top_spending_report():
    """Get top spending categories with customizable date ranges."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            user_id = session['user_id']
            range_type = request.args.get('range', 'monthly')
            year = request.args.get('year', datetime.now().year, type=int)
            month = request.args.get('month', datetime.now().month, type=int)
            limit = request.args.get('limit', 10, type=int)

            if range_type == 'weekly':
                # Get top spending for the specified month using view
                cursor.execute("""
                               SELECT category,
                                      type,
                                      total_spent,
                                      transaction_count,
                                      avg_amount
                               FROM v_top_spending
                               WHERE user_id = %s AND year = %s AND month = %s
                               ORDER BY total_spent DESC
                                   LIMIT %s
                               """, (user_id, year, month, limit))
            elif range_type == 'yearly':
                # Get top spending for the specified year using view
                cursor.execute("""
                               SELECT category,
                                      type,
                                      SUM(total_spent)       as total_spent,
                                      SUM(transaction_count) as transaction_count,
                                      AVG(avg_amount)        as avg_amount
                               FROM v_top_spending
                               WHERE user_id = %s AND year = %s
                               GROUP BY category_id, category, type
                               ORDER BY total_spent DESC
                                   LIMIT %s
                               """, (user_id, year, limit))
            else:  # monthly
                # Get top spending for the specified month using view
                cursor.execute("""
                               SELECT category,
                                      type,
                                      total_spent,
                                      transaction_count,
                                      avg_amount
                               FROM v_top_spending
                               WHERE user_id = %s AND year = %s AND month = %s
                               ORDER BY total_spent DESC
                                   LIMIT %s
                               """, (user_id, year, month, limit))

            top_spending = cursor.fetchall()
            return jsonify(top_spending)

        except Error as e:
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/reports/forecast')
@login_required
def forecast_report():
    """Predict next month's spending based on historical data."""
    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            user_id = session['user_id']
            months_to_analyze = request.args.get('months', 6, type=int)

            # Get historical data for the last N months using view
            cursor.execute("""
                           SELECT
                               year, month, month_name, cash_in as total_income, cash_out as total_expenses, net_flow as net_savings
                           FROM v_cash_flow
                           WHERE user_id = %s
                           ORDER BY year DESC, month DESC
                               LIMIT %s
                           """, (user_id, months_to_analyze))

            historical_data = cursor.fetchall()

            # Get category-wise spending patterns using view
            cursor.execute("""
                           SELECT category,
                                  AVG(total_spent)    as avg_monthly_spending,
                                  MIN(total_spent)    as min_spending,
                                  MAX(total_spent)    as max_spending,
                                  STDDEV(total_spent) as std_deviation
                           FROM v_top_spending
                           WHERE user_id = %s
                           GROUP BY category_id, category
                           ORDER BY avg_monthly_spending DESC
                           """, (user_id,))

            category_forecast = cursor.fetchall()

            # Calculate simple averages for forecasting
            if historical_data:
                avg_income = sum(float(row['total_income']) for row in historical_data) / len(historical_data)
                avg_expenses = sum(float(row['total_expenses']) for row in historical_data) / len(historical_data)
                avg_savings = sum(float(row['net_savings']) for row in historical_data) / len(historical_data)

                # Calculate trend (simple linear)
                if len(historical_data) >= 2:
                    recent_avg_expenses = sum(float(row['total_expenses']) for row in historical_data[:3]) / min(3,
                                                                                                                 len(historical_data))
                    older_avg_expenses = sum(float(row['total_expenses']) for row in historical_data[-3:]) / min(3,
                                                                                                                 len(historical_data))
                    trend = ((
                                     recent_avg_expenses - older_avg_expenses) / older_avg_expenses * 100) if older_avg_expenses > 0 else 0
                else:
                    trend = 0

                forecast = {
                    'next_month_forecast': {
                        'predicted_income': round(avg_income, 2),
                        'predicted_expenses': round(avg_expenses, 2),
                        'predicted_savings': round(avg_savings, 2),
                        'expense_trend': round(trend, 2),
                        'confidence': 'medium' if len(historical_data) >= 3 else 'low'
                    },
                    'historical_average': {
                        'avg_income': round(avg_income, 2),
                        'avg_expenses': round(avg_expenses, 2),
                        'avg_savings': round(avg_savings, 2)
                    },
                    'category_forecast': category_forecast,
                    'based_on_months': len(historical_data)
                }
            else:
                forecast = {
                    'next_month_forecast': {
                        'predicted_income': 0,
                        'predicted_expenses': 0,
                        'predicted_savings': 0,
                        'expense_trend': 0,
                        'confidence': 'no_data'
                    },
                    'historical_average': {
                        'avg_income': 0,
                        'avg_expenses': 0,
                        'avg_savings': 0
                    },
                    'category_forecast': [],
                    'based_on_months': 0
                }

            return jsonify(forecast)

        except Error as e:
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


@app.route('/api/payment-methods', methods=['GET', 'POST'])
@login_required
def payment_methods():
    """Get or create payment methods."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        if request.method == 'GET':
            cursor.execute("""
                           SELECT *
                           FROM payment_methods
                           WHERE user_id = %s
                             AND is_active = TRUE
                           ORDER BY type, name
                           """, (user_id,))

            methods = cursor.fetchall()
            return jsonify(methods)

        else:  # POST
            data = request.get_json()
            cursor.execute("""
                           INSERT INTO payment_methods (user_id, name, type, color)
                           VALUES (%s, %s, %s, %s)
                           """, (
                user_id,
                data.get('name'),
                data.get('type', 'credit_card'),
                data.get('color', '#007bff')
            ))

            connection.commit()
            return jsonify({'message': 'Payment method added successfully', 'id': cursor.lastrowid}), 201

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/payment-methods/<int:method_id>', methods=['DELETE'])
@login_required
def delete_payment_method(method_id):
    """Delete a payment method."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        cursor.execute("""
                       UPDATE payment_methods
                       SET is_active = FALSE
                       WHERE id = %s
                         AND user_id = %s
                       """, (method_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Payment method deleted successfully'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/mark-done', methods=['POST'])
@login_required
def mark_transaction_done(transaction_id):
    """Mark a transaction as done with payment method."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        data = request.get_json()
        payment_method_id = data.get('payment_method_id')

        cursor.execute("""
                       UPDATE transactions
                       SET is_done           = TRUE,
                           payment_method_id = %s,
                           marked_done_at    = CURRENT_TIMESTAMP
                       WHERE id = %s
                         AND monthly_record_id IN
                             (SELECT id FROM monthly_records WHERE user_id = %s)
                       """, (payment_method_id, transaction_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Transaction marked as done'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/mark-undone', methods=['POST'])
@login_required
def mark_transaction_undone(transaction_id):
    """Mark a transaction as not done."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        cursor.execute("""
                       UPDATE transactions
                       SET is_done           = FALSE,
                           payment_method_id = NULL,
                           marked_done_at    = NULL
                       WHERE id = %s
                         AND monthly_record_id IN
                             (SELECT id FROM monthly_records WHERE user_id = %s)
                       """, (transaction_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Transaction marked as not done'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/mark-paid', methods=['POST'])
@login_required
def mark_transaction_paid(transaction_id):
    """Mark a transaction as paid (when description cell is clicked)."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        data = request.get_json()
        payment_method_id = data.get('payment_method_id')

        cursor.execute("""
                       UPDATE transactions
                       SET is_done           = TRUE,
                           is_paid           = TRUE,
                           payment_method_id = %s,
                           marked_done_at    = CURRENT_TIMESTAMP,
                           paid_at           = CURRENT_TIMESTAMP
                       WHERE id = %s
                         AND monthly_record_id IN
                             (SELECT id FROM monthly_records WHERE user_id = %s)
                       """, (payment_method_id, transaction_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Transaction marked as paid'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/transactions/<int:transaction_id>/mark-unpaid', methods=['POST'])
@login_required
def mark_transaction_unpaid(transaction_id):
    """Unmark a transaction as paid (reverse the paid status)."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        cursor.execute("""
                       UPDATE transactions
                       SET is_done           = FALSE,
                           is_paid           = FALSE,
                           payment_method_id = NULL,
                           marked_done_at    = NULL,
                           paid_at           = NULL
                       WHERE id = %s
                         AND monthly_record_id IN
                             (SELECT id FROM monthly_records WHERE user_id = %s)
                       """, (transaction_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Transaction marked as unpaid'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/payment-method-totals', methods=['GET'])
@login_required
def get_payment_method_totals():
    """Get totals for each payment method for the current month."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)

        # Get monthly record
        cursor.execute("""
                       SELECT id
                       FROM monthly_records
                       WHERE user_id = %s AND year = %s AND month = %s
                       """, (user_id, year, month))

        monthly_record = cursor.fetchone()

        if not monthly_record:
            return jsonify([])

        # Get totals by payment method
        cursor.execute("""
                       SELECT pm.id,
                              pm.name,
                              pm.type,
                              pm.color,
                              COUNT(t.id)                                       as transaction_count,
                              SUM(t.debit)                                      as total_debit,
                              SUM(t.credit)                                     as total_credit,
                              SUM(COALESCE(t.debit, 0) - COALESCE(t.credit, 0)) as net_amount
                       FROM payment_methods pm
                                LEFT JOIN transactions t ON pm.id = t.payment_method_id
                           AND t.monthly_record_id = %s
                           AND t.is_done = TRUE
                       WHERE pm.user_id = %s
                         AND pm.is_active = TRUE
                       GROUP BY pm.id, pm.name, pm.type, pm.color
                       ORDER BY pm.type, pm.name
                       """, (monthly_record['id'], user_id))

        totals = cursor.fetchall()
        return jsonify(totals)

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


@app.route('/api/clone-month-transactions', methods=['POST'])
@login_required
def clone_month_transactions():
    """Clone all transactions from one month to another."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        data = request.get_json()
        from_year = data.get('from_year')
        from_month = data.get('from_month')
        to_year = data.get('to_year')
        to_month = data.get('to_month')
        include_payments = data.get('include_payments', False)

        # Validate inputs
        if not all([from_year, from_month, to_year, to_month]):
            return jsonify({'error': 'All date fields are required'}), 400

        if from_year == to_year and from_month == to_month:
            return jsonify({'error': 'Source and target months cannot be the same'}), 400

        # Get source monthly record
        cursor.execute("""
                       SELECT id
                       FROM monthly_records
                       WHERE user_id = %s AND year = %s AND month = %s
                       """, (user_id, from_year, from_month))

        source_record = cursor.fetchone()

        if not source_record:
            return jsonify({'error': 'Source month has no transactions'}), 404

        # Get or create target monthly record
        month_name = calendar.month_name[to_month]
        cursor.execute("""
                       INSERT INTO monthly_records (user_id, year, month, month_name)
                       VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                       UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                       """, (user_id, to_year, to_month, month_name))

        target_record = {'id': cursor.lastrowid}

        # Get all transactions from source month (preserve order)
        cursor.execute("""
                       SELECT description,
                              category_id,
                              debit,
                              credit,
                              notes,
                              payment_method_id,
                              is_done,
                              is_paid,
                              display_order
                       FROM transactions
                       WHERE monthly_record_id = %s
                       ORDER BY display_order ASC, id ASC
                       """, (source_record['id'],))

        source_transactions = cursor.fetchall()

        if not source_transactions:
            return jsonify({'error': 'No transactions found in source month'}), 404

        # Clone transactions using a single multi-row INSERT instead
        # of one INSERT per transaction.
        clone_date = datetime.now().date()
        insert_values = []
        insert_params = []

        for trans in source_transactions:
            debit = Decimal(str(trans['debit'])) if trans['debit'] else Decimal('0')
            credit = Decimal(str(trans['credit'])) if trans['credit'] else Decimal('0')

            payment_method_id = trans['payment_method_id'] if include_payments else None
            is_done = trans['is_done'] if include_payments else False
            is_paid = trans['is_paid'] if include_payments else False

            insert_values.append("(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
            insert_params.extend([
                target_record['id'],
                trans['description'],
                trans['category_id'],
                debit if debit > 0 else None,
                credit if credit > 0 else None,
                clone_date,
                trans['notes'],
                payment_method_id,
                is_done,
                is_paid,
                trans['display_order'],
            ])

        cursor.execute(
            "INSERT INTO transactions "
            "(monthly_record_id, description, category_id, debit, credit, "
            "transaction_date, notes, payment_method_id, is_done, is_paid, display_order) VALUES "
            + ", ".join(insert_values),
            insert_params,
        )
        cloned_count = len(insert_values)

        connection.commit()

        return jsonify({
            'message': f'Successfully cloned {cloned_count} transactions',
            'count': cloned_count
        }), 200

    except Error as e:
        connection.rollback()
        logger.error(f"Error cloning transactions: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


# ==================================================
# TAX CALCULATION ROUTES - Registered from services/tax_service.py
# ==================================================
# All tax calculation routes are now handled by the tax_service
# This reduces the size of app.py and improves code organization
# Routes include: POST /api/tax-calculations, GET /api/tax-calculations,
# GET /api/tax-calculations/<id>, DELETE /api/tax-calculations/<id>,
# PUT /api/tax-calculations/<id>/set-active


# ==================================================
# EXCHANGE RATE ROUTES - Registered from services/exchange_rate_routes.py
# ==================================================
# All exchange rate routes are now handled by the exchange_rate_routes service
# This reduces the size of app.py and improves code organization
# Routes include: /api/exchange-rate, /api/exchange-rate/month, /api/exchange-rate/import-csv,
# /api/exchange-rate/bulk-cache, /api/exchange-rate/hnb/current, /api/exchange-rate/pb/current,
# /api/exchange-rate/sampath/current, /api/exchange-rate/refresh-all, /api/exchange-rate/banks,
# /api/exchange-rate/bank/<bank_code>, /api/exchange-rate/pb, /api/exchange-rate/trends/all,
# /api/exchange-rate/ai-insights, /api/exchange-rate/intraday-logs


# Global error handlers (must be outside if __name__ block)
@app.errorhandler(500)
def internal_error(error):
    logger.error(f"Internal server error: {str(error)}")
    # For API requests, return JSON
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Internal server error'}), 500
    # For page requests, render error template or return HTML
    return render_template('error.html', error_code=500, error_message='Internal Server Error'), 500


@app.errorhandler(404)
def not_found(error):
    logger.warning(f"404 error: {request.path}")
    # For API requests, return JSON
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Not found'}), 404
    # For page requests, redirect to dashboard or login
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.errorhandler(Exception)
def handle_exception(error):
    logger.error(f"Unhandled exception: {str(error)}", exc_info=True)
    # For API requests, return JSON
    if request.path.startswith('/api/'):
        return jsonify({'error': 'An unexpected error occurred'}), 500
    # For page requests, render error template
    return render_template('error.html', error_code=500, error_message='An unexpected error occurred'), 500


if __name__ == '__main__':
    logger.info("Starting Personal Finance Budget application...")
    logger.info(f"Debug mode: False (Production)")
    logger.info(f"Host: 0.0.0.0, Port: 5003")
    app.run(debug=False, host='0.0.0.0', port=5003)
