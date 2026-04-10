"""
Transaction Service
Handles all Flask routes and business logic for transaction functionality
"""

import calendar
import csv
import io
import logging
import os
import uuid
from datetime import datetime
from decimal import Decimal

import re
import time
import threading

from flask import request, jsonify, session, make_response, url_for, Response
from mysql.connector import Error

from db import get_db_connection
from services.google_drive_file_service import get_google_drive_file_service
from services.gemini_bill_scanner import get_gemini_bill_scanner

logger = logging.getLogger(__name__)

# Try importing optional dependencies
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from PIL import Image

    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

# Initialize Google Drive file service
file_service = get_google_drive_file_service()


# ==================================================
# HELPER FUNCTIONS
# ==================================================

def fix_image_orientation(img):
    """
    Fix image orientation based on EXIF data.

    Mobile devices often store images in the wrong orientation and use EXIF
    tags to indicate how they should be rotated. This function reads the EXIF
    orientation tag and physically rotates/transposes the image accordingly.

    Args:
        img: PIL Image object

    Returns:
        PIL Image object with corrected orientation
    """
    try:
        # Get EXIF data
        exif = img.getexif()

        if exif is not None:
            # EXIF orientation tag is 0x0112 (274 in decimal)
            orientation = exif.get(0x0112, 1)

            # Apply rotation/transpose based on orientation tag
            # See: http://sylvana.net/jpegcrop/exif_orientation.html
            if orientation == 2:
                # Mirrored horizontally
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
            elif orientation == 3:
                # Rotated 180 degrees
                img = img.rotate(180, expand=True)
            elif orientation == 4:
                # Mirrored vertically
                img = img.transpose(Image.FLIP_TOP_BOTTOM)
            elif orientation == 5:
                # Mirrored horizontally then rotated 90 CCW
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
                img = img.rotate(90, expand=True)
            elif orientation == 6:
                # Rotated 90 degrees CCW (or 270 CW)
                img = img.rotate(270, expand=True)
            elif orientation == 7:
                # Mirrored horizontally then rotated 90 CW
                img = img.transpose(Image.FLIP_LEFT_RIGHT)
                img = img.rotate(270, expand=True)
            elif orientation == 8:
                # Rotated 90 degrees CW
                img = img.rotate(90, expand=True)

            # Reset orientation tag to normal after applying the rotation
            if orientation != 1:
                logger.info(f"Fixed image orientation (EXIF tag: {orientation})")
                # Remove EXIF orientation tag by creating new EXIF data
                exif[0x0112] = 1

    except (AttributeError, KeyError, IndexError) as e:
        # Image doesn't have EXIF data or orientation tag, that's fine
        logger.debug(f"No EXIF orientation data found: {str(e)}")

    return img


def optimize_file_for_upload(file_data, file_ext, original_filename):
    """
    Optimize file size for upload while maintaining quality.

    For images: Fixes orientation, resizes and compresses if >1MB

    Args:
        file_data: bytes - The original file data
        file_ext: str - File extension (e.g., 'jpg', 'pdf')
        original_filename: str - Original filename for logging

    Returns:
        tuple: (optimized_data: bytes, was_optimized: bool, original_size: int, new_size: int)
    """
    original_size = len(file_data)
    original_size_mb = original_size / (1024 * 1024)

    # Define size thresholds
    IMAGE_THRESHOLD_MB = 1.0  # Optimize images larger than 1MB

    # Image optimization
    if file_ext in ['jpg', 'jpeg', 'png', 'gif', 'webp'] and PIL_AVAILABLE:
        try:
            # ALWAYS load and fix orientation for images, regardless of size
            # This ensures mobile photos are displayed correctly
            logger.info(f"Processing image {original_filename}: {original_size_mb:.2f}MB")

            # Load image
            img = Image.open(io.BytesIO(file_data))

            # Fix orientation FIRST (critical for mobile photos)
            img = fix_image_orientation(img)

            needs_optimization = original_size_mb > IMAGE_THRESHOLD_MB

            # Resize if needed
            if needs_optimization:
                # Calculate new dimensions (max 2000px on longest side)
                max_dimension = 2000
                ratio = min(max_dimension / img.width, max_dimension / img.height, 1.0)

                if ratio < 1.0:
                    new_size = (int(img.width * ratio), int(img.height * ratio))
                    img = img.resize(new_size, Image.Resampling.LANCZOS)
                    logger.info(f"  Resized from original to {new_size}")

            # Convert to RGB if necessary (for JPEG)
            if img.mode in ('RGBA', 'P', 'LA'):
                # Create white background
                rgb_img = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                rgb_img.paste(img, mask=img.split()[-1] if img.mode in ('RGBA', 'LA') else None)
                img = rgb_img

            # Save processed version (with orientation fix applied)
            output = io.BytesIO()

            # Use appropriate format and quality
            if file_ext in ['jpg', 'jpeg']:
                img.save(output, format='JPEG', quality=85, optimize=True)
            elif file_ext == 'png':
                img.save(output, format='PNG', optimize=True, compress_level=6)
            elif file_ext == 'webp':
                img.save(output, format='WEBP', quality=85)
            else:
                img.save(output, format=img.format or 'JPEG', quality=85, optimize=True)

            processed_data = output.getvalue()
            new_size = len(processed_data)
            new_size_mb = new_size / (1024 * 1024)

            # Return processed image (with orientation fix always applied)
            if needs_optimization and new_size < original_size:
                reduction_pct = ((original_size - new_size) / original_size) * 100
                logger.info(
                    f"  ✓ Image optimized: {original_size_mb:.2f}MB → {new_size_mb:.2f}MB ({reduction_pct:.1f}% reduction)")
                return (processed_data, True, original_size, new_size)
            else:
                # Even if size didn't reduce, we still fixed orientation
                logger.info(f"  ✓ Image processed (orientation fixed): {original_size_mb:.2f}MB")
                return (processed_data, True, original_size, new_size)

        except Exception as e:
            logger.warning(f"Image processing failed: {str(e)}, using original")
            return (file_data, False, original_size, original_size)

    # No optimization needed or possible
    return (file_data, False, original_size, original_size)


def apply_percentage_markup(amount, category_id, payment_method_id, user_id, connection):
    """
    Apply percentage markup to an amount based on flexible markup rules.

    Rules are checked in the following priority order:
    1. Exact match (both category and payment method)
    2. Payment method only
    3. Category only

    For example, if a rule has 1.5% markup and amount is 1000:
    - Result will be 1000 + (1000 * 1.5%) = 1015

    Args:
        amount: Decimal - The base amount
        category_id: int or None - Category ID
        payment_method_id: int or None - Payment method ID
        user_id: int - User ID to fetch user-specific rules
        connection: MySQL connection object

    Returns:
        Decimal - Amount with markup applied (or original amount if no markup)
    """
    if not amount or amount == 0:
        return amount

    try:
        cursor = connection.cursor(dictionary=True)

        # Try to find the best matching rule with priority:
        # 1. Exact match (category + payment method)
        # 2. Payment method only
        # 3. Category only

        query = """
            SELECT 
                id, rule_name, category_id, payment_method_id, 
                percentage_markup, priority
            FROM markup_rules
            WHERE user_id = %s 
                AND is_active = TRUE
                AND percentage_markup > 0
                AND (
                    -- Exact match: both category and payment method
                    (category_id = %s AND payment_method_id = %s)
                    OR
                    -- Payment method only match
                    (category_id IS NULL AND payment_method_id = %s)
                    OR
                    -- Category only match
                    (category_id = %s AND payment_method_id IS NULL)
                )
            ORDER BY 
                -- Priority 1: Both category and payment method (most specific)
                CASE WHEN category_id IS NOT NULL AND payment_method_id IS NOT NULL THEN 1 ELSE 2 END,
                -- Priority 2: Sort by rule priority field
                priority DESC,
                -- Priority 3: Newer rules first
                created_at DESC
            LIMIT 1
        """

        cursor.execute(query, (
            user_id,
            category_id, payment_method_id,  # Exact match
            payment_method_id,                # Payment method only
            category_id                        # Category only
        ))

        result = cursor.fetchone()
        cursor.close()

        if result:
            markup_percentage = Decimal(str(result['percentage_markup']))
            markup_amount = (amount * markup_percentage) / Decimal('100')
            final_amount = amount + markup_amount

            rule_type = "combination"
            if result['category_id'] and result['payment_method_id']:
                rule_type = "category + payment method"
            elif result['payment_method_id']:
                rule_type = "payment method only"
            elif result['category_id']:
                rule_type = "category only"

            logger.info(
                f"Applied {markup_percentage}% markup ({rule_type} rule '{result['rule_name']}'): "
                f"{amount} + {markup_amount} = {final_amount}"
            )
            return final_amount

    except Exception as e:
        logger.error(f"Error applying markup: {str(e)}")

    return amount


def auto_categorize_transaction(description, connection):
    """
    Auto-categorize a transaction based on keyword matching.

    Searches the category_keywords table for keywords that match the description.
    Returns the first matching category_id, or None if no match found.

    Args:
        description: str - The transaction description
        connection: MySQL connection object

    Returns:
        int or None - Category ID if match found, None otherwise
    """
    if not description:
        return None

    try:
        cursor = connection.cursor(dictionary=True)

        # Get all keywords and their categories
        cursor.execute("""
            SELECT category_id, keyword 
            FROM category_keywords
            ORDER BY LENGTH(keyword) DESC
        """)

        keywords = cursor.fetchall()
        cursor.close()

        # Convert description to lowercase for case-insensitive matching
        description_lower = description.lower()

        # Check each keyword for a match
        for row in keywords:
            keyword = row['keyword'].lower()
            if keyword in description_lower:
                category_id = row['category_id']
                logger.info(f"Auto-categorized '{description}' to category {category_id} (matched keyword: '{keyword}')")
                return category_id

        logger.debug(f"No category match found for description: '{description}'")
        return None

    except Exception as e:
        logger.error(f"Error in auto-categorization: {str(e)}")
        return None


def log_transaction_audit(cursor, transaction_id, user_id, action, field_name=None, old_value=None, new_value=None):
    """
    Log transaction changes to audit trail.

    Args:
        cursor: Database cursor
        transaction_id: ID of the transaction (None for DELETE after completion)
        user_id: ID of the user making the change
        action: Type of action (CREATE, UPDATE, DELETE)
        field_name: Name of the field changed (None for CREATE/DELETE)
        old_value: Previous value (None for CREATE)
        new_value: New value (None for DELETE)
    """
    try:
        # Get request metadata
        ip_address = request.remote_addr if request else None
        user_agent = request.headers.get('User-Agent') if request else None

        # Convert values to strings for storage
        old_value_str = str(old_value) if old_value is not None else None
        new_value_str = str(new_value) if new_value is not None else None

        cursor.execute("""
                       INSERT INTO transaction_audit_logs
                       (transaction_id, user_id, action, field_name, old_value, new_value, ip_address, user_agent)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                       """, (transaction_id, user_id, action, field_name, old_value_str, new_value_str, ip_address,
                             user_agent))

        logger.info(f"Audit log created: {action} on transaction {transaction_id} by user {user_id}")
    except Exception as e:
        logger.error(f"Failed to create audit log: {e}")
        # Don't fail the main transaction if audit logging fails


# ==================================================
# EXPORT HELPERS
# ==================================================

def generate_csv(transactions, year, month):
    """Generate CSV file from transactions."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(
        ['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance', 'Notes', 'Payment Method', 'Done', 'Paid',
         'Paid At'])

    # Transactions come in DESC order (oldest first in downloads)
    # Calculate running balance sequentially from oldest to newest
    balance = 0
    for t in transactions:
        debit = float(t['debit']) if t['debit'] else 0
        credit = float(t['credit']) if t['credit'] else 0
        balance += debit - credit

        writer.writerow([
            t['transaction_date'],
            t['description'],
            t['category'] or '',
            f"{debit:.2f}" if debit > 0 else '',
            f"{credit:.2f}" if credit > 0 else '',
            f"{balance:.2f}",
            t['notes'] or '',
            t['payment_method'] or '',
            'Yes' if t['is_done'] else 'No',
            'Yes' if t['is_paid'] else 'No',
            t['paid_at'] if t['paid_at'] else ''
        ])

    # Create response
    output.seek(0)
    month_name = calendar.month_name[month]
    filename = f'transactions_{month_name}_{year}.csv'

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


def generate_excel(transactions, year, month):
    """Generate Excel file from transactions."""
    if not EXCEL_AVAILABLE:
        # Fallback to CSV
        return generate_csv(transactions, year, month)

    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    month_name = calendar.month_name[month]
    ws.title = f"{month_name} {year}"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write header row
    headers = ['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance', 'Notes', 'Payment Method', 'Done',
               'Paid', 'Paid At']
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Transactions come in DESC order (oldest first in downloads)
    # Calculate running balance sequentially from oldest to newest
    balance = 0
    for t in transactions:
        debit = float(t['debit']) if t['debit'] else 0
        credit = float(t['credit']) if t['credit'] else 0
        balance += debit - credit

        ws.append([
            str(t['transaction_date']),
            t['description'],
            t['category'] or '',
            debit if debit > 0 else '',
            credit if credit > 0 else '',
            balance,
            t['notes'] or '',
            t['payment_method'] or '',
            'Yes' if t['is_done'] else 'No',
            'Yes' if t['is_paid'] else 'No',
            str(t['paid_at']) if t['paid_at'] else ''
        ])

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 20

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'transactions_{month_name}_{year}.xlsx'

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


def generate_pdf(transactions, year, month):
    """Generate PDF file from transactions."""
    if not PDF_AVAILABLE:
        # Fallback to CSV
        return generate_csv(transactions, year, month)

    month_name = calendar.month_name[month]
    output = io.BytesIO()

    # Create the PDF document
    doc = SimpleDocTemplate(output, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # Add title
    title = Paragraph(f"<b>Transaction Report - {month_name} {year}</b>", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))

    # Create table data
    table_data = [['Date', 'Description', 'Category', 'Debit', 'Credit', 'Balance']]

    # Transactions come in DESC order (oldest first in downloads)
    # Calculate running balance sequentially from oldest to newest
    balance = 0
    for t in transactions:
        debit = float(t['debit']) if t['debit'] else 0
        credit = float(t['credit']) if t['credit'] else 0
        balance += debit - credit

        table_data.append([
            str(t['transaction_date']),
            t['description'][:30],  # Truncate long descriptions
            (t['category'] or '')[:15],
            f"{debit:.2f}" if debit > 0 else '',
            f"{credit:.2f}" if credit > 0 else '',
            f"{balance:.2f}"
        ])

    # Create table
    table = Table(table_data, colWidths=[1.0 * inch, 2.5 * inch, 1.2 * inch, 0.9 * inch, 0.9 * inch, 1.0 * inch])

    # Style the table
    table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

        # Body styling
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Date
        ('ALIGN', (1, 1), (2, -1), 'LEFT'),  # Description, Category
        ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),  # Amounts
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)
    output.seek(0)

    filename = f'transactions_{month_name}_{year}.pdf'

    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'attachment; filename={filename}'

    return response


# ==================================================
# TRANSACTION ROUTE HANDLERS
# ==================================================

def transactions_handler():
    """Get or create transactions."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        if request.method == 'GET':
            # Get query parameters
            year = request.args.get('year', datetime.now().year, type=int)
            month = request.args.get('month', datetime.now().month, type=int)
            search_all = request.args.get('searchAll', 'false').lower() == 'true'

            # Get filter parameters
            description = request.args.get('description', '')
            notes_filter = request.args.get('notes', '')
            categories = request.args.get('categories', '')  # comma-separated IDs
            payment_methods = request.args.get('paymentMethods', '')  # comma-separated IDs
            types = request.args.get('types', '')  # comma-separated: income,expense
            statuses = request.args.get('statuses', '')  # comma-separated: done,not_done,paid,unpaid
            min_amount = request.args.get('minAmount', type=float)
            max_amount = request.args.get('maxAmount', type=float)
            start_date = request.args.get('startDate', '')
            end_date = request.args.get('endDate', '')

            # Check if any filters are active
            has_filters = any([
                description, notes_filter, categories, payment_methods,
                types, statuses, min_amount is not None, max_amount is not None,
                start_date, end_date
            ])

            # Build dynamic WHERE clause
            where_clauses = []
            params = []

            # Use a JOIN on monthly_records instead of a separate
            # query to fetch IDs first (eliminates one round-trip).
            # The mr.user_id filter is always applied via the JOIN.
            where_clauses.append("mr.user_id = %s")
            params.append(user_id)

            if search_all or has_filters:
                # Parse date range to extract year and month if provided
                if start_date:
                    try:
                        start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                        where_clauses.append("(mr.year > %s OR (mr.year = %s AND mr.month >= %s))")
                        params.extend([start_dt.year, start_dt.year, start_dt.month])
                    except ValueError:
                        pass

                if end_date:
                    try:
                        end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                        where_clauses.append("(mr.year < %s OR (mr.year = %s AND mr.month <= %s))")
                        params.extend([end_dt.year, end_dt.year, end_dt.month])
                    except ValueError:
                        pass
            else:
                # Normal behavior - limit to specific month
                where_clauses.append("mr.year = %s")
                where_clauses.append("mr.month = %s")
                params.extend([year, month])

            # Continue with filter building only if we have WHERE clauses
            if where_clauses:

                # Description filter
                if description:
                    where_clauses.append("LOWER(t.description) LIKE %s")
                    params.append(f"%{description.lower()}%")

                # Notes filter
                if notes_filter:
                    where_clauses.append("LOWER(t.notes) LIKE %s")
                    params.append(f"%{notes_filter.lower()}%")

                # Category filter
                if categories:
                    cat_ids = [int(cid) for cid in categories.split(',') if cid.strip()]
                    if cat_ids:
                        placeholders = ','.join(['%s'] * len(cat_ids))
                        where_clauses.append(f"t.category_id IN ({placeholders})")
                        params.extend(cat_ids)

                # Payment method filter
                if payment_methods:
                    pm_ids = [int(pmid) for pmid in payment_methods.split(',') if pmid.strip()]
                    if pm_ids:
                        placeholders = ','.join(['%s'] * len(pm_ids))
                        where_clauses.append(f"t.payment_method_id IN ({placeholders})")
                        params.extend(pm_ids)

                # Transaction type filter
                if types:
                    type_list = [t.strip() for t in types.split(',') if t.strip()]
                    type_conditions = []
                    if 'income' in type_list:
                        type_conditions.append("t.debit > 0")
                    if 'expense' in type_list:
                        type_conditions.append("t.credit > 0")
                    if type_conditions:
                        where_clauses.append(f"({' OR '.join(type_conditions)})")

                # Status filter
                if statuses:
                    status_list = [s.strip() for s in statuses.split(',') if s.strip()]
                    status_conditions = []
                    if 'done' in status_list:
                        status_conditions.append("t.is_done = TRUE")
                    if 'not_done' in status_list:
                        status_conditions.append("t.is_done = FALSE OR t.is_done IS NULL")
                    if 'paid' in status_list:
                        status_conditions.append("t.is_paid = TRUE")
                    if 'unpaid' in status_list:
                        status_conditions.append("t.is_paid = FALSE OR t.is_paid IS NULL")
                    if status_conditions:
                        where_clauses.append(f"({' OR '.join(status_conditions)})")

                # Amount range filter
                if min_amount is not None:
                    where_clauses.append("(COALESCE(t.debit, 0) >= %s OR COALESCE(t.credit, 0) >= %s)")
                    params.extend([min_amount, min_amount])

                if max_amount is not None:
                    where_clauses.append("(COALESCE(t.debit, 0) <= %s OR COALESCE(t.credit, 0) <= %s)")
                    params.extend([max_amount, max_amount])

                # Date range filter is now handled via monthly_records filtering above
                # No need to filter on transaction dates here

                # Combine WHERE clauses
                where_sql = " AND ".join(where_clauses)

                query = f"""
                    SELECT
                        t.*,
                        c.name as category_name,
                        pm.name as payment_method_name,
                        pm.type as payment_method_type,
                        pm.color as payment_method_color,
                        COALESCE(t.is_paid, FALSE) as is_paid
                    FROM transactions t
                    INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                    LEFT JOIN categories c ON t.category_id = c.id
                    LEFT JOIN payment_methods pm ON t.payment_method_id = pm.id
                    WHERE {where_sql}
                    ORDER BY t.display_order ASC, t.id ASC
                """

                cursor.execute(query, params)
                transactions = cursor.fetchall()
                return jsonify(transactions)
            else:
                return jsonify([])

        else:  # POST - Create new transaction
            # Check if request is multipart/form-data (with image) or JSON (without image)
            is_multipart = request.content_type and 'multipart/form-data' in request.content_type

            if is_multipart:
                # Get form data
                data = request.form.to_dict()
                # Convert string numbers to proper types
                for key in ['debit', 'credit', 'category_id', 'year', 'month', 'payment_method_id']:
                    if key in data and data[key]:
                        try:
                            if key in ['debit', 'credit']:
                                data[key] = float(data[key]) if data[key] else None
                            else:
                                data[key] = int(data[key]) if data[key] else None
                        except (ValueError, TypeError):
                            data[key] = None
            else:
                data = request.get_json()

            print(f"[DEBUG] Received transaction data: {data}")

            # Handle bill image(s) upload to Google Drive (if provided)
            attachment_guids = []
            attachments_value = None

            # Check if attachments were already uploaded (new sequential upload flow)
            if 'attachments' in data and data['attachments']:
                # Images already uploaded sequentially, use provided GUIDs
                logger.info(f"Using pre-uploaded attachments: {data['attachments']}")
                attachments_value = data['attachments']
            else:
                # Legacy flow: handle multipart file uploads
                # Check for multiple images first, then fall back to single image (backward compatible)
                bill_images = []
                if is_multipart:
                    if 'bill_images' in request.files:
                        bill_images = request.files.getlist('bill_images')
                    elif 'bill_image' in request.files:
                        bill_images = [request.files['bill_image']]

                # Process each image
                for idx, bill_image in enumerate(bill_images):
                    if not bill_image or not bill_image.filename:
                        continue

                    # Validate file type
                    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}
                    file_ext = bill_image.filename.rsplit('.', 1)[-1].lower() if '.' in bill_image.filename else ''

                    if file_ext not in allowed_extensions:
                        logger.warning(f"Skipping file {idx + 1} with invalid extension: {file_ext}")
                        continue

                    # Read file data
                    file_data = bill_image.read()

                    if len(file_data) == 0:
                        logger.warning(f"Skipping empty file {idx + 1}")
                        continue

                    if file_service.is_available():
                        try:
                            # Optimize file if needed (resize images, compress PDFs)
                            optimized_data, was_optimized, original_size, final_size = optimize_file_for_upload(
                                file_data,
                                file_ext,
                                bill_image.filename
                            )

                            # Use optimized data for upload
                            image_data = optimized_data

                            # Log optimization results
                            if was_optimized:
                                reduction_pct = ((original_size - final_size) / original_size) * 100
                                logger.info(
                                    f"✓ Image {idx + 1}/{len(bill_images)} optimized: {original_size / (1024 * 1024):.2f}MB → {final_size / (1024 * 1024):.2f}MB ({reduction_pct:.1f}% smaller)")
                            else:
                                logger.info(
                                    f"Image {idx + 1}/{len(bill_images)} size OK, no optimization needed: {original_size / (1024 * 1024):.2f}MB")

                            # Generate pure GUID for filename
                            attachment_guid = str(uuid.uuid4())
                            filename = f"{attachment_guid}.{file_ext}"

                            # Check file size
                            file_size_mb = len(image_data) / (1024 * 1024)
                            logger.info(
                                f"Uploading image {idx + 1}/{len(bill_images)} to Google Drive: {filename}, type: {file_ext}, size: {len(image_data)} bytes ({file_size_mb:.2f}MB)")

                            # Check first bytes for validation
                            first_bytes = image_data[:8] if len(image_data) >= 8 else image_data
                            logger.info(f"First 8 bytes: {first_bytes} (hex: {first_bytes.hex()})")

                            # Validate PDF header if uploading PDF
                            if file_ext == 'pdf':
                                if not first_bytes.startswith(b'%PDF'):
                                    logger.error(
                                        f"WARNING: PDF file does NOT start with %PDF header before upload!")
                                    logger.error(f"First 20 bytes: {image_data[:20]}")
                                else:
                                    logger.info(f"✓ PDF header valid before upload")

                            # Warn if file is still unusually large after optimization
                            if file_size_mb > 10:
                                logger.warning(
                                    f"Large file upload: {file_size_mb:.2f}MB - may take time to process")

                            # Upload to Google Drive using the file service
                            success, error, result = file_service.upload_file(
                                image_data,
                                attachment_guid,
                                filename
                            )

                            if success:
                                drive_file_id = result.get('id', attachment_guid) if result else attachment_guid
                                stored_size = result.get('sizeOriginal', 'N/A') if result else 'N/A'
                                logger.info(
                                    f"✓ Image {idx + 1}/{len(bill_images)} uploaded successfully: {drive_file_id}, stored size: {stored_size}")
                                # Add to list of uploaded file IDs
                                attachment_guids.append(drive_file_id)
                            else:
                                logger.error(f"Failed to upload bill image {idx + 1} to Google Drive: {error}")
                                # Continue with other images even if one fails
                        except Exception as e:
                            logger.error(f"Failed to process and upload bill image {idx + 1}: {str(e)}")
                            # Continue with other images even if one fails
                    else:
                        logger.warning(f"File storage not configured for image {idx + 1}")

                # Store comma-separated GUIDs in attachments field (or None if no uploads succeeded)
                attachments_value = ','.join(attachment_guids) if attachment_guids else None
                if attachments_value:
                    logger.info(f"Stored {len(attachment_guids)} attachment(s): {attachments_value}")

            # Get or create monthly record
            year = data.get('year', datetime.now().year)
            month = data.get('month', datetime.now().month)
            month_name = calendar.month_name[month]

            cursor.execute("""
                           INSERT INTO monthly_records (user_id, year, month, month_name)
                           VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                           UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                           """, (user_id, year, month, month_name))

            monthly_record = {'id': cursor.lastrowid}

            # Convert to Decimal to avoid float/Decimal arithmetic errors
            debit_value = data.get('debit')
            credit_value = data.get('credit')

            debit = Decimal(str(debit_value)) if debit_value else Decimal('0')
            credit = Decimal(str(credit_value)) if credit_value else Decimal('0')

            # Get category_id from data, or auto-categorize if not provided
            category_id = data.get('category_id') or None
            if not category_id:
                category_id = auto_categorize_transaction(data.get('description'), connection)

            # Get payment_method_id
            payment_method_id = data.get('payment_method_id')

            # Apply flexible percentage markup based on category, payment method, or both
            if category_id or payment_method_id:
                if credit > 0:
                    credit = apply_percentage_markup(credit, category_id, payment_method_id, user_id, connection)
                elif debit > 0:
                    debit = apply_percentage_markup(debit, category_id, payment_method_id, user_id, connection)

            print(f"[DEBUG] Debit: {debit}, Credit: {credit}")

            # Use current date if no transaction_date provided
            transaction_date = data.get('transaction_date')
            if not transaction_date:
                transaction_date = datetime.now().date()

            # Push all existing transactions down by incrementing their display_order
            # This makes room for the new transaction at position 1 (top)
            cursor.execute("""
                           UPDATE transactions
                           SET display_order = display_order + 1
                           WHERE monthly_record_id = %s
                           """, (monthly_record['id'],))

            # New transaction gets display_order = 1 (appears at top)
            next_display_order = 1

            # Get bill content if provided (from scanned bills)
            bill_content = data.get('bill_content')

            # Get payment status from frontend (explicit is_paid flag)
            payment_method_id = data.get('payment_method_id')
            is_paid = data.get('is_paid', False)  # Use explicit flag from frontend

            # Get paid_at timestamp if provided
            paid_at = None
            if is_paid:
                paid_at_str = data.get('paid_at')
                if paid_at_str:
                    # Parse ISO timestamp from frontend
                    try:
                        paid_at = datetime.fromisoformat(paid_at_str.replace('Z', '+00:00'))
                    except (ValueError, AttributeError):
                        # If parsing fails, use current timestamp
                        paid_at = datetime.now()
                else:
                    # No paid_at provided but marked as paid, use current timestamp
                    paid_at = datetime.now()

            # Insert transaction with attachments field
            insert_values = (
                monthly_record['id'],
                data.get('description'),
                category_id,
                debit if debit > 0 else None,
                credit if credit > 0 else None,
                transaction_date,
                data.get('notes'),
                next_display_order,
                bill_content,
                attachments_value,  # Store comma-separated GUIDs in attachments column
                payment_method_id,  # Add payment method
                is_paid,  # Mark as paid if payment method is selected
                paid_at  # Timestamp when paid
            )
            print(f"[DEBUG] Inserting transaction with values: {insert_values}")

            cursor.execute("""
                           INSERT INTO transactions
                           (monthly_record_id, description, category_id, debit, credit, transaction_date, notes,
                            display_order, bill_content, attachments, payment_method_id, is_paid, paid_at)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                           """, insert_values)

            transaction_id = cursor.lastrowid
            print(f"[DEBUG] Transaction inserted with ID: {transaction_id}")

            connection.commit()
            print(f"[DEBUG] Transaction committed successfully")

            response = {'message': 'Transaction created successfully', 'id': transaction_id}
            if attachments_value:
                response['attachments'] = attachments_value

            return jsonify(response), 201

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def filter_transactions_handler():
    """Filter transactions across all data with advanced criteria."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session.get('user_id')

        # Get filter parameters
        date_from = request.args.get('dateFrom')
        date_to = request.args.get('dateTo')
        category_id = request.args.get('category')
        payment_method_id = request.args.get('paymentMethod')
        amount_min = request.args.get('amountMin')
        amount_max = request.args.get('amountMax')
        transaction_type = request.args.get('transactionType')  # 'debit' or 'credit'
        search_text = request.args.get('searchText')
        done_status = request.args.get('doneStatus')  # 'done', 'not_done', or empty
        paid_status = request.args.get('paidStatus')  # 'paid', 'not_paid', or empty

        # Build SQL query with filters
        query = """
                SELECT t.*,
                       c.name   as category_name,
                       pm.name  as payment_method_name,
                       pm.type  as payment_method_type,
                       pm.color as payment_method_color,
                       mr.year,
                       mr.month,
                       mr.month_name
                FROM transactions t
                         LEFT JOIN categories c ON t.category_id = c.id
                         LEFT JOIN payment_methods pm ON t.payment_method_id = pm.id
                         INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                WHERE mr.user_id = %s \
                """

        params = [user_id]

        # Add date range filter (based on monthly_records year/month)
        if date_from:
            try:
                start_dt = datetime.strptime(date_from, '%Y-%m-%d')
                start_year = start_dt.year
                start_month = start_dt.month
                query += " AND (mr.year > %s OR (mr.year = %s AND mr.month >= %s))"
                params.extend([start_year, start_year, start_month])
            except ValueError:
                pass

        if date_to:
            try:
                end_dt = datetime.strptime(date_to, '%Y-%m-%d')
                end_year = end_dt.year
                end_month = end_dt.month
                query += " AND (mr.year < %s OR (mr.year = %s AND mr.month <= %s))"
                params.extend([end_year, end_year, end_month])
            except ValueError:
                pass

        # Add category filter
        if category_id:
            query += " AND t.category_id = %s"
            params.append(int(category_id))

        # Add payment method filter
        if payment_method_id:
            query += " AND t.payment_method_id = %s"
            params.append(int(payment_method_id))

        # Add amount filter based on transaction type
        if transaction_type == 'debit':
            # Income transactions (debit > 0)
            query += " AND t.debit > 0"
            if amount_min:
                query += " AND t.debit >= %s"
                params.append(float(amount_min))
            if amount_max:
                query += " AND t.debit <= %s"
                params.append(float(amount_max))
        elif transaction_type == 'credit':
            # Expense transactions (credit > 0)
            query += " AND t.credit > 0"
            if amount_min:
                query += " AND t.credit >= %s"
                params.append(float(amount_min))
            if amount_max:
                query += " AND t.credit <= %s"
                params.append(float(amount_max))
        else:
            # Both types - check either debit or credit
            if amount_min or amount_max:
                amount_conditions = []
                if amount_min:
                    amount_conditions.append("(t.debit >= %s OR t.credit >= %s)")
                    params.extend([float(amount_min), float(amount_min)])
                if amount_max:
                    amount_conditions.append("(t.debit <= %s OR t.credit <= %s)")
                    params.extend([float(amount_max), float(amount_max)])
                if amount_conditions:
                    query += " AND " + " AND ".join(amount_conditions)

        # Add text search filter (search in description and notes)
        if search_text:
            query += " AND (t.description LIKE %s OR t.notes LIKE %s)"
            search_pattern = f"%{search_text}%"
            params.extend([search_pattern, search_pattern])

        # Add done status filter
        if done_status == 'done':
            query += " AND t.is_done = 1"
        elif done_status == 'not_done':
            query += " AND t.is_done = 0"

        # Add paid status filter
        if paid_status == 'paid':
            query += " AND t.is_paid = 1"
        elif paid_status == 'not_paid':
            query += " AND t.is_paid = 0"

        # Order by date descending (most recent first)
        query += " ORDER BY t.transaction_date DESC, t.id DESC"

        # Limit results to prevent overload (max 500 transactions)
        query += " LIMIT 500"

        cursor.execute(query, params)
        transactions = cursor.fetchall()

        # Calculate running balance for filtered transactions
        running_balance = 0
        for trans in reversed(transactions):
            running_balance += (trans['debit'] or 0) - (trans['credit'] or 0)
            trans['balance'] = running_balance

        # Reverse back to show most recent first
        transactions.reverse()

        return jsonify(transactions)

    except Exception as e:
        print(f"Error filtering transactions: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def manage_transaction_handler(transaction_id):
    """Update or delete a transaction."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        if request.method == 'PUT':
            data = request.get_json()

            # Validate request data
            if not data:
                return jsonify({'error': 'No data provided'}), 400

            print(f"[DEBUG] Updating transaction {transaction_id} with data: {data}")

            # Get the current transaction data for audit trail
            dict_cursor = connection.cursor(dictionary=True)
            dict_cursor.execute("""
                                SELECT t.*
                                FROM transactions t
                                WHERE t.id = %s
                                  AND t.monthly_record_id IN
                                      (SELECT id FROM monthly_records WHERE user_id = %s)
                                """, (transaction_id, session['user_id']))

            old_transaction = dict_cursor.fetchone()
            dict_cursor.close()

            if not old_transaction:
                print(f"[DEBUG] Transaction {transaction_id} not found for user {session['user_id']}")
                return jsonify({'error': 'Transaction not found'}), 404

            monthly_record_id = old_transaction['monthly_record_id']
            print(f"[DEBUG] Found monthly_record_id: {monthly_record_id}")

            # Convert to Decimal to avoid float/Decimal arithmetic errors
            debit_value = data.get('debit')
            credit_value = data.get('credit')

            debit = Decimal(str(debit_value)) if debit_value else Decimal('0')
            credit = Decimal(str(credit_value)) if credit_value else Decimal('0')

            # Get category_id
            category_id = data.get('category_id')

            # Get payment_method_id
            payment_method_id = data.get('payment_method_id')

            # Apply flexible percentage markup based on category, payment method, or both
            if category_id or payment_method_id:
                if credit > 0:
                    credit = apply_percentage_markup(credit, category_id, payment_method_id, session['user_id'], connection)
                elif debit > 0:
                    debit = apply_percentage_markup(debit, category_id, payment_method_id, session['user_id'], connection)

            print(f"[DEBUG] Debit: {debit}, Credit: {credit}")

            # Handle transaction_date - use current date if not provided or empty
            transaction_date = data.get('transaction_date')
            if not transaction_date or transaction_date == '':
                transaction_date = datetime.now().date()

            # Handle payment method and is_done logic
            # If payment_method_id is provided (and not empty), set is_done to TRUE
            # If payment_method_id is None/empty, set is_done to FALSE
            payment_method_id = data.get('payment_method_id')
            if payment_method_id:
                is_done = True
            else:
                is_done = False
                payment_method_id = None

            # Get payment status from frontend (explicit is_paid flag)
            is_paid = data.get('is_paid', False)  # Use explicit flag from frontend

            # Get paid_at timestamp if provided
            paid_at = None
            if is_paid:
                paid_at_str = data.get('paid_at')
                if paid_at_str:
                    # Parse ISO timestamp from frontend
                    try:
                        paid_at = datetime.fromisoformat(paid_at_str.replace('Z', '+00:00'))
                    except (ValueError, AttributeError):
                        # If parsing fails, use current timestamp
                        paid_at = datetime.now()
                else:
                    # No paid_at provided but marked as paid, use current timestamp
                    paid_at = datetime.now()

            # Update transaction (balance will be calculated on frontend)
            cursor.execute("""
                           UPDATE transactions
                           SET description      = %s,
                               category_id      = %s,
                               debit            = %s,
                               credit           = %s,
                               transaction_date = %s,
                               notes            = %s,
                               payment_method_id = %s,
                               is_done          = %s,
                               is_paid          = %s,
                               paid_at          = %s
                           WHERE id = %s
                           """, (
                data.get('description'),
                data.get('category_id'),
                debit if debit > 0 else None,
                credit if credit > 0 else None,
                transaction_date,
                data.get('notes'),
                payment_method_id,
                is_done,
                is_paid,
                paid_at,
                transaction_id
            ))

            print(f"[DEBUG] Transaction {transaction_id} updated successfully")

            # Log audit trail for each changed field
            user_id = session['user_id']

            # Track field changes
            new_debit = debit if debit > 0 else None
            new_credit = credit if credit > 0 else None

            # Normalize values for comparison to avoid false positives
            # Convert category_id to int for comparison (handle None)
            old_category = old_transaction['category_id']
            new_category = int(data.get('category_id')) if data.get('category_id') else None

            # Compare description
            if old_transaction['description'] != data.get('description'):
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'description',
                                      old_transaction['description'], data.get('description'))

            # Compare category_id (normalized)
            if old_category != new_category:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'category_id',
                                      old_category, new_category)

            # Compare debit (Decimal comparison)
            old_debit_normalized = old_transaction['debit'] if old_transaction['debit'] else None
            if old_debit_normalized != new_debit:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'debit',
                                      old_transaction['debit'], new_debit)

            # Compare credit (Decimal comparison)
            old_credit_normalized = old_transaction['credit'] if old_transaction['credit'] else None
            if old_credit_normalized != new_credit:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'credit',
                                      old_transaction['credit'], new_credit)

            # Compare transaction_date
            if str(old_transaction['transaction_date']) != str(transaction_date):
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'transaction_date',
                                      old_transaction['transaction_date'], transaction_date)

            # Compare notes (handle None/empty string)
            old_notes = old_transaction['notes'] if old_transaction['notes'] else None
            new_notes = data.get('notes') if data.get('notes') else None
            if old_notes != new_notes:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'notes',
                                      old_transaction['notes'], data.get('notes'))

            # Compare payment_method_id (handle None)
            old_payment_method = old_transaction['payment_method_id']
            new_payment_method = int(data.get('payment_method_id')) if data.get('payment_method_id') else None
            if old_payment_method != new_payment_method:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'payment_method_id',
                                      old_payment_method, new_payment_method)

            # Compare is_done status
            old_is_done = old_transaction.get('is_done', False)
            # Calculate new is_done based on payment method
            new_is_done = True if new_payment_method else False
            if old_is_done != new_is_done:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'is_done',
                                      old_is_done, new_is_done)

            # Compare is_paid status
            old_is_paid = old_transaction.get('is_paid', False)
            if old_is_paid != is_paid:
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'is_paid',
                                      old_is_paid, is_paid)

            # Compare paid_at timestamp
            old_paid_at = old_transaction.get('paid_at')
            if str(old_paid_at) != str(paid_at):
                log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'paid_at',
                                      old_paid_at, paid_at)

            connection.commit()
            print(f"[DEBUG] Transaction update committed successfully")
            return jsonify({'message': 'Transaction updated successfully'})

        else:  # DELETE
            # Log audit trail before deleting
            user_id = session['user_id']

            # Check if transaction has an attachment and delete it from Google Drive
            dict_cursor = connection.cursor(dictionary=True)
            dict_cursor.execute("""
                SELECT t.attachments
                FROM transactions t
                WHERE t.id = %s
                  AND t.monthly_record_id IN
                      (SELECT id FROM monthly_records WHERE user_id = %s)
            """, (transaction_id, user_id))

            transaction = dict_cursor.fetchone()
            dict_cursor.close()

            if transaction and transaction['attachments']:
                attachments_value = transaction['attachments']
                # Split comma-separated GUIDs (supports both single and multiple attachments)
                attachment_guids = [guid.strip() for guid in attachments_value.split(',') if guid.strip()]

                # Delete all attachments from storage
                if file_service.is_available():
                    for attachment_guid in attachment_guids:
                        success, error = file_service.delete_file(attachment_guid)
                        if success:
                            logger.info(
                                f"Deleted attachment {attachment_guid} from Google Drive for transaction {transaction_id}")
                        else:
                            logger.warning(f"Failed to delete attachment {attachment_guid}: {error}")
                            # Continue with other deletions even if one fails

            log_transaction_audit(cursor, transaction_id, user_id, 'DELETE')

            cursor.execute("""
                           DELETE
                           FROM transactions
                           WHERE id = %s
                             AND monthly_record_id IN
                                 (SELECT id FROM monthly_records WHERE user_id = %s)
                           """, (transaction_id, user_id))

            connection.commit()
            return jsonify({'message': 'Transaction deleted successfully'})

    except Error as e:
        print(f"[ERROR] Database error in manage_transaction: {str(e)}")
        connection.rollback()
        return jsonify({'error': f'Database error: {str(e)}'}), 500
    except Exception as e:
        print(f"[ERROR] Unexpected error in manage_transaction: {str(e)}")
        connection.rollback()
        return jsonify({'error': f'Server error: {str(e)}'}), 500
    finally:
        cursor.close()
        connection.close()


def get_transaction_audit_logs_handler(transaction_id):
    """Get audit logs for a specific transaction."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']

        # Verify the transaction belongs to the user
        cursor.execute("""
                       SELECT t.id
                       FROM transactions t
                                INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                       WHERE t.id = %s
                         AND mr.user_id = %s
                       """, (transaction_id, user_id))

        transaction = cursor.fetchone()
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404

        # Fetch audit logs for this transaction
        cursor.execute("""
                       SELECT tal.id,
                              tal.action,
                              tal.field_name,
                              tal.old_value,
                              tal.new_value,
                              tal.created_at,
                              u.username
                       FROM transaction_audit_logs tal
                                INNER JOIN users u ON tal.user_id = u.id
                       WHERE tal.transaction_id = %s
                       ORDER BY tal.created_at DESC
                       """, (transaction_id,))

        audit_logs = cursor.fetchall()
        return jsonify(audit_logs)

    except Error as e:
        logger.error(f"Error fetching audit logs: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def move_transaction_handler(transaction_id):
    """Move a transaction to a different month."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']
        data = request.get_json()

        target_year = data.get('target_year')
        target_month = data.get('target_month')

        if not target_year or not target_month:
            return jsonify({'error': 'Target year and month are required'}), 400

        # Verify the transaction belongs to the user
        cursor.execute("""
                       SELECT t.*, mr.year, mr.month
                       FROM transactions t
                                INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                       WHERE t.id = %s
                         AND mr.user_id = %s
                       """, (transaction_id, user_id))

        transaction = cursor.fetchone()
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404

        # Check if moving to the same month
        if transaction['year'] == target_year and transaction['month'] == target_month:
            return jsonify({'error': 'Transaction is already in this month'}), 400

        # Get or create target monthly record
        month_name = calendar.month_name[target_month]
        cursor.execute("""
                       INSERT INTO monthly_records (user_id, year, month, month_name)
                       VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                       UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                       """, (user_id, target_year, target_month, month_name))

        target_record_id = cursor.lastrowid

        # Update transaction's monthly_record_id and date
        new_date = datetime(target_year, target_month, 1).date()
        cursor.execute("""
                       UPDATE transactions
                       SET monthly_record_id = %s,
                           transaction_date  = %s
                       WHERE id = %s
                       """, (target_record_id, new_date, transaction_id))

        # Log audit trail
        log_transaction_audit(cursor, transaction_id, user_id, 'UPDATE', 'moved_to_month',
                              f"{transaction['year']}-{transaction['month']:02d}",
                              f"{target_year}-{target_month:02d}")

        connection.commit()

        return jsonify({
            'message': f'Transaction moved to {month_name} {target_year} successfully',
            'target_year': target_year,
            'target_month': target_month
        })

    except Error as e:
        logger.error(f"Error moving transaction: {e}")
        connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def copy_transaction_handler(transaction_id):
    """Copy a transaction to a different month."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']
        data = request.get_json()

        target_year = data.get('target_year')
        target_month = data.get('target_month')

        if not target_year or not target_month:
            return jsonify({'error': 'Target year and month are required'}), 400

        # Verify the transaction belongs to the user and get its data
        cursor.execute("""
                       SELECT t.*
                       FROM transactions t
                                INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
                       WHERE t.id = %s
                         AND mr.user_id = %s
                       """, (transaction_id, user_id))

        transaction = cursor.fetchone()
        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404

        # Get or create target monthly record
        month_name = calendar.month_name[target_month]
        cursor.execute("""
                       INSERT INTO monthly_records (user_id, year, month, month_name)
                       VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                       UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                       """, (user_id, target_year, target_month, month_name))

        target_record_id = cursor.lastrowid

        # Push all existing transactions down in the target month
        cursor.execute("""
                       UPDATE transactions
                       SET display_order = display_order + 1
                       WHERE monthly_record_id = %s
                       """, (target_record_id,))

        # Create a copy of the transaction in the target month at position 1 (top)
        new_date = datetime(target_year, target_month, 1).date()
        debit = Decimal(str(transaction['debit'])) if transaction['debit'] else None
        credit = Decimal(str(transaction['credit'])) if transaction['credit'] else None

        cursor.execute("""
                       INSERT INTO transactions
                       (monthly_record_id, description, category_id, debit, credit,
                        transaction_date, notes, payment_method_id, display_order)
                       VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                       """, (
            target_record_id,
            transaction['description'],
            transaction['category_id'],
            debit,
            credit,
            new_date,
            transaction['notes'],
            transaction['payment_method_id'],
            1  # Display at top
        ))

        new_transaction_id = cursor.lastrowid

        # Log audit trail for the new transaction
        log_transaction_audit(cursor, new_transaction_id, user_id, 'CREATE')
        log_transaction_audit(cursor, new_transaction_id, user_id, 'UPDATE', 'copied_from_transaction',
                              None, str(transaction_id))

        connection.commit()

        return jsonify({
            'message': f'Transaction copied to {month_name} {target_year} successfully',
            'new_transaction_id': new_transaction_id,
            'target_year': target_year,
            'target_month': target_month
        })

    except Error as e:
        logger.error(f"Error copying transaction: {e}")
        connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def manage_transaction_attachment_handler(transaction_id):
    """Get or delete a transaction's attachment."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']

        # Verify transaction belongs to user
        cursor.execute("""
            SELECT t.attachments, t.monthly_record_id
            FROM transactions t
            INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
            WHERE t.id = %s AND mr.user_id = %s
        """, (transaction_id, user_id))

        transaction = cursor.fetchone()

        if not transaction:
            return jsonify({'error': 'Transaction not found'}), 404

        attachments_value = transaction['attachments']

        if not attachments_value:
            return jsonify({'error': 'No attachment found for this transaction'}), 404

        # Split comma-separated GUIDs (supports both single and multiple attachments)
        attachment_guids = [guid.strip() for guid in attachments_value.split(',') if guid.strip()]

        if request.method == 'GET':
            # Return attachment info(s) via proxy URLs
            if file_service.is_available():
                attachments_list = []

                for attachment_guid in attachment_guids:
                    # Get file metadata using the service
                    file_name, mime_type = file_service.get_file_metadata(attachment_guid)

                    # Return proxy URLs that stream files (avoids loading into memory)
                    # Proxy handles authentication and streams response
                    proxy_url = url_for('serve_attachment', transaction_id=transaction_id,
                                        attachment_guid=attachment_guid, _external=True)

                    attachments_list.append({
                        'attachment_guid': attachment_guid,
                        'file_url': proxy_url,
                        'download_url': proxy_url + '?download=1',
                        'file_name': file_name,
                        'mime_type': mime_type
                    })

                    logger.info(f"Generated proxy URL for {attachment_guid}: {proxy_url}")

                # Return list of attachments (backward compatible: single item list for single attachment)
                return jsonify({
                    'attachments': attachments_list,
                    'count': len(attachments_list)
                }), 200
            else:
                return jsonify({'error': 'File storage not available'}), 500

        elif request.method == 'DELETE':
            # Delete attachment(s) from Google Drive and update transaction
            # If specific_guid is provided in request, delete only that one; otherwise delete all
            specific_guid = request.args.get('guid')

            if file_service.is_available():
                guids_to_delete = [specific_guid] if specific_guid else attachment_guids

                for guid in guids_to_delete:
                    success, error = file_service.delete_file(guid)
                    if success:
                        logger.info(f"Deleted attachment {guid} from Google Drive")
                    else:
                        logger.error(f"Failed to delete attachment {guid}: {error}")
                        # Continue with other deletions even if one fails

            # Update database: remove deleted GUID(s) from attachments field
            if specific_guid and specific_guid in attachment_guids:
                # Remove only the specific GUID
                remaining_guids = [g for g in attachment_guids if g != specific_guid]
                new_attachments_value = ','.join(remaining_guids) if remaining_guids else None
            else:
                # Delete all attachments
                new_attachments_value = None

            cursor.execute("""
                UPDATE transactions
                SET attachments = %s
                WHERE id = %s
            """, (new_attachments_value, transaction_id))

            connection.commit()

            logger.info(f"Updated attachments for transaction {transaction_id}")

            return jsonify({'message': 'Attachment(s) deleted successfully'}), 200

    except Error as e:
        logger.error(f"Error managing attachment: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def serve_attachment_handler(transaction_id):
    """Proxy endpoint to serve attachment files with authentication."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = None
    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']

        # Verify transaction belongs to user and get attachment GUID(s)
        cursor.execute("""
            SELECT t.attachments
            FROM transactions t
            INNER JOIN monthly_records mr ON t.monthly_record_id = mr.id
            WHERE t.id = %s AND mr.user_id = %s
        """, (transaction_id, user_id))

        transaction = cursor.fetchone()

        if not transaction:
            return "Transaction not found", 404

        attachments_value = transaction['attachments']

        if not attachments_value:
            return "No attachment found", 404

        # Get specific GUID from query parameter, or default to first one
        requested_guid = request.args.get('attachment_guid')
        attachment_guids = [guid.strip() for guid in attachments_value.split(',') if guid.strip()]

        if requested_guid:
            # Validate requested GUID is in the list
            if requested_guid not in attachment_guids:
                return "Attachment not found", 404
            attachment_guid = requested_guid
        else:
            # Default to first attachment
            attachment_guid = attachment_guids[0]

        if not file_service.is_available():
            return "Storage not available", 500

        # Get file metadata using the service
        file_name, mime_type = file_service.get_file_metadata(attachment_guid)
        logger.info(f"Fetching attachment {attachment_guid}, mime: {mime_type}")

        # Download the file content using the service
        file_content, status_code, error = file_service.download_file(attachment_guid)

        if not file_content:
            return error or "Failed to download file", status_code

        # Create response with the file content
        flask_response = Response(file_content, content_type=mime_type)
        flask_response.headers['Cache-Control'] = 'public, max-age=31536000'  # Cache for 1 year

        # Check if download parameter is present
        if request.args.get('download') == '1':
            flask_response.headers['Content-Disposition'] = f'attachment; filename="{file_name}"'
        else:
            flask_response.headers['Content-Disposition'] = 'inline'

        logger.info(f"Serving attachment {attachment_guid}: Content-Type={mime_type}, Mode={'download' if request.args.get('download') == '1' else 'inline'}, Size={len(file_content)} bytes")
        return flask_response


    except Error as e:
        logger.error(f"Error serving attachment: {str(e)}")
        return f"Database error: {str(e)}", 500
    finally:
        cursor.close()
        connection.close()


def export_transactions_handler():
    """Export transactions to CSV, PDF, or Excel format."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    try:
        cursor = connection.cursor(dictionary=True)
        user_id = session['user_id']

        # Get parameters
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)
        export_format = request.args.get('format', 'csv')

        # Get monthly record
        cursor.execute("""
                       SELECT id
                       FROM monthly_records
                       WHERE user_id = %s AND year = %s AND month = %s
                       """, (user_id, year, month))

        monthly_record = cursor.fetchone()

        if not monthly_record:
            # Return empty file if no transactions
            transactions = []
        else:
            # Fetch transactions (DESC order for downloads - oldest first)
            cursor.execute("""
                           SELECT t.id,
                                  t.transaction_date,
                                  t.description,
                                  c.name  as category,
                                  t.debit,
                                  t.credit,
                                  t.notes,
                                  pm.name as payment_method,
                                  t.is_done,
                                  t.is_paid,
                                  t.paid_at
                           FROM transactions t
                                    LEFT JOIN categories c ON t.category_id = c.id
                                    LEFT JOIN payment_methods pm ON t.payment_method_id = pm.id
                           WHERE t.monthly_record_id = %s
                           ORDER BY t.display_order DESC, t.id DESC
                           """, (monthly_record['id'],))

            transactions = cursor.fetchall()

        # Generate file based on format
        if export_format == 'csv':
            return generate_csv(transactions, year, month)
        elif export_format == 'excel':
            return generate_excel(transactions, year, month)
        elif export_format == 'pdf':
            return generate_pdf(transactions, year, month)
        else:
            return jsonify({'error': 'Invalid format'}), 400

    except Error as e:
        logger.error(f"Error exporting transactions: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def reorder_transactions_handler():
    """Reorder transactions based on new order array of transaction IDs."""
    data = request.get_json()
    transaction_ids = data.get('transaction_ids', [])

    if not transaction_ids or not isinstance(transaction_ids, list):
        return jsonify({'error': 'Invalid transaction_ids. Must be a non-empty array'}), 400

    connection = get_db_connection()
    if connection:
        cursor = connection.cursor(dictionary=True)
        try:
            user_id = session.get('user_id')

            # Build a single UPDATE … CASE statement to set all
            # display_order values in one round-trip instead of N queries.
            case_clauses = []
            params = []
            for index, transaction_id in enumerate(transaction_ids):
                case_clauses.append("WHEN %s THEN %s")
                params.extend([transaction_id, index + 1])

            # Append the IN-list params
            params.extend(transaction_ids)
            placeholders = ','.join(['%s'] * len(transaction_ids))

            cursor.execute(
                f"UPDATE transactions SET display_order = CASE id "
                f"{' '.join(case_clauses)} END "
                f"WHERE id IN ({placeholders})",
                params,
            )

            connection.commit()
            return jsonify({
                'success': True,
                'message': 'Transaction order updated successfully'
            })

        except Error as e:
            connection.rollback()
            return jsonify({'error': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    return jsonify({'error': 'Database connection failed'}), 500


def mark_transaction_done_handler(transaction_id):
    """Mark a transaction as done with payment method."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        data = request.get_json()
        payment_method_id = data.get('payment_method_id')

        cursor.execute("""
                       UPDATE transactions
                       SET is_done           = TRUE,
                           payment_method_id = %s,
                           marked_done_at    = CURRENT_TIMESTAMP
                       WHERE id = %s
                         AND monthly_record_id IN
                             (SELECT id FROM monthly_records WHERE user_id = %s)
                       """, (payment_method_id, transaction_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Transaction marked as done'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def mark_transaction_undone_handler(transaction_id):
    """Mark a transaction as not done."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        cursor.execute("""
                       UPDATE transactions
                       SET is_done           = FALSE,
                           payment_method_id = NULL,
                           marked_done_at    = NULL
                       WHERE id = %s
                         AND monthly_record_id IN
                             (SELECT id FROM monthly_records WHERE user_id = %s)
                       """, (transaction_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Transaction marked as not done'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def mark_transaction_paid_handler(transaction_id):
    """Mark a transaction as paid (when description cell is clicked)."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        data = request.get_json()
        payment_method_id = data.get('payment_method_id')

        cursor.execute("""
                       UPDATE transactions
                       SET is_done           = TRUE,
                           is_paid           = TRUE,
                           payment_method_id = %s,
                           marked_done_at    = CURRENT_TIMESTAMP,
                           paid_at           = CURRENT_TIMESTAMP
                       WHERE id = %s
                         AND monthly_record_id IN
                             (SELECT id FROM monthly_records WHERE user_id = %s)
                       """, (payment_method_id, transaction_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Transaction marked as paid'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def mark_transaction_unpaid_handler(transaction_id):
    """Unmark a transaction as paid (reverse the paid status)."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor()

    try:
        cursor.execute("""
                       UPDATE transactions
                       SET is_done           = FALSE,
                           is_paid           = FALSE,
                           payment_method_id = NULL,
                           marked_done_at    = NULL,
                           paid_at           = NULL
                       WHERE id = %s
                         AND monthly_record_id IN
                             (SELECT id FROM monthly_records WHERE user_id = %s)
                       """, (transaction_id, session['user_id']))

        connection.commit()
        return jsonify({'message': 'Transaction marked as unpaid'})

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def get_payment_method_totals_handler():
    """Get totals for each payment method for the current month."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        year = request.args.get('year', datetime.now().year, type=int)
        month = request.args.get('month', datetime.now().month, type=int)

        # Get monthly record
        cursor.execute("""
                       SELECT id
                       FROM monthly_records
                       WHERE user_id = %s AND year = %s AND month = %s
                       """, (user_id, year, month))

        monthly_record = cursor.fetchone()

        if not monthly_record:
            return jsonify([])

        # Get totals by payment method
        cursor.execute("""
                       SELECT pm.id,
                              pm.name,
                              pm.type,
                              pm.color,
                              COUNT(t.id)                                       as transaction_count,
                              SUM(t.debit)                                      as total_debit,
                              SUM(t.credit)                                     as total_credit,
                              SUM(COALESCE(t.debit, 0) - COALESCE(t.credit, 0)) as net_amount
                       FROM payment_methods pm
                                LEFT JOIN transactions t ON pm.id = t.payment_method_id
                           AND t.monthly_record_id = %s
                           AND t.is_done = TRUE
                       WHERE pm.user_id = %s
                         AND pm.is_active = TRUE
                       GROUP BY pm.id, pm.name, pm.type, pm.color
                       ORDER BY pm.type, pm.name
                       """, (monthly_record['id'], user_id))

        totals = cursor.fetchall()
        return jsonify(totals)

    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def clone_month_transactions_handler():
    """Clone all transactions from one month to another."""
    connection = get_db_connection()
    if not connection:
        return jsonify({'error': 'Database connection failed'}), 500

    cursor = connection.cursor(dictionary=True)
    user_id = session['user_id']

    try:
        data = request.get_json()
        from_year = data.get('from_year')
        from_month = data.get('from_month')
        to_year = data.get('to_year')
        to_month = data.get('to_month')
        include_payments = data.get('include_payments', False)

        # Validate inputs
        if not all([from_year, from_month, to_year, to_month]):
            return jsonify({'error': 'All date fields are required'}), 400

        if from_year == to_year and from_month == to_month:
            return jsonify({'error': 'Source and target months cannot be the same'}), 400

        # Get source monthly record
        cursor.execute("""
                       SELECT id
                       FROM monthly_records
                       WHERE user_id = %s AND year = %s AND month = %s
                       """, (user_id, from_year, from_month))

        source_record = cursor.fetchone()

        if not source_record:
            return jsonify({'error': 'Source month has no transactions'}), 404

        # Get or create target monthly record
        month_name = calendar.month_name[to_month]
        cursor.execute("""
                       INSERT INTO monthly_records (user_id, year, month, month_name)
                       VALUES (%s, %s, %s, %s) ON DUPLICATE KEY
                       UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
                       """, (user_id, to_year, to_month, month_name))

        target_record = {'id': cursor.lastrowid}

        # Get all transactions from source month (preserve order)
        cursor.execute("""
                       SELECT description,
                              category_id,
                              debit,
                              credit,
                              notes,
                              payment_method_id,
                              is_done,
                              is_paid,
                              display_order
                       FROM transactions
                       WHERE monthly_record_id = %s
                       ORDER BY display_order ASC, id ASC
                       """, (source_record['id'],))

        source_transactions = cursor.fetchall()

        if not source_transactions:
            return jsonify({'error': 'No transactions found in source month'}), 404

        # Clone transactions using a single multi-row INSERT instead
        # of one INSERT per transaction.
        clone_date = datetime.now().date()
        insert_values = []
        insert_params = []

        for trans in source_transactions:
            debit = Decimal(str(trans['debit'])) if trans['debit'] else Decimal('0')
            credit = Decimal(str(trans['credit'])) if trans['credit'] else Decimal('0')

            payment_method_id = trans['payment_method_id'] if include_payments else None
            is_done = trans['is_done'] if include_payments else False
            is_paid = trans['is_paid'] if include_payments else False

            insert_values.append("(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)")
            insert_params.extend([
                target_record['id'],
                trans['description'],
                trans['category_id'],
                debit if debit > 0 else None,
                credit if credit > 0 else None,
                clone_date,
                trans['notes'],
                payment_method_id,
                is_done,
                is_paid,
                trans['display_order'],
            ])

        cursor.execute(
            "INSERT INTO transactions "
            "(monthly_record_id, description, category_id, debit, credit, "
            "transaction_date, notes, payment_method_id, is_done, is_paid, display_order) VALUES "
            + ", ".join(insert_values),
            insert_params,
        )
        cloned_count = len(insert_values)

        connection.commit()

        return jsonify({
            'message': f'Successfully cloned {cloned_count} transactions',
            'count': cloned_count
        }), 200

    except Error as e:
        connection.rollback()
        logger.error(f"Error cloning transactions: {str(e)}")
        return jsonify({'error': str(e)}), 500
    finally:
        cursor.close()
        connection.close()


def scan_bill_handler():
    """Scan a bill image or PDF using Gemini AI to extract shop name, amount, discounts, and line items."""
    try:
        # Check for multiple images first, then fall back to single image (backward compatible)
        bill_images = request.files.getlist('bill_images')

        # Backward compatibility: check for single 'bill_image' if no 'bill_images'
        if not bill_images:
            if 'bill_image' in request.files:
                bill_images = [request.files['bill_image']]
            else:
                return jsonify({'error': 'No bill file provided'}), 400

        # Validate we have at least one file
        if not bill_images or all(img.filename == '' for img in bill_images):
            return jsonify({'error': 'No file selected'}), 400

        # Validate maximum number of images (5 images max for API performance)
        MAX_IMAGES = 5
        if len(bill_images) > MAX_IMAGES:
            return jsonify({'error': f'Maximum {MAX_IMAGES} images allowed per scan'}), 400

        # Process and validate each image
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}
        processed_images = []
        total_size = 0

        for idx, bill_image in enumerate(bill_images):
            # Check if file is empty
            if bill_image.filename == '':
                continue

            # Validate file type
            file_ext = bill_image.filename.rsplit('.', 1)[-1].lower() if '.' in bill_image.filename else ''

            if file_ext not in allowed_extensions:
                return jsonify({
                    'error': f'Invalid file type for image {idx + 1}: "{bill_image.filename}". Allowed: {", ".join(allowed_extensions)}'
                }), 400

            # Read file data
            image_data = bill_image.read()

            if len(image_data) == 0:
                return jsonify({'error': f'Empty file: "{bill_image.filename}"'}), 400

            total_size += len(image_data)

            # Check combined file size (16MB Flask limit)
            if total_size > 16 * 1024 * 1024:
                return jsonify({'error': 'Combined file size exceeds 16MB limit'}), 413

            # Fix image orientation if it's an image file (not PDF)
            if file_ext in ['jpg', 'jpeg', 'png', 'gif', 'webp'] and PIL_AVAILABLE:
                try:
                    img = Image.open(io.BytesIO(image_data))
                    img = fix_image_orientation(img)

                    # Convert back to bytes
                    output = io.BytesIO()
                    if file_ext in ['jpg', 'jpeg']:
                        img.save(output, format='JPEG', quality=95)
                    elif file_ext == 'png':
                        img.save(output, format='PNG')
                    elif file_ext == 'webp':
                        img.save(output, format='WEBP', quality=95)
                    else:
                        img.save(output, format=img.format or 'JPEG', quality=95)

                    image_data = output.getvalue()
                    logger.info(f"Image {idx + 1} orientation corrected for scanning")
                except Exception as e:
                    logger.warning(f"Failed to fix orientation for image {idx + 1}, using original: {str(e)}")

            processed_images.append(image_data)

        # Validate we have at least one valid image
        if not processed_images:
            return jsonify({'error': 'No valid images to process'}), 400

        # Get Gemini scanner instance
        scanner = get_gemini_bill_scanner()

        if not scanner:
            return jsonify({
                'error': 'Bill scanning service not configured. Please set Gemini API key in integration settings.'
            }), 503

        # Scan the bill with Gemini AI
        image_input = processed_images[0] if len(processed_images) == 1 else processed_images
        logger.info(f"Scanning {len(processed_images)} bill image(s) for user {session['user_id']}")
        result = scanner.scan_bill(image_input)

        # Check if there was an error
        if 'error' in result:
            logger.error(f"Bill scanning error: {result['error']}")
            return jsonify({
                'success': False,
                'shop_name': result.get('shop_name', 'Unknown Store'),
                'amount': result.get('amount', '0'),
                'subtotal': result.get('subtotal', '0'),
                'discounts': result.get('discounts', []),
                'items': result.get('items', []),
                'error': result['error'],
                'raw_response': result.get('raw_response', '')
            }), 200

        # Return successful result
        logger.info(
            f"Bill scanned successfully ({len(processed_images)} image(s)): {result['shop_name']} - {result['amount']} - {len(result.get('items', []))} items")

        return jsonify({
            'success': True,
            'shop_name': result['shop_name'],
            'amount': result['amount'],
            'subtotal': result.get('subtotal', '0'),
            'discounts': result.get('discounts', []),
            'items': result.get('items', []),
            'raw_response': result.get('raw_response', '')
        }), 200

    except Exception as e:
        logger.error(f"Error scanning bill: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'Failed to scan bill: {str(e)}',
            'shop_name': 'Unknown Store',
            'amount': '0',
            'subtotal': '0',
            'discounts': [],
            'items': []
        }), 500


def upload_bill_attachment_handler():
    """Upload a single bill image/PDF to Google Drive and return its file ID."""
    try:
        if 'bill_image' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400

        bill_image = request.files['bill_image']

        if not bill_image or not bill_image.filename:
            return jsonify({'success': False, 'error': 'Empty file'}), 400

        # Validate file type
        allowed_extensions = {'png', 'jpg', 'jpeg', 'gif', 'webp', 'pdf'}
        file_ext = bill_image.filename.rsplit('.', 1)[-1].lower() if '.' in bill_image.filename else ''

        if file_ext not in allowed_extensions:
            return jsonify({'success': False, 'error': f'Invalid file type: {file_ext}'}), 400

        # Read file data
        file_data = bill_image.read()

        if len(file_data) == 0:
            return jsonify({'success': False, 'error': 'Empty file data'}), 400

        if not file_service.is_available():
            return jsonify({'success': False, 'error': 'File storage not configured'}), 503

        # Optimize file if needed
        optimized_data, was_optimized, original_size, final_size = optimize_file_for_upload(
            file_data,
            file_ext,
            bill_image.filename
        )

        # Log optimization
        if was_optimized:
            reduction_pct = ((original_size - final_size) / original_size) * 100
            logger.info(f"✓ Image optimized: {original_size / (1024 * 1024):.2f}MB → {final_size / (1024 * 1024):.2f}MB ({reduction_pct:.1f}% smaller)")

        # Generate GUID for filename
        attachment_guid = str(uuid.uuid4())
        filename = f"{attachment_guid}.{file_ext}"

        logger.info(f"Uploading to Google Drive: {filename}, size: {final_size / (1024 * 1024):.2f}MB")

        # Upload to Google Drive using the file service
        success, error, result = file_service.upload_file(
            optimized_data,
            attachment_guid,
            filename
        )

        if not success:
            raise Exception(error or "Upload failed")

        # Use Google Drive file ID as the attachment identifier
        drive_file_id = result.get('id', attachment_guid) if result else attachment_guid
        logger.info(f"✓ Uploaded successfully: {drive_file_id}")

        return jsonify({
            'success': True,
            'attachment_guid': drive_file_id
        }), 200

    except Exception as e:
        logger.error(f"Error uploading bill attachment: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': f'Upload failed: {str(e)}'
        }), 500


def create_transaction_handler():
    """Create a new transaction with token authentication."""
    try:
        # Get user ID from the token (set by token_required decorator)
        user_id = request.current_user['user_id']

        # Get request data
        data = request.get_json()

        if not data:
            return jsonify({'error': 'Request body is required'}), 400

        # Validate required fields
        description = data.get('description')
        credit_value = data.get('credit')

        if not description:
            return jsonify({'error': 'Description is required'}), 400

        if credit_value is None or credit_value == '':
            return jsonify({'error': 'Credit (expense) amount is required'}), 400

        # Use current date as transaction date
        transaction_date = datetime.now().date()

        # Convert credit to Decimal
        try:
            credit = Decimal(str(credit_value))
            if credit <= 0:
                return jsonify({'error': 'Credit amount must be greater than 0'}), 400
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid credit amount'}), 400

        # Auto-categorize from description, or use explicitly provided category_id
        category_id = data.get('category_id')
        category_name = None

        if category_id is not None:
            # Caller explicitly provided a category - validate it
            try:
                category_id = int(category_id)
            except (ValueError, TypeError):
                return jsonify({'error': 'Invalid category_id'}), 400
        else:
            # Auto-detect category from description
            category_id = auto_categorize_transaction(description)

        # Get database connection
        connection = get_db_connection()
        if not connection:
            return jsonify({'error': 'Database connection failed'}), 500

        cursor = connection.cursor(dictionary=True)

        try:
            # If we have a category_id (explicit or auto-detected), verify it exists
            if category_id is not None:
                cursor.execute("SELECT id, name FROM categories WHERE id = %s", (category_id,))
                category_row = cursor.fetchone()
                if category_row:
                    category_name = category_row['name']
                else:
                    # Invalid category_id - clear it rather than failing
                    category_id = None

            # Extract year and month from transaction date
            year = transaction_date.year
            month = transaction_date.month
            month_name = calendar.month_name[month]

            # Create or get monthly record for the transaction date
            cursor.execute("""
                INSERT INTO monthly_records (user_id, year, month, month_name)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE id = LAST_INSERT_ID(id), updated_at = CURRENT_TIMESTAMP
            """, (user_id, year, month, month_name))

            monthly_record = {'id': cursor.lastrowid}

            if not monthly_record['id']:
                return jsonify({'error': 'Failed to create or retrieve monthly record'}), 500

            # Push all existing transactions down by incrementing their display_order
            cursor.execute("""
                UPDATE transactions
                SET display_order = display_order + 1
                WHERE monthly_record_id = %s
            """, (monthly_record['id'],))

            # New transaction gets display_order = 1 (appears at top)
            next_display_order = 1

            # Apply flexible percentage markup (no payment method for mobile quick add)
            if category_id:
                credit = apply_percentage_markup(credit, category_id, None, user_id, connection)

            # Insert the transaction with auto-categorized category
            cursor.execute("""
                INSERT INTO transactions
                (monthly_record_id, description, category_id, debit, credit, transaction_date, display_order)
                VALUES (%s, %s, %s, NULL, %s, %s, %s)
            """, (monthly_record['id'], description, category_id, credit, transaction_date, next_display_order))

            transaction_id = cursor.lastrowid

            # Log the transaction creation in audit logs
            category_info = f", Category: {category_name} (auto)" if category_name and not data.get('category_id') else \
                f", Category: {category_name}" if category_name else ""
            log_transaction_audit(
                cursor,
                transaction_id,
                user_id,
                'CREATE',
                None,
                None,
                f"Created via API - Description: {description}, Credit: {credit}, Date: {transaction_date}{category_info}"
            )

            connection.commit()

            logger.info(f"Transaction created via API - User: {user_id}, ID: {transaction_id}, "
                        f"Description: {description}, Category: {category_name or 'Uncategorized'}")

            response = {
                'message': 'Transaction created successfully',
                'transaction_id': transaction_id,
                'description': description,
                'credit': float(credit),
                'transaction_date': transaction_date.isoformat(),
                'year': year,
                'month': month,
                'category_id': category_id,
                'category_name': category_name,
                'auto_categorized': category_id is not None and data.get('category_id') is None
            }

            return jsonify(response), 201

        except Error as e:
            connection.rollback()
            logger.error(f"Database error creating transaction: {str(e)}")
            return jsonify({'error': 'Failed to create transaction', 'details': str(e)}), 500
        finally:
            cursor.close()
            connection.close()

    except Exception as e:
        logger.error(f"Error creating transaction: {str(e)}")
        return jsonify({'error': 'Failed to create transaction', 'details': str(e)}), 500


# ==================================================
# ROUTE REGISTRATION
# ==================================================

def register_transaction_routes(app, login_required, limiter, rate_limit_api, token_required=None):
    """
    Register all transaction routes with the Flask app.

    Args:
        app: Flask application instance
        login_required: Login required decorator
        limiter: Flask-Limiter instance
        rate_limit_api: Rate limit string for API endpoints
        token_required: Token required decorator for API token auth
    """

    @app.route('/api/transactions', methods=['GET', 'POST'])
    @login_required
    @limiter.limit(rate_limit_api)
    def transactions():
        return transactions_handler()

    @app.route('/api/transactions/filter', methods=['GET'])
    @login_required
    def filter_transactions():
        return filter_transactions_handler()

    @app.route('/api/transactions/<int:transaction_id>', methods=['PUT', 'DELETE'])
    @login_required
    def manage_transaction(transaction_id):
        return manage_transaction_handler(transaction_id)

    @app.route('/api/transactions/<int:transaction_id>/audit-logs', methods=['GET'])
    @login_required
    def get_transaction_audit_logs(transaction_id):
        return get_transaction_audit_logs_handler(transaction_id)

    @app.route('/api/transactions/<int:transaction_id>/move', methods=['POST'])
    @login_required
    def move_transaction(transaction_id):
        return move_transaction_handler(transaction_id)

    @app.route('/api/transactions/<int:transaction_id>/copy', methods=['POST'])
    @login_required
    def copy_transaction(transaction_id):
        return copy_transaction_handler(transaction_id)

    @app.route('/api/transactions/<int:transaction_id>/attachment', methods=['GET', 'DELETE'])
    @login_required
    def manage_transaction_attachment(transaction_id):
        return manage_transaction_attachment_handler(transaction_id)

    @app.route('/api/transactions/<int:transaction_id>/attachment/view')
    @login_required
    def serve_attachment(transaction_id):
        return serve_attachment_handler(transaction_id)

    @app.route('/api/transactions/export', methods=['GET'])
    @login_required
    def export_transactions():
        return export_transactions_handler()

    @app.route('/api/transactions/reorder', methods=['POST'])
    @login_required
    def reorder_transactions():
        return reorder_transactions_handler()

    @app.route('/api/transactions/<int:transaction_id>/mark-done', methods=['POST'])
    @login_required
    def mark_transaction_done(transaction_id):
        return mark_transaction_done_handler(transaction_id)

    @app.route('/api/transactions/<int:transaction_id>/mark-undone', methods=['POST'])
    @login_required
    def mark_transaction_undone(transaction_id):
        return mark_transaction_undone_handler(transaction_id)

    @app.route('/api/transactions/<int:transaction_id>/mark-paid', methods=['POST'])
    @login_required
    def mark_transaction_paid(transaction_id):
        return mark_transaction_paid_handler(transaction_id)

    @app.route('/api/transactions/<int:transaction_id>/mark-unpaid', methods=['POST'])
    @login_required
    def mark_transaction_unpaid(transaction_id):
        return mark_transaction_unpaid_handler(transaction_id)

    @app.route('/api/payment-method-totals', methods=['GET'])
    @login_required
    def get_payment_method_totals():
        return get_payment_method_totals_handler()

    @app.route('/api/clone-month-transactions', methods=['POST'])
    @login_required
    def clone_month_transactions():
        return clone_month_transactions_handler()

    @app.route('/api/scan-bill', methods=['POST'])
    @login_required
    def scan_bill():
        return scan_bill_handler()

    @app.route('/api/upload-bill-attachment', methods=['POST'])
    @login_required
    def upload_bill_attachment():
        return upload_bill_attachment_handler()

    if token_required:
        @app.route('/api/transactions/create', methods=['POST'])
        @token_required
        def create_transaction():
            return create_transaction_handler()

    logger.info("Transaction routes registered successfully")
